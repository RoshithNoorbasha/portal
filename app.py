# app.py
import io
import re
import json
import hashlib
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from datetime import datetime
from pathlib import Path
import restore  # Restore & TAT module
import storage1  # Shared backend storage (uploads, users, audit log)
import analysis  # Multi-day SCADA string fault analysis page (admin/manager/super-admin only)
from functools import lru_cache

# ==========================================
# 1. PAGE CONFIGURATION & STYLING
# ==========================================
st.set_page_config(
    page_title="PV String Analytics",
    page_icon="☀️",  # Streamlit page-tab icons only accept an emoji/image, not Font Awesome markup.
    layout="wide",
    initial_sidebar_state="expanded"
)

# Font Awesome CDN
st.markdown("""
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
    .main .block-container { padding-top: 1.2rem; padding-bottom: 2rem; }

    /* ---------- Global card / metric polish ---------- */
    .stMetric {
        background-color: #0f172a;
        border: 1px solid #1e293b;
        padding: 1rem;
        border-radius: 0.75rem;
    }
    div[data-testid="stMetricValue"] {
        font-size: 1.7rem;
        font-weight: 700;
        color: #38bdf8;
    }
    .stDataFrame thead th {
        background-color: #1e293b;
        color: white;
        font-weight: 600;
    }

    /* ---------- Role badges ---------- */
    .user-badge-admin {
        background: linear-gradient(135deg, #8b5cf6, #6d28d9);
        color: white; padding: 3px 10px;
        border-radius: 999px; font-size: 0.72rem; font-weight: 700;
        letter-spacing: 0.02em;
    }
    .user-badge-manager {
        background: linear-gradient(135deg, #f59e0b, #d97706);
        color: white; padding: 3px 10px;
        border-radius: 999px; font-size: 0.72rem; font-weight: 700;
        letter-spacing: 0.02em;
    }
    .user-badge-engineer {
        background: linear-gradient(135deg, #3b82f6, #2563eb);
        color: white; padding: 3px 10px;
        border-radius: 999px; font-size: 0.72rem; font-weight: 700;
        letter-spacing: 0.02em;
    }
    .fa-icon { margin-right: 8px; }

    /* ---------- Header: welcome banner ---------- */
    .app-header {
        display: flex; align-items: center; justify-content: space-between;
        flex-wrap: wrap; gap: 12px;
        background: linear-gradient(120deg, #0b1220 0%, #101a30 55%, #0f2942 100%);
        border: 1px solid rgba(56, 189, 248, 0.22);
        border-radius: 20px;
        padding: 20px 26px;
        margin-bottom: 16px;
        box-shadow: 0 20px 45px rgba(2, 6, 23, 0.35), inset 0 1px 0 rgba(255,255,255,0.03);
        position: relative;
        overflow: hidden;
    }
    .app-header::before {
        content: "";
        position: absolute; top: -60%; right: -10%;
        width: 260px; height: 260px; border-radius: 50%;
        background: radial-gradient(circle, rgba(56,189,248,0.20), transparent 70%);
        pointer-events: none;
    }
    .app-header-left { display: flex; align-items: center; gap: 14px; z-index: 1; }
    .app-header-avatar {
        width: 46px; height: 46px; border-radius: 14px;
        background: linear-gradient(135deg, #38bdf8, #0ea5e9);
        display: flex; align-items: center; justify-content: center;
        font-size: 1.3rem; color: #04121f; font-weight: 800;
        box-shadow: 0 8px 18px rgba(56,189,248,0.35);
    }
    .app-header-greeting { font-size: 1.05rem; color: #f8fafc; font-weight: 600; margin: 0; }
    .app-header-sub { font-size: 0.78rem; color: #94a3b8; margin-top: 2px; }
    .app-header-right { display: flex; align-items: center; gap: 10px; z-index: 1; }

    /* ---------- Header: snapshot / calendar banner ---------- */
    .calendar-banner {
        background: linear-gradient(120deg, rgba(15, 23, 42, 0.97) 0%, rgba(15, 40, 64, 0.9) 100%);
        border: 1px solid rgba(56, 189, 248, 0.20);
        border-radius: 16px;
        padding: 14px 18px; margin-bottom: 14px;
        box-shadow: 0 12px 28px rgba(2, 6, 23, 0.2);
        display: flex; align-items: center; gap: 10px;
    }
    .calendar-banner strong { color: #f8fafc; font-size: 0.95rem; }
    .calendar-banner .calendar-subtitle { color: #94a3b8; font-size: 0.8rem; margin-top: 2px; }
    .calendar-icon-wrap {
        width: 34px; height: 34px; border-radius: 10px;
        background: rgba(56,189,248,0.14); color: #38bdf8;
        display: flex; align-items: center; justify-content: center; font-size: 1rem;
        flex-shrink: 0;
    }

    .negative-report-card {
        background: rgba(15, 23, 42, 0.72);
        border: 1px solid rgba(248, 113, 113, 0.22);
        border-radius: 16px;
        padding: 10px 14px;
    }

    /* ---------- Processing status card ---------- */
    div[data-testid="stStatusWidget"] {
        border-radius: 14px !important;
        border: 1px solid rgba(56,189,248,0.18) !important;
    }

    /* ---------- Filter form panel (PV String Details) ---------- */
    div[data-testid="stForm"] {
        background: rgba(15, 23, 42, 0.55);
        border: 1px solid rgba(148, 163, 184, 0.15);
        border-radius: 16px;
        padding: 16px 18px 6px 18px;
    }

    /* ---------- Login page: mobile-friendly card ---------- */
    .login-card {
        background: linear-gradient(160deg, rgba(15,23,42,0.9) 0%, rgba(15,40,64,0.85) 100%);
        border: 1px solid rgba(56, 189, 248, 0.22);
        border-radius: 20px;
        padding: 28px 24px 20px 24px;
        max-width: 420px;
        margin: 12px auto 0 auto;
        box-shadow: 0 20px 45px rgba(2, 6, 23, 0.35);
    }
    .login-title {
        text-align: center; color: #f8fafc; font-size: 1.4rem;
        font-weight: 700; margin-bottom: 2px;
    }
    .login-subtitle {
        text-align: center; color: #94a3b8; font-size: 0.85rem; margin-bottom: 18px;
    }
    /* Larger, easier-to-tap inputs/buttons on phones */
    @media (max-width: 640px) {
        .login-card { padding: 20px 16px 14px 16px; border-radius: 16px; margin-top: 4px; }
        .login-title { font-size: 1.2rem; }
        div[data-testid="stTextInput"] input {
            font-size: 1rem !important; padding-top: 0.6rem !important; padding-bottom: 0.6rem !important;
        }
        .stButton button, .stFormSubmitButton button {
            font-size: 1rem !important; padding-top: 0.6rem !important; padding-bottom: 0.6rem !important;
        }
        .main .block-container { padding-left: 0.6rem; padding-right: 0.6rem; }
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# 2. CONFIGURATION
# ==========================================
DEFAULT_TOTAL_ACTIVE_STRINGS = 19
WORKING_CURRENT_THRESHOLD = 0.5
# "Low performance" (yellow) = a working string sitting 20% or more below
# its own inverter's average working current (previously 23%, 30%, and
# 20% before that).
LOW_PERFORMANCE_DROP_PCT = 0.20
# A string only counts as a "negative value" fault (blinking red) at -1A or
# below - values between -1A and 0.5A are still just "Failed" (red), not
# flagged as a sensor/wiring-fault negative reading.
NEGATIVE_VALUE_THRESHOLD = -1.0
PV_CURRENT_COLUMNS = [f"PV-I{i}" for i in range(1, 29)]

ACTIVE_STRING_OVERRIDES = {
    "P2": {"IB1": 18, "IB3": 17, "IB4": 18, "IB5": 18},
    "P6": {"IB1": 18, "IB2": 18, "IB3": 18, "IB5": 18, "IB6": 18, "IB7": 18},
}

INVERTER_ID_COLS = [
    "Inverter ID", "Inverter_ID", "Inverter", "ID",
    "Device Name", "String Inverter", "Inverters"
]

# Duplicate Inverter IDs beyond this count in a single sheet block the
# upload entirely (data quality issue too big to silently work around).
# At or below this count, the upload still proceeds - duplicates are
# de-duplicated (first occurrence kept) and reported as a warning so the
# user can see exactly which Inverter IDs repeated.
MAX_ALLOWED_DUPLICATE_INVERTERS = 6


class InverterIDValidationError(Exception):
    """Raised when a workbook's Inverter ID column is missing, contains
    values that don't look like Inverter IDs, or has too many duplicate
    Inverter IDs to safely auto-resolve."""
    pass

MANUAL_SCADA_COLUMNS = [
    "String Inverter", "MBUS", "Grid", "E-Daily(KWH)", "Active Power", "Reactive Power",
    "PV1", "PV2", "PV3", "PV4", "PV5", "PV6", "PV7", "PV8", "PV9", "PV10",
    "PV11", "PV12", "PV13", "PV14", "PV15", "PV16", "PV17", "PV18", "PV19", "PV20",
    "PV21", "PV22", "PV23", "PV24", "PV25", "PV26", "PV27", "PV28",
    "PV-I1", "PV-I2", "PV-I3", "PV-I4", "PV-I5", "PV-I6", "PV-I7", "PV-I8", "PV-I9", "PV-I10",
    "PV-I11", "PV-I12", "PV-I13", "PV-I14", "PV-I15", "PV-I16", "PV-I17", "PV-I18", "PV-I19", "PV-I20",
    "PV-I21", "PV-I22", "PV-I23", "PV-I24", "PV-I25", "PV-I26", "PV-I27", "PV-I28",
    "VAB", "VBC", "VCA", "IA", "IB", "IC"
]

ROLE_BADGES = {
    "admin": "👑 Admin",
    "manager": "🧭 Manager",
    "engineer": "🔧 AM/Engineer",
}

# ==========================================
# 3. USER / SESSION HELPERS (delegates to storage.py)
# ==========================================
def get_current_user():
    return st.session_state.get("user")

def is_admin():
    user = get_current_user()
    return bool(user) and user.get("role") == "admin"

def is_manager():
    user = get_current_user()
    return bool(user) and user.get("role") == "manager"

def is_admin_or_manager():
    user = get_current_user()
    return bool(user) and user.get("role") in ("admin", "manager")

def is_engineer():
    user = get_current_user()
    return bool(user) and user.get("role") == "engineer"

def is_super_admin():
    user = get_current_user()
    return bool(user) and storage1.is_super_admin(user.get("username"))

# ==========================================
# 4. OPTIMIZED HELPERS WITH CACHING
# ==========================================
def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip().upper()

def clean_manual_columns(col_list):
    cleaned = []
    for col in col_list:
        col = str(col).strip()
        if col and col.lower() != "nan":
            cleaned.append(col)
    return cleaned

@lru_cache(maxsize=128)
def extract_plot_cached(inverter_id_str):
    if isinstance(inverter_id_str, str):
        parts = inverter_id_str.split("-")
        if len(parts) > 0:
            return parts[0].strip()
    return "Unknown Plot"

@lru_cache(maxsize=128)
def extract_block_cached(inverter_id_str):
    if isinstance(inverter_id_str, str):
        parts = inverter_id_str.split("-")
        if len(parts) > 1:
            return parts[1].strip()
    return "Unknown Block"

# @lru_cache(maxsize=128)
# def map_inverter_to_sacu_cached(inverter_id_str):
#     if not isinstance(inverter_id_str, str):
#         return "Invalid Inverter ID"

#     match = re.search(r'-(\d[\.\-]\d)-', inverter_id_str)
#     if match:
#         sacu_identifier = match.group(1)
#         try:
#             if "." in sacu_identifier:
#                 first_digit_str = sacu_identifier.split(".")[0]
#             else:
#                 first_digit_str = sacu_identifier.split("-")[0]
#             first_digit = int(first_digit_str)
#             if first_digit in [1, 2]:
#                 return "SACU-1"
#             elif first_digit in [3, 4]:
#                 return "SACU-2"
#         except ValueError:
#             pass
#     return "Unknown SACU"
@lru_cache(maxsize=128)
def map_inverter_to_sacu_cached(inverter_id_str):
    if not isinstance(inverter_id_str, str):
        return "Invalid Inverter ID"

    # Standardize to uppercase for safe checking
    upper_id = inverter_id_str.upper()

    # Check for LT designations first
    if "LT1" in upper_id or "LT2" in upper_id:
        return "SACU-1"
    elif "LT3" in upper_id or "LT4" in upper_id:
        return "SACU-2"

    # Fallback to the original regex pattern for numeric block identifiers
    match = re.search(r'-(\d[\.\-]\d)-', inverter_id_str)
    if match:
        sacu_identifier = match.group(1)
        try:
            if "." in sacu_identifier:
                first_digit_str = sacu_identifier.split(".")[0]
            else:
                first_digit_str = sacu_identifier.split("-")[0]
            first_digit = int(first_digit_str)
            if first_digit in [1, 2]:
                return "SACU-1"
            elif first_digit in [3, 4]:
                return "SACU-2"
        except ValueError:
            pass
            
    return "Unknown SACU"

def sorted_filter_options(series):
    """Return non-null unique values sorted safely for Streamlit filters."""
    values = [value for value in series.dropna().unique()]

    def natural_key(value):
        text = str(value).strip()
        return tuple(
            (0, int(part)) if part.isdigit() else (1, part.casefold())
            for part in re.split(r"(\d+)", text)
        )

    return sorted(values, key=natural_key)

def get_total_active_strings(plot, block):
    plot_key = normalize_text(plot)
    block_key = normalize_text(block)
    if plot_key in ACTIVE_STRING_OVERRIDES and block_key in ACTIVE_STRING_OVERRIDES[plot_key]:
        return ACTIVE_STRING_OVERRIDES[plot_key][block_key]
    return DEFAULT_TOTAL_ACTIVE_STRINGS

@st.cache_data(ttl=3600)
def get_available_pv_columns_cached(df):
    """Cache PV columns detection"""
    normalized_map = {str(col).strip().upper(): col for col in df.columns}
    available_columns = []
    for col in PV_CURRENT_COLUMNS:
        if col.upper() in normalized_map:
            available_columns.append(normalized_map[col.upper()])
    return available_columns

@st.cache_data(ttl=3600)
def apply_string_metrics_cached(df):
    """Cache the string metrics calculation"""
    df_copy = df.copy()
    pv_columns = get_available_pv_columns_cached(df_copy)

    df_copy["Total Active Strings"] = df_copy.apply(
        lambda row: get_total_active_strings(row.get("Plot"), row.get("Block")), axis=1
    )

    if pv_columns:
        df_copy["Working String Count"] = df_copy.apply(
            lambda row: calculate_working_string_count(row, pv_columns), axis=1
        )
    else:
        df_copy["Working String Count"] = 0

    df_copy["Failed String Count"] = (df_copy["Total Active Strings"] - df_copy["Working String Count"]).clip(lower=0)
    df_copy["Availability (%)"] = ((df_copy["Working String Count"] / df_copy["Total Active Strings"]) * 100).fillna(0).round(2)
    df_copy["Failure Percentage (%)"] = ((df_copy["Failed String Count"] / df_copy["Total Active Strings"]) * 100).fillna(0).round(2)
    return df_copy

def calculate_working_string_count(row, pv_columns):
    count = 0
    for col in pv_columns:
        value = pd.to_numeric(row.get(col), errors="coerce")
        if pd.notna(value) and value > WORKING_CURRENT_THRESHOLD:
            count += 1
    return count

@st.cache_data(ttl=3600)
def get_pv_string_columns_cached(df):
    """Cache PV string columns detection"""
    pv_voltage_cols, pv_current_cols = [], []
    for col in df.columns:
        col_str = str(col).strip()
        if col_str.startswith("PV-I"):
            pv_current_cols.append(col)
        elif col_str.startswith("PV") and col_str != "PV" and not col_str.startswith("PV-I"):
            try:
                num = int(col_str[2:])
                if 1 <= num <= 28:
                    pv_voltage_cols.append(col)
            except Exception:
                pass
    return pv_voltage_cols, pv_current_cols

def get_string_health_color(value):
    if pd.isna(value):
        return "#64748b"
    if value > 5.0:
        return "#10b981"
    elif value > 3.0:
        return "#34d399"
    elif value > 1.5:
        return "#fbbf24"
    elif value > 0.5:
        return "#f59e0b"
    else:
        return "#ef4444"

def get_column_header_color(value):
    if pd.isna(value):
        return "#64748b"
    if value >= 80:
        return "#10b981"
    elif value >= 60:
        return "#f59e0b"
    elif value >= 40:
        return "#f97316"
    else:
        return "#ef4444"

# ==========================================
# 5. OPTIMIZED PARSER WITH CACHING
# ==========================================
def _find_inverter_column(df):
    """Locate the Inverter ID column by known header names (case-insensitive)."""
    df_columns_lower_map = {str(c).strip().lower(): c for c in df.columns}
    for col in INVERTER_ID_COLS:
        if col in df.columns:
            return col
        elif col.strip().lower() in df_columns_lower_map:
            return df_columns_lower_map[col.strip().lower()]
    return None


def _validate_inverter_id_values(df, col, sheet_name):
    """
    Guard against the case where a column matching a known Inverter ID
    header name was found, but the *values* in it clearly aren't Inverter
    IDs (e.g. the workbook's columns got shifted, or the wrong column was
    labeled 'ID'). A real Inverter ID looks like 'P1-IB1-...' (Plot-Block-...),
    so anything that mostly lacks that shape is treated as a wrong-column
    upload and rejected with a clear exception instead of silently
    producing a dashboard full of "Unknown Plot"/"Unknown Block" rows.
    """
    raw_values = df[col].dropna().astype(str).str.strip()
    raw_values = raw_values[~raw_values.str.lower().isin(["", "nan", "none"])]

    if raw_values.empty:
        raise InverterIDValidationError(
            f"Sheet '{sheet_name}': the detected Inverter ID column ('{col}') is empty. "
            f"Please check that the correct column is labeled as the Inverter ID "
            f"(expected one of: {', '.join(INVERTER_ID_COLS)}) and re-upload."
        )

    looks_like_id = raw_values.str.contains("-", regex=False)
    invalid_ratio = 1 - looks_like_id.mean()

    if invalid_ratio > 0.5:
        sample_bad_values = raw_values[~looks_like_id].unique()[:5].tolist()
        raise InverterIDValidationError(
            f"Sheet '{sheet_name}': column '{col}' was matched as the Inverter ID column, "
            f"but its values don't look like Inverter IDs (expected a format like 'P1-IB1-...'). "
            f"Found values such as: {sample_bad_values}. Please check the column headers/values "
            f"and re-upload."
        )


def _check_duplicate_inverter_ids(df, col, sheet_name):
    """
    Identify duplicate Inverter IDs in a sheet. Returns the sorted list of
    duplicated IDs (empty if none). Raises InverterIDValidationError if the
    number of distinct duplicated IDs exceeds MAX_ALLOWED_DUPLICATE_INVERTERS.
    """
    id_series = df[col].astype(str).str.strip()
    id_series = id_series[~id_series.str.lower().isin(["", "nan", "none"])]

    counts = id_series.value_counts()
    duplicated_ids = sorted(counts[counts > 1].index.tolist())

    if len(duplicated_ids) > MAX_ALLOWED_DUPLICATE_INVERTERS:
        raise InverterIDValidationError(
            f"Sheet '{sheet_name}': found {len(duplicated_ids)} duplicate Inverter ID(s), "
            f"which exceeds the allowed limit of {MAX_ALLOWED_DUPLICATE_INVERTERS}. "
            f"Duplicate Inverter IDs: {', '.join(duplicated_ids)}. "
            f"Please fix the duplicates in the source file and re-upload."
        )

    return duplicated_ids


@st.cache_data(show_spinner=False, ttl=3600)
def process_scada_excel_bytes(file_bytes, filename_hash=None):
    """
    Process SCADA file with caching based on file content.

    Returns a tuple: (processed_dfs, duplicate_warnings)
      - processed_dfs: {sheet_name: DataFrame}, one row per UNIQUE Inverter ID
        (duplicates de-duplicated, first occurrence kept).
      - duplicate_warnings: {sheet_name: [duplicate_inverter_id, ...]} for any
        sheet that had duplicates within the allowed limit (>MAX raises instead).

    Raises InverterIDValidationError if a sheet's Inverter ID column can't be
    found, its values don't look like Inverter IDs, or duplicates exceed the
    allowed limit.
    """
    file_stream = io.BytesIO(file_bytes)
    excel_file = pd.ExcelFile(file_stream, engine="openpyxl")
    processed_dfs = {}
    duplicate_warnings = {}
    sheet_skip_reasons = []

    for sheet_name in excel_file.sheet_names:
        try:
            df = read_sheet_with_fallback(file_stream, sheet_name)
        except Exception:
            continue

        df.dropna(how="all", inplace=True)
        if df.empty:
            continue
        df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed:", case=False, regex=True)]
        df = df.loc[:, ~df.columns.duplicated()].copy()

        actual_inverter_col = _find_inverter_column(df)

        if not actual_inverter_col:
            sheet_skip_reasons.append(
                f"Sheet '{sheet_name}': no Inverter ID column found. "
                f"Expected one of: {', '.join(INVERTER_ID_COLS)}. "
                f"Found columns: {', '.join(str(c) for c in df.columns)}."
            )
            continue

        # Reject sheets where the matched column's *values* don't look like
        # Inverter IDs (wrong column mapped to the expected header name).
        _validate_inverter_id_values(df, actual_inverter_col, sheet_name)

        # Detect duplicate Inverter IDs. Small numbers of duplicates are
        # tolerated (de-duplicated + reported); too many blocks the upload.
        duplicated_ids = _check_duplicate_inverter_ids(df, actual_inverter_col, sheet_name)
        if duplicated_ids:
            duplicate_warnings[sheet_name] = duplicated_ids
            # Only ever calculate metrics against UNIQUE Inverter IDs.
            df = df.drop_duplicates(subset=[actual_inverter_col], keep="first").reset_index(drop=True)

        df["Plot"] = df[actual_inverter_col].apply(extract_plot_cached)
        df["Block"] = df[actual_inverter_col].apply(extract_block_cached)
        df["SACU"] = df[actual_inverter_col].apply(map_inverter_to_sacu_cached)
        df = apply_string_metrics_cached(df)

        preferred_columns = [
            "Plot", "Block", actual_inverter_col, "SACU",
            "Total Active Strings", "Working String Count", "Failed String Count",
            "Availability (%)", "Failure Percentage (%)"
        ]
        remaining_cols = [c for c in df.columns if c not in preferred_columns]
        final_columns = [c for c in preferred_columns if c in df.columns] + remaining_cols
        df = df[final_columns]
        processed_dfs[sheet_name] = df

    if not processed_dfs and sheet_skip_reasons:
        raise InverterIDValidationError(
            "Could not find a valid Inverter ID column in any sheet.\n" + "\n".join(sheet_skip_reasons)
        )

    return processed_dfs, duplicate_warnings


def process_scada_excel_with_status(file_bytes, filename_hash=None, source_label="SCADA workbook"):
    """Run the cached parser behind a polished, step-by-step Streamlit status card."""
    steps = [
        ("📖", "Reading workbook", "Opening the file and detecting sheet structure..."),
        ("🧭", "Mapping identifiers", "Extracting Plot / Block / SACU from inverter IDs..."),
        ("⚡", "Calculating metrics", "Computing working/failed string counts and availability..."),
        ("💾", "Saving snapshot", "Persisting the processed snapshot for dashboards & history..."),
    ]

    duplicate_warnings = {}
    with st.status(f"Processing **{source_label}**", expanded=True) as status:
        progress = st.progress(0, text=f"{steps[0][0]}  {steps[0][1]}...")
        status.write(f"{steps[0][0]} **{steps[0][1]}** — {steps[0][2]}")
        progress.progress(20, text=f"{steps[1][0]}  {steps[1][1]}...")
        status.write(f"{steps[1][0]} **{steps[1][1]}** — {steps[1][2]}")
        progress.progress(55, text=f"{steps[2][0]}  {steps[2][1]}...")
        status.write(f"{steps[2][0]} **{steps[2][1]}** — {steps[2][2]}")

        try:
            processed_dfs, duplicate_warnings = process_scada_excel_bytes(file_bytes, filename_hash=filename_hash)
        except InverterIDValidationError as exc:
            progress.progress(100, text="Failed")
            status.update(label=f"❌ Could not process {source_label}", state="error", expanded=True)
            status.write(f"⚠️ {exc}")
            return {}

        progress.progress(85, text=f"{steps[3][0]}  {steps[3][1]}...")
        status.write(f"{steps[3][0]} **{steps[3][1]}** — {steps[3][2]}")

        if processed_dfs:
            progress.progress(100, text="Done")
            sheet_count = len(processed_dfs)
            row_count = sum(len(d) for d in processed_dfs.values())
            status.update(
                label=f"✅ {source_label} processed — {sheet_count} sheet(s), {row_count:,} row(s)",
                state="complete", expanded=False,
            )
            for sheet_name, dup_ids in duplicate_warnings.items():
                status.write(
                    f"⚠️ Sheet '{sheet_name}': {len(dup_ids)} duplicate Inverter ID(s) found and "
                    f"de-duplicated (first occurrence kept): {', '.join(dup_ids)}"
                )
        else:
            progress.progress(100, text="Failed")
            status.update(label=f"❌ Could not process {source_label}", state="error", expanded=True)

    if duplicate_warnings:
        for sheet_name, dup_ids in duplicate_warnings.items():
            st.warning(
                f"⚠️ Sheet '{sheet_name}': {len(dup_ids)} duplicate Inverter ID(s) found "
                f"(de-duplicated, first occurrence kept): {', '.join(dup_ids)}"
            )

    return processed_dfs

@st.cache_data(ttl=3600)
def find_header_row_index_cached(file_bytes, sheet_name, possible_header_columns, max_rows_to_check=100):
    """Cache header row detection"""
    file_stream = io.BytesIO(file_bytes)
    temp_df = pd.read_excel(file_stream, sheet_name=sheet_name, header=None,
                             nrows=max_rows_to_check, engine="openpyxl")
    possible_headers_lower = [str(col).strip().lower() for col in possible_header_columns]

    for i, row in temp_df.iterrows():
        row_values = [str(val).strip() for val in row.dropna()]
        row_values_lower = [v.lower() for v in row_values]
        if any(col in row_values_lower for col in possible_headers_lower):
            return i
    return None

def read_sheet_with_fallback(file_stream, sheet_name):
    file_bytes = file_stream.getvalue()
    header_row_index = find_header_row_index_cached(file_bytes, sheet_name, INVERTER_ID_COLS)
    file_stream.seek(0)
    if header_row_index is not None:
        df = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=header_row_index,
                            header=0, engine="openpyxl")
    else:
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=None, engine="openpyxl")
        df = assign_manual_headers(df, MANUAL_SCADA_COLUMNS)
    return df

def assign_manual_headers(df, manual_headers):
    manual_headers = clean_manual_columns(manual_headers)
    if len(df.columns) >= len(manual_headers):
        df = df.iloc[:, :len(manual_headers)].copy()
        df.columns = manual_headers
    else:
        df.columns = manual_headers[:len(df.columns)]
    return df

MANUAL_SCADA_COLUMNS = [
    "String Inverter", "MBUS", "Grid", "E-Daily(KWH)", "Active Power", "Reactive Power",
    "PV1", "PV2", "PV3", "PV4", "PV5", "PV6", "PV7", "PV8", "PV9", "PV10",
    "PV11", "PV12", "PV13", "PV14", "PV15", "PV16", "PV17", "PV18", "PV19", "PV20",
    "PV21", "PV22", "PV23", "PV24", "PV25", "PV26", "PV27", "PV28",
    "PV-I1", "PV-I2", "PV-I3", "PV-I4", "PV-I5", "PV-I6", "PV-I7", "PV-I8", "PV-I9", "PV-I10",
    "PV-I11", "PV-I12", "PV-I13", "PV-I14", "PV-I15", "PV-I16", "PV-I17", "PV-I18", "PV-I19", "PV-I20",
    "PV-I21", "PV-I22", "PV-I23", "PV-I24", "PV-I25", "PV-I26", "PV-I27", "PV-I28",
    "VAB", "VBC", "VCA", "IA", "IB", "IC"
]

ROLE_BADGES = {
    "admin": '<i class="fas fa-crown"></i> Admin',
    "manager": '<i class="fas fa-compass"></i> Manager',
    "engineer": '<i class="fas fa-wrench"></i> AM/Engineer',
}

def process_and_save_upload(file_bytes, filename, snapshot_date, username, role):
    snapshot_date_str = str(snapshot_date)

    try:
        processed, duplicate_warnings = process_scada_excel_bytes(file_bytes)
    except InverterIDValidationError as exc:
        return False, f"Upload rejected — {exc}"

    if not processed:
        return False, "Could not process this workbook - no valid sheets/inverter column found."

    upload_id, msg = storage1.save_preprocessed_upload(
        file_bytes=file_bytes, original_filename=filename,
        processed_dataframes=processed, snapshot_date=snapshot_date_str,
        uploaded_by=username,
    )

    if upload_id is None:
        return False, msg

    first_sheet = next(iter(processed))
    restore.update_string_history(processed[first_sheet].copy(), snapshot_date_str)
    restore.clear_snapshot_caches()

    if not msg.lower().startswith("file already exists"):
        audit_details = {"filename": filename, "snapshot_date": snapshot_date_str, "upload_id": upload_id}
        if duplicate_warnings:
            # Persist which Inverter IDs were duplicated so it can be shown
            # later on the Dashboard, not just in the upload status card.
            audit_details["duplicate_inverter_ids"] = duplicate_warnings
        storage1.log_audit_event(username, role, "file_uploaded", audit_details)

    if duplicate_warnings:
        dup_notes = []
        for sheet_name, dup_ids in duplicate_warnings.items():
            dup_notes.append(f"Sheet '{sheet_name}': {len(dup_ids)} duplicate Inverter ID(s) — {', '.join(dup_ids)}")
        msg = msg + "\n\n⚠️ Duplicate Inverter IDs were found and de-duplicated (first occurrence kept):\n" + "\n".join(dup_notes)

    return True, msg

def get_duplicate_inverter_warnings_for_snapshot(snapshot_date):
    """
    Look up the audit log for the 'file_uploaded' event that produced this
    snapshot date and return any recorded duplicate Inverter IDs
    ({sheet_name: [duplicate_id, ...]}). Returns {} if none were recorded.
    This is how duplicate-inverter details persist on the Dashboard after
    the upload itself has finished, instead of only flashing during upload.
    """
    if not snapshot_date:
        return {}
    snapshot_date_str = str(snapshot_date)
    try:
        entries = storage1.get_audit_log()
    except Exception:
        return {}

    matches = [
        e for e in entries
        if e.get("event_type") == "file_uploaded"
        and str((e.get("details") or {}).get("snapshot_date")) == snapshot_date_str
        and (e.get("details") or {}).get("duplicate_inverter_ids")
    ]
    if not matches:
        return {}

    matches.sort(key=lambda e: e.get("timestamp", ""), reverse=True)
    return matches[0]["details"]["duplicate_inverter_ids"]


def create_excel_download(dataframes_dict):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            safe_sheet_name = str(sheet_name)[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
    buffer.seek(0)
    return buffer.getvalue()


def _excel_cell_fill_hex(col_name, value):
    """Mirror the on-screen color rules (color_availability / color_health_status /
    color_failed_strings / get_string_health_color) as openpyxl fill colors, so
    exported Excel sheets look the same as the dashboard."""
    col_name = str(col_name)
    try:
        if col_name in ("Availability (%)", "Availability"):
            v = float(value)
            if v >= 90: return "10B981"
            if v >= 70: return "34D399"
            if v >= 50: return "FBBF24"
            if v >= 30: return "F59E0B"
            return "EF4444"
        if col_name == "Health Status":
            return {"Excellent": "10B981", "Good": "34D399", "Fair": "FBBF24", "Poor": "EF4444"}.get(str(value))
        if col_name in ("Failed Strings", "Failed", "Failed String Count"):
            v = float(value)
            if v == 0: return "10B981"
            if v <= 2: return "FBBF24"
            if v <= 5: return "F59E0B"
            return "EF4444"
        if col_name.startswith("PV-I"):
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return None
            v = float(value)
            if v > 5.0: return "10B981"
            if v > 3.0: return "34D399"
            if v > 1.5: return "FBBF24"
            if v > 0.5: return "F59E0B"
            return "EF4444"
    except (TypeError, ValueError):
        return None
    return None


def create_colored_excel_download(dataframes_dict):
    """Same as create_excel_download, but colors cells to match the on-screen
    UI (availability / health status / failed counts / PV string health)."""
    from openpyxl.styles import PatternFill, Font

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            safe_sheet_name = str(sheet_name)[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            worksheet = writer.sheets[safe_sheet_name]

            col_names = {idx + 1: col for idx, col in enumerate(df.columns)}
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx, col_name in col_names.items():
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    fill_hex = _excel_cell_fill_hex(col_name, cell.value)
                    if fill_hex:
                        cell.fill = PatternFill("solid", fgColor=fill_hex)
                        cell.font = Font(bold=True, color="FFFFFF")
    buffer.seek(0)
    return buffer.getvalue()


def display_negative_pv_values(df_result, sheet_name="Sheet1"):
    """Show PV-I channels with negative current values for the selected sheet."""
    st.markdown("---")
    st.markdown(
        "<div class='negative-report-card'><strong><i class='fas fa-circle-exclamation'></i> Negative PV Current Details</strong>"
        "<div class='calendar-subtitle'>Flags any PV-I channel below zero so sensor or parsing issues can be reviewed quickly.</div></div>",
        unsafe_allow_html=True,
    )

    if df_result is None or df_result.empty:
        st.info("No processed data available for negative PV inspection.")
        return

    inverter_col = get_inverter_column_cached(df_result)
    if not inverter_col:
        st.info("No inverter identifier column was found in this sheet.")
        return

    available_pv_cols = [col for col in PV_CURRENT_COLUMNS if col in df_result.columns]
    if not available_pv_cols:
        st.info("No PV-I columns were found in this sheet.")
        return

    negative_pv_values = []
    for _, row in df_result.iterrows():
        for pv_col in available_pv_cols:
            pv_value = pd.to_numeric(row.get(pv_col), errors="coerce")
            if pd.notna(pv_value) and pv_value < 0:
                negative_pv_values.append({
                    "Plot": row.get("Plot", ""),
                    "Block": row.get("Block", ""),
                    "SACU": row.get("SACU", ""),
                    "Inverter Name": row.get(inverter_col, ""),
                    "MPPT PV No": pv_col,
                    "PV Value": float(pv_value),
                    "Sheet": sheet_name,
                })

    if not negative_pv_values:
        st.success("No negative values found in the PV-I columns for this sheet.")
        return

    df_negative_pv = pd.DataFrame(negative_pv_values).sort_values(
        ["Plot", "Block", "SACU", "Inverter Name", "MPPT PV No"],
        ascending=True,
    ).reset_index(drop=True)

    metric_cols = st.columns(3)
    metric_cols[0].metric("Affected Inverters", df_negative_pv["Inverter Name"].nunique())
    metric_cols[1].metric("Negative Readings", len(df_negative_pv))
    metric_cols[2].metric("PV Columns Flagged", df_negative_pv["MPPT PV No"].nunique())

    styled_negative = df_negative_pv.style.map(
        lambda _: "background-color: #7f1d1d; color: white; font-weight: 700;",
        subset=["PV Value"],
    )
    st.dataframe(styled_negative, use_container_width=True, height=340)
    st.download_button(
        label="Download Negative PV Details (CSV)",
        data=df_negative_pv.to_csv(index=False),
        file_name=f"negative_pv_values_{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv",
        use_container_width=True,
        key=f"negative_pv_download_{sheet_name}",
    )

# ==========================================
# 6. UI - USER MANAGEMENT
# ==========================================
def user_management_ui():
    """
    admin   -> create/delete any user, change roles, assign plots, full access
    manager -> create engineer users only, CANNOT delete anyone, can assign plots
    engineer-> no access (menu not shown)

    NOTE ON ICONS: st.button / st.expander / st.selectbox labels render as
    PLAIN TEXT in Streamlit (there is no unsafe_allow_html for widget
    labels), so Font Awesome <i> tags never actually rendered here - they
    just showed up as literal text. Those labels now use plain text/emoji
    instead. Font Awesome is still used everywhere via st.markdown, which
    does support it.
    """
    current_user = get_current_user()
    if not current_user:
        return
    role = current_user.get("role")
    if not storage1.can_manage_users(role):
        return

    st.sidebar.markdown("---")
    st.sidebar.markdown('<i class="fas fa-users"></i> <b>User Management</b>', unsafe_allow_html=True)

    users = storage1.load_users()
    allowed_roles = storage1.creatable_roles(role)

    with st.sidebar.expander("👥 Manage Users", expanded=False):
        # ---- Create user ----
        st.write("### ➕ Create New User")
        new_full_name = st.text_input("Full Name", key="new_user_fullname")
        new_username = st.text_input("Username", key="new_user")
        new_password = st.text_input("Password", type="password", key="new_pass")
        new_role = st.selectbox("Role", allowed_roles, key="new_role")
        default_plots = storage1.ALL_PLOTS if new_role in ("admin", "manager") else storage1.ALL_PLOTS[:3]
        new_plots = st.multiselect("Assign Plots", storage1.ALL_PLOTS, default=default_plots, key="new_user_plots")

        if st.button("➕ Create User", key="create_user_btn", use_container_width=True):
            ok, msg = storage1.create_user(
                username=new_username, password=new_password, role=new_role,
                full_name=new_full_name, assigned_plots=new_plots,
                created_by=current_user.get("username"),
            )
            if ok:
                storage1.log_audit_event(current_user["username"], role, "user_created",
                                         {"created_user": new_username, "assigned_role": new_role})
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

        # ---- Existing users ----
        st.write("### 📋 Existing Users")
        for username, user_data in users.items():
            if username == current_user.get("username"):
                continue
            # Manager cannot see/manage admin accounts
            if role == "manager" and user_data.get("role") == "admin":
                continue

            col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
            with col1:
                st.write(f"**{user_data.get('full_name', username)}** (`{username}` - {user_data['role']})")
            with col2:
                if storage1.can_delete_users(role):
                    if st.button("🗑️", key=f"del_{username}", help="Delete this user"):
                        storage1.delete_user(username)
                        storage1.log_audit_event(current_user["username"], role, "user_deleted",
                                                 {"deleted_user": username})
                        st.rerun()
                else:
                    st.caption("no delete")
            with col3:
                if storage1.can_assign_plots(role) and user_data["role"] in ("engineer", "manager"):
                    if st.button("✏️", key=f"assign_{username}", help="Assign plots"):
                        st.session_state.assign_user = username
                        st.rerun()
            with col4:
                if storage1.can_reset_password_for(role, user_data["role"]):
                    if st.button("🔑", key=f"resetpwd_{username}", help="Reset password"):
                        st.session_state.reset_pwd_user = username
                        st.rerun()

        # ---- Password reset for another user (lost password) ----
        if "reset_pwd_user" in st.session_state:
            target_username = st.session_state.reset_pwd_user
            target_data = users.get(target_username)
            if target_data and storage1.can_reset_password_for(role, target_data["role"]):
                st.write(f"### 🔑 Reset Password for {target_data.get('full_name', target_username)}")
                admin_new_pw = st.text_input("New Password", type="password", key="admin_reset_pw_1")
                admin_new_pw_confirm = st.text_input("Confirm New Password", type="password", key="admin_reset_pw_2")
                if st.button("Confirm Reset", key="admin_reset_pw_confirm_btn"):
                    if not admin_new_pw or admin_new_pw != admin_new_pw_confirm:
                        st.error("Passwords don't match or are empty.")
                    else:
                        ok, msg = storage1.reset_password(target_username, admin_new_pw)
                        if ok:
                            storage1.log_audit_event(current_user["username"], role, "password_reset_admin",
                                                     {"target_user": target_username})
                            st.success(msg)
                            del st.session_state.reset_pwd_user
                            st.rerun()
                        else:
                            st.error(msg)
                if st.button("Cancel", key="admin_reset_pw_cancel_btn"):
                    del st.session_state.reset_pwd_user
                    st.rerun()
            else:
                del st.session_state.reset_pwd_user

        # ---- Plot assignment ----
        if "assign_user" in st.session_state:
            username = st.session_state.assign_user
            user_data = users.get(username)
            if user_data:
                st.write(f"### ✏️ Assign Plots for {user_data.get('full_name', username)}")
                assigned = user_data.get("assigned_plots", [])
                selected_plots = st.multiselect(
                    f"Select plots for {username}", options=storage1.ALL_PLOTS, default=assigned,
                    key="assign_plots_multiselect",
                )
                if st.button("Save Assignments", key="save_assign_plots"):
                    storage1.update_user_plots(username, selected_plots)
                    storage1.log_audit_event(current_user["username"], role, "plots_assigned",
                                             {"target_user": username, "plots": selected_plots})
                    st.success(f"Plots assigned for {username}")
                    del st.session_state.assign_user
                    st.rerun()
                if st.button("Cancel", key="cancel_assign_plots"):
                    del st.session_state.assign_user
                    st.rerun()

    # ---- Super-admin-only panel ----
    if storage1.is_super_admin(current_user.get("username")):
        render_super_admin_panel(current_user, role)


def render_super_admin_panel(current_user, role):
    """
    Super-admin-only sidebar panel:
      - export/download a fresh login-data backup
      - restore missing users from an uploaded backup
      - browse, download, and delete server-side saved backup snapshots
      - danger zone: reset ALL application data except user accounts
    """
    with st.sidebar.expander("🛡️ Super Admin Panel", expanded=False):
        st.caption("Restricted to the super admin account. Backs up login/user records so they survive app updates.")

        backup_bytes = storage1.export_users_backup()
        st.download_button(
            label="⬇️ Download New Backup",
            data=backup_bytes,
            file_name=f"users_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            key="download_users_backup_btn",
            use_container_width=True,
            on_click=_log_download, args=("users_backup",),
        )

        st.markdown("**♻️ Restore Missing Users**")
        st.caption("Only usernames absent from the current system are added back. Existing accounts are left untouched.")
        backup_upload = st.file_uploader("Upload backup file (.json)", type=["json"], key="users_backup_upload")
        if backup_upload is not None:
            if st.button("Restore Missing Users", key="restore_users_backup_btn", use_container_width=True):
                ok, msg, restored = storage1.import_users_backup(backup_upload.getvalue())
                if ok:
                    storage1.log_audit_event(
                        current_user["username"], role, "users_backup_restored",
                        {"restored_users": restored},
                    )
                    st.success(msg)
                    if restored:
                        st.rerun()
                else:
                    st.error(msg)

        st.markdown("---")
        st.markdown("**📦 Manage Saved Backups**")
        saved_backups = storage1.list_user_backups()
        if not saved_backups:
            st.caption("No server-side backup snapshots yet. Download a backup above to create one.")
        else:
            st.caption(f"{len(saved_backups)} snapshot(s) stored on the server.")
            for b in saved_backups:
                bcol1, bcol2, bcol3 = st.columns([3, 1, 1])
                with bcol1:
                    st.write(f"`{b['filename']}`")
                    st.caption(f"{b['created']} · {b['size_kb']} KB · {b.get('user_count', '?')} users")
                with bcol2:
                    backup_data = storage1.get_user_backup_bytes(b["filename"])
                    if backup_data:
                        st.download_button(
                            "⬇️", data=backup_data, file_name=b["filename"], mime="application/json",
                            key=f"dl_backup_{b['filename']}", help="Download this snapshot",
                        )
                with bcol3:
                    if st.button("🗑️", key=f"del_backup_{b['filename']}", help="Delete this snapshot"):
                        ok, msg = storage1.delete_user_backup(b["filename"])
                        storage1.log_audit_event(current_user["username"], role, "backup_deleted", {"filename": b["filename"]})
                        if ok:
                            st.success(msg)
                        else:
                            st.error(msg)
                        st.rerun()

            if st.button("🗑️ Delete ALL Saved Backups", key="delete_all_backups_btn", use_container_width=True):
                st.session_state.confirm_delete_all_backups = True
                st.rerun()

            if st.session_state.get("confirm_delete_all_backups"):
                st.warning("This will permanently delete every saved backup snapshot. This cannot be undone.")
                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("Yes, delete all", key="confirm_delete_all_backups_yes", use_container_width=True):
                        ok, msg = storage1.delete_all_user_backups()
                        storage1.log_audit_event(current_user["username"], role, "all_backups_deleted", {})
                        st.success(msg)
                        del st.session_state.confirm_delete_all_backups
                        st.rerun()
                with cc2:
                    if st.button("Cancel", key="confirm_delete_all_backups_no", use_container_width=True):
                        del st.session_state.confirm_delete_all_backups
                        st.rerun()

        st.markdown("---")
        st.markdown(
            "<div style='background:#7f1d1d; border:1px solid #ef4444; border-radius:10px; padding:10px 12px;'>"
            "<strong style='color:#fecaca;'>⚠️ Danger Zone — Reset Application</strong>"
            "<div style='color:#fecaca; font-size:0.82rem; margin-top:4px;'>"
            "Wipes ALL uploaded SCADA workbooks, preprocessed snapshots, and the audit log. "
            "<b>User accounts &amp; passwords are NEVER touched</b> by this action.</div></div>",
            unsafe_allow_html=True,
        )
        reset_confirm_text = st.text_input(
            "Type RESET to confirm", key="reset_app_confirm_text",
            placeholder="Type RESET to enable the button below",
        )
        reset_disabled = reset_confirm_text.strip().upper() != "RESET"
        if st.button("🧨 Reset Application Data", key="reset_app_data_btn",
                     use_container_width=True, disabled=reset_disabled, type="primary"):
            ok, msg, details = storage1.reset_application_data()
            storage1.log_audit_event(current_user["username"], role, "application_reset", details)
            if ok:
                st.success(msg)
            else:
                st.error(msg)
            del st.session_state["reset_app_confirm_text"]
            st.rerun()

def audit_log_tab():
    """Role-scoped audit log, shown in its own dashboard tab"""
    current_user = get_current_user()
    role = current_user.get("role")
    username = current_user.get("username")

    if not storage1.can_view_audit_log(role):
        st.info("Audit log isn't available for your role.")
        return

    st.markdown('<i class="fas fa-search"></i> Audit Log', unsafe_allow_html=True)

    intact = storage1.verify_audit_chain()
    if intact:
        st.markdown(
            '<div style="padding:0.75rem 1rem; border-radius:0.5rem; background:#064e3b; color:#ecfdf5; font-weight:600;">'
            '<i class="fas fa-circle-check"></i> Log integrity verified (hash chain intact)</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div style="padding:0.75rem 1rem; border-radius:0.5rem; background:#7f1d1d; color:#fef2f2; font-weight:600;">'
            '<i class="fas fa-triangle-exclamation"></i> Log integrity check FAILED - entries may have been altered</div>',
            unsafe_allow_html=True,
        )

    entries = storage1.get_audit_log_for(username, role)
    if not entries:
        st.info("No audit events to show.")
        return

    df_audit = pd.DataFrame(entries)[["timestamp", "username", "role", "event_type", "details", "entry_hash"]]
    df_audit = df_audit.sort_values("timestamp", ascending=False)

    # Details come back as raw dicts (e.g. {"duplicate_inverter_ids": {...}}),
    # which Streamlit's Arrow-backed st.dataframe can't render directly and
    # was silently breaking this tab. Flatten to a readable JSON string.
    df_audit["details"] = df_audit["details"].apply(
        lambda d: json.dumps(d, default=str, ensure_ascii=False) if isinstance(d, dict) else ("" if d is None else str(d))
    )

    scope_label = "All Users" if role == "admin" else "Engineers" if role == "manager" else "My Activity"

    filter_col1, filter_col2, metric_col = st.columns([1, 1, 1])
    with filter_col1:
        event_types = ["All"] + sorted_filter_options(df_audit["event_type"])
        selected_event = st.selectbox("Filter by Event Type", event_types, key="audit_event_filter")
    with filter_col2:
        if role == "admin":
            users_in_log = ["All"] + sorted_filter_options(df_audit["username"])
            selected_user = st.selectbox("Filter by User", users_in_log, key="audit_user_filter")
        else:
            selected_user = "All"
    with metric_col:
        st.metric("Scope", scope_label)

    filtered = df_audit.copy()
    if selected_event != "All":
        filtered = filtered[filtered["event_type"] == selected_event]
    if selected_user != "All":
        filtered = filtered[filtered["username"] == selected_user]

    st.caption(f"Showing {len(filtered)} of {len(df_audit)} event(s) you have access to.")
    st.dataframe(filtered.drop(columns=["entry_hash"]), use_container_width=True, height=450)

    # ---- Super-admin-only: delete audit log entries (any user) ----
    if storage1.is_super_admin(username):
        with st.expander("🛡️ Super Admin - Delete Audit Log Entries", expanded=False):
            st.caption(
                "Permanently remove one or more log entries below, for any user. "
                "The hash chain is automatically rebuilt afterwards so the integrity check above still passes."
            )
            if filtered.empty:
                st.info("No entries to delete for the current filter selection.")
            else:
                options_df = filtered.copy()
                options_df["_label"] = (
                    options_df["timestamp"] + " · " + options_df["username"] + " · " + options_df["event_type"]
                )
                label_to_hash = dict(zip(options_df["_label"], options_df["entry_hash"]))
                to_delete_labels = st.multiselect(
                    "Select entries to delete", options=options_df["_label"].tolist(), key="audit_delete_select",
                )
                if st.button("🗑️ Delete Selected Entries", key="audit_delete_btn", type="primary"):
                    if not to_delete_labels:
                        st.warning("Select at least one entry first.")
                    else:
                        hashes_to_delete = [label_to_hash[lbl] for lbl in to_delete_labels if lbl in label_to_hash]
                        ok, del_msg, deleted_count = storage1.delete_audit_entries(hashes_to_delete)
                        if ok:
                            storage1.log_audit_event(
                                username, role, "audit_entries_deleted",
                                {"deleted_count": deleted_count},
                            )
                            st.success(del_msg)
                            st.rerun()
                        else:
                            st.error(del_msg)

# ==========================================
# 7. UI - OPTIMIZED MAIN TABS WITH CACHING
# ==========================================

# Cache the inverter column detection
@st.cache_data(ttl=3600)
def get_inverter_column_cached(df):
    """Cache the inverter column detection"""
    df_columns_lower_map = {str(c).strip().lower(): c for c in df.columns}
    for col in INVERTER_ID_COLS:
        if col in df.columns:
            return col
        elif col.strip().lower() in df_columns_lower_map:
            return df_columns_lower_map[col.strip().lower()]
    return None

# Cache filter options with hash
def _fast_df_fingerprint(df):
    """
    Fast content fingerprint used as a Streamlit cache key for DataFrame
    arguments. Hashing only df.shape[0] (row count) is unsafe: two different
    filter selections (or a freshly re-uploaded snapshot) that happen to
    produce the same row count would silently reuse another selection's
    stale cached result - a real source of wrong numbers elsewhere in this
    tab. pd.util.hash_pandas_object is vectorized and still far cheaper than
    Streamlit's default full-object hash, but actually reflects the data.
    """
    if df is None or df.empty:
        return 0
    try:
        return int(pd.util.hash_pandas_object(df, index=True).sum())
    except Exception:
        return (df.shape, tuple(df.columns))


@st.cache_data(ttl=3600, hash_funcs={pd.DataFrame: _fast_df_fingerprint})
def get_filter_options_cached(df, column):
    """Cache filter options for a specific column with dataframe hash"""
    if column in df.columns:
        return sorted_filter_options(df[column])
    return []

# Cache summary metrics calculation with hash - fully vectorized (no iterrows).
# The previous per-row `.iterrows()` loop was the main reason the PV String
# Details tab felt slow: every filter change re-ran that loop over the whole
# (filtered) sheet. Vectorized pandas ops give the same output far faster.
@st.cache_data(ttl=300, hash_funcs={pd.DataFrame: _fast_df_fingerprint})
def calculate_summary_metrics_cached(filtered_df, inverter_col, pv_voltage_cols, pv_current_cols):
    """Vectorized summary metrics calculation."""
    if filtered_df.empty:
        return pd.DataFrame()

    df = filtered_df.copy()
    n = len(df)

    has_precomputed = "Total Active Strings" in df.columns and "Working String Count" in df.columns

    if has_precomputed:
        total_strings = pd.to_numeric(df["Total Active Strings"], errors="coerce").fillna(0).astype(int)
        working_strings = pd.to_numeric(df["Working String Count"], errors="coerce").fillna(0).astype(int)
        failed_strings = pd.to_numeric(
            df.get("Failed String Count", total_strings - working_strings), errors="coerce"
        ).fillna(0).astype(int)
        availability = pd.to_numeric(df.get("Availability (%)"), errors="coerce").fillna(0)
    else:
        present_current_cols = [c for c in pv_current_cols if c in df.columns]
        if present_current_cols:
            numeric_currents = df[present_current_cols].apply(pd.to_numeric, errors="coerce")
            total_strings = numeric_currents.notna().sum(axis=1)
            working_strings = (numeric_currents > WORKING_CURRENT_THRESHOLD).sum(axis=1)
        else:
            total_strings = pd.Series(0, index=df.index)
            working_strings = pd.Series(0, index=df.index)
        failed_strings = (total_strings - working_strings).clip(lower=0)
        availability = (working_strings / total_strings.replace(0, pd.NA) * 100).fillna(0)

    present_voltage_cols = [c for c in pv_voltage_cols if c in df.columns]
    if present_voltage_cols:
        numeric_voltage = df[present_voltage_cols].apply(pd.to_numeric, errors="coerce")
        avg_voltage = numeric_voltage.mean(axis=1).fillna(0)
    else:
        avg_voltage = pd.Series(0.0, index=df.index)

    present_current_cols_avg = [c for c in pv_current_cols if c in df.columns]
    if present_current_cols_avg:
        numeric_current_avg = df[present_current_cols_avg].apply(pd.to_numeric, errors="coerce")
        avg_current = numeric_current_avg.mean(axis=1).fillna(0)
    else:
        avg_current = pd.Series(0.0, index=df.index)

    health_status = pd.cut(
        availability, bins=[-0.1, 50, 70, 90, 100.1],
        labels=["Poor", "Fair", "Good", "Excellent"],
    ).astype(str)

    summary_df = pd.DataFrame({
        "Inverter ID": df[inverter_col].values if inverter_col in df.columns else [""] * n,
        "Plot": df.get("Plot", pd.Series([""] * n)).values,
        "Block": df.get("Block", pd.Series([""] * n)).values,
        "SACU": df.get("SACU", pd.Series([""] * n)).values,
        "Total Strings": total_strings.values,
        "Working Strings": working_strings.values,
        "Failed Strings": failed_strings.values,
        "Availability (%)": availability.round(2).values,
        "Health Status": health_status.values,
        "Avg PV Voltage (V)": avg_voltage.round(1).values,
        "Avg PV Current (A)": avg_current.round(2).values,
        "Grid": df.get("Grid", pd.Series([""] * n)).values,
        "E-Daily (KWH)": df.get("E-Daily(KWH)", pd.Series([""] * n)).values,
        "Active Power (KW)": df.get("Active Power", pd.Series([""] * n)).values,
        "Reactive Power (KVAR)": df.get("Reactive Power", pd.Series([""] * n)).values,
    })

    return summary_df

# Cache the filtered dataframe with hash
@st.cache_data(ttl=60, hash_funcs={pd.DataFrame: _fast_df_fingerprint})
def apply_filters_cached(df, selected_plot, selected_block, selected_sacu, selected_inverter,
                         selected_grid, selected_status, inverter_col):
    """Cache the filtered dataframe result"""
    filtered_df = df.copy()

    if selected_plot != "All":
        filtered_df = filtered_df[filtered_df["Plot"] == selected_plot]
    if selected_block != "All":
        filtered_df = filtered_df[filtered_df["Block"] == selected_block]
    if selected_sacu != "All":
        filtered_df = filtered_df[filtered_df["SACU"] == selected_sacu]
    if selected_inverter != "All":
        filtered_df = filtered_df[filtered_df[inverter_col] == selected_inverter]
    if selected_grid != "All" and "Grid" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["Grid"] == selected_grid]

    if selected_status != "All" and "Availability (%)" in filtered_df.columns:
        if selected_status == "Working (>=70%)":
            filtered_df = filtered_df[filtered_df["Availability (%)"] >= 70]
        elif selected_status == "Fair (50-70%)":
            filtered_df = filtered_df[(filtered_df["Availability (%)"] >= 50) & (filtered_df["Availability (%)"] < 70)]
        elif selected_status == "Failed (<50%)":
            filtered_df = filtered_df[filtered_df["Availability (%)"] < 50]
        elif selected_status == "Critical (<30%)":
            filtered_df = filtered_df[filtered_df["Availability (%)"] < 30]

    return filtered_df

# Color functions defined at module level for reuse
def color_availability(val):
    if isinstance(val, (int, float)):
        if val >= 90: return 'background-color: #10b981; color: white; font-weight: bold'
        elif val >= 70: return 'background-color: #34d399; color: white; font-weight: bold'
        elif val >= 50: return 'background-color: #fbbf24; color: black; font-weight: bold'
        elif val >= 30: return 'background-color: #f59e0b; color: white; font-weight: bold'
        else: return 'background-color: #ef4444; color: white; font-weight: bold'
    return ''

def color_health_status(val):
    if val == "Excellent": return 'background-color: #10b981; color: white; font-weight: bold'
    elif val == "Good": return 'background-color: #34d399; color: white; font-weight: bold'
    elif val == "Fair": return 'background-color: #fbbf24; color: black; font-weight: bold'
    elif val == "Poor": return 'background-color: #ef4444; color: white; font-weight: bold'
    return ''

def color_failed_strings(val):
    if isinstance(val, (int, float)):
        if val == 0: return 'background-color: #10b981; color: white; font-weight: bold'
        elif val <= 2: return 'background-color: #fbbf24; color: black; font-weight: bold'
        elif val <= 5: return 'background-color: #f59e0b; color: white; font-weight: bold'
        else: return 'background-color: #ef4444; color: white; font-weight: bold'
    return ''

@st.cache_data(ttl=300, hash_funcs={pd.DataFrame: _fast_df_fingerprint})
def find_low_performance_strings_cached(filtered_df, inverter_col, pv_current_cols):
    """
    Return PV string rows whose current is >= LOW_PERFORMANCE_DROP_PCT (30%)
    below the average of the same inverter's own working strings - matches
    the Individual String Details tab's classification exactly. Vectorized
    (no per-row Python loop) so it stays fast on large sheets.
    """
    present_cols = [c for c in pv_current_cols if c in filtered_df.columns]
    if filtered_df is None or filtered_df.empty or not present_cols or inverter_col not in filtered_df.columns:
        return pd.DataFrame()

    numeric = filtered_df[present_cols].apply(pd.to_numeric, errors="coerce")
    working_mask = numeric > WORKING_CURRENT_THRESHOLD
    average_current = numeric.where(working_mask).mean(axis=1)
    low_performance_threshold = average_current * (1 - LOW_PERFORMANCE_DROP_PCT)

    low_mask = working_mask & numeric.lt(low_performance_threshold, axis=0) & (average_current > 0)
    if not low_mask.to_numpy().any():
        return pd.DataFrame()

    context_cols = [c for c in [inverter_col, "Plot", "Block", "SACU", "Grid"] if c in filtered_df.columns]
    context = filtered_df[context_cols].copy()
    context["Inverter Avg Current (Working Strings)"] = average_current.round(2)
    context[f"Low Performance Threshold ({LOW_PERFORMANCE_DROP_PCT*100:.0f}% below avg)"] = low_performance_threshold.round(2)

    values_only = numeric.where(low_mask)
    melted = values_only.join(context).reset_index(drop=True).melt(
        id_vars=context_cols + ["Inverter Avg Current (Working Strings)",
                                 f"Low Performance Threshold ({LOW_PERFORMANCE_DROP_PCT*100:.0f}% below avg)"],
        value_vars=present_cols, var_name="MPPT PV No", value_name="PV Value",
    )
    melted = melted.dropna(subset=["PV Value"]).reset_index(drop=True)
    melted["PV Value"] = melted["PV Value"].round(2)
    if inverter_col in melted.columns and inverter_col != "Inverter Name":
        melted = melted.rename(columns={inverter_col: "Inverter Name"})
    return melted



def create_pv_string_tab(df):
    """Create the inverter-wise PV string details tab with optimized caching"""

    st.markdown("""
    <div style='display: flex; align-items: center; margin-bottom: 0px;'>
        <h2 style='margin: 0; color: #38bdf8;'><i class='fas fa-plug'></i> Inverter-wise PV String Details</h2>
        <span style='margin-left: 12px; font-size: 0.9rem; color: #94a3b8;'><i class='fas fa-chevron-right' style='font-size: 0.7rem;'></i> Individual string performance</span>
    </div>
    <p style='color: #94a3b8; font-size: 0.85rem; margin-top: 4px;'>Color-coded headers show string health status</p>
    """, unsafe_allow_html=True)

    # Use cached inverter column detection
    inverter_col = get_inverter_column_cached(df)
    if not inverter_col:
        st.warning("No inverter ID column found in the dataset")
        return

    # Use cached PV columns detection
    pv_voltage_cols, pv_current_cols = get_pv_string_columns_cached(df)
    if not pv_voltage_cols and not pv_current_cols:
        st.warning("No PV string data columns found in the dataset")
        return

    # ==========================================
    # FILTERS SECTION - live & cascading (Plot -> Block -> SACU -> Inverter).
    #
    # These used to sit behind a "Search" button, with every dropdown's
    # options coming from the whole sheet regardless of the other
    # selections - so choosing a Plot didn't narrow the Block list, and
    # nothing actually applied until a second click on Search. Both of
    # those are fixed here: each selectbox reruns immediately (no form),
    # and each one's options are derived from the *already-selected*
    # filters above it, so the dropdowns stay in sync with each other and
    # with what's actually shown below.
    # ==========================================

    if "pv_filters" not in st.session_state:
        st.session_state.pv_filters = {
            "plot": "All", "block": "All", "sacu": "All", "inverter": "All",
            "grid": "All", "status": "All", "show_voltage": False, "show_current": True,
        }

    status_options = ["All", "Working (>=70%)", "Fair (50-70%)", "Failed (<50%)", "Critical (<30%)"]

    st.markdown(
        "<div style='font-size:0.85rem; color:#94a3b8; margin-bottom:4px;'>"
        "<i class='fas fa-filter'></i> Filters apply immediately as you change them.</div>",
        unsafe_allow_html=True,
    )
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        plot_choices = ["All"] + get_filter_options_cached(df, "Plot")
        prev_plot = st.session_state.pv_filters["plot"]
        selected_plot = st.selectbox(
            "Plot", plot_choices,
            index=plot_choices.index(prev_plot) if prev_plot in plot_choices else 0,
            key="pv_plot_filter",
        )

    # Each following dropdown's options are narrowed by everything selected so far.
    df_after_plot = df if selected_plot == "All" else df[df["Plot"] == selected_plot]

    with col2:
        block_choices = ["All"] + sorted_filter_options(df_after_plot["Block"]) if "Block" in df_after_plot.columns else ["All"]
        prev_block = st.session_state.pv_filters["block"]
        selected_block = st.selectbox(
            "Block", block_choices,
            index=block_choices.index(prev_block) if prev_block in block_choices else 0,
            key="pv_block_filter",
        )

    df_after_block = df_after_plot if selected_block == "All" else df_after_plot[df_after_plot["Block"] == selected_block]

    with col3:
        sacu_choices = ["All"] + sorted_filter_options(df_after_block["SACU"]) if "SACU" in df_after_block.columns else ["All"]
        prev_sacu = st.session_state.pv_filters["sacu"]
        selected_sacu = st.selectbox(
            "SACU", sacu_choices,
            index=sacu_choices.index(prev_sacu) if prev_sacu in sacu_choices else 0,
            key="pv_sacu_filter",
        )

    df_after_sacu = df_after_block if selected_sacu == "All" else df_after_block[df_after_block["SACU"] == selected_sacu]

    with col4:
        inverter_choices = ["All"] + sorted_filter_options(df_after_sacu[inverter_col])
        prev_inverter = st.session_state.pv_filters["inverter"]
        selected_inverter = st.selectbox(
            "Inverter", inverter_choices,
            index=inverter_choices.index(prev_inverter) if prev_inverter in inverter_choices else 0,
            key="pv_inverter_filter",
        )
    with col5:
        show_voltage = st.checkbox("Show Voltage", value=st.session_state.pv_filters["show_voltage"], key="show_voltage")
        show_current = st.checkbox("Show Current", value=st.session_state.pv_filters["show_current"], key="show_current")

    col6, col7 = st.columns(2)
    with col6:
        available_grids = get_filter_options_cached(df, "Grid") if "Grid" in df.columns else []
        grid_choices = ["All"] + available_grids
        prev_grid = st.session_state.pv_filters["grid"]
        selected_grid = st.selectbox(
            "Grid", grid_choices,
            index=grid_choices.index(prev_grid) if prev_grid in grid_choices else 0,
            key="pv_grid_filter",
        )
    with col7:
        prev_status = st.session_state.pv_filters["status"]
        selected_status = st.selectbox(
            "Status", status_options,
            index=status_options.index(prev_status) if prev_status in status_options else 0,
            key="pv_status_filter",
        )

    st.session_state.pv_filters.update({
        "plot": selected_plot, "block": selected_block, "sacu": selected_sacu,
        "inverter": selected_inverter, "grid": selected_grid, "status": selected_status,
        "show_voltage": show_voltage, "show_current": show_current,
    })

    active_chips = [f"{label}: {value}" for label, value in
                    [("Plot", selected_plot), ("Block", selected_block), ("SACU", selected_sacu),
                     ("Inverter", selected_inverter), ("Grid", selected_grid), ("Status", selected_status)]
                    if value != "All"]
    if active_chips:
        st.caption("🔎 Active filters — " + " · ".join(active_chips))
    else:
        st.caption("No filters applied — showing all data.")

    # Apply filters with caching
    filter_key = f"{selected_plot}_{selected_block}_{selected_sacu}_{selected_inverter}_{selected_grid}_{selected_status}"

    if st.session_state.get("pv_filter_key") != filter_key:
        st.session_state.filtered_df = apply_filters_cached(
            df, selected_plot, selected_block, selected_sacu, selected_inverter,
            selected_grid, selected_status, inverter_col
        )
        st.session_state.pv_filter_key = filter_key

    filtered_df = st.session_state.filtered_df

    if filtered_df.empty:
        st.warning("No data available for the selected filters")
        return

    # ==========================================
    # BUILD SUMMARY METRICS - Cached
    # ==========================================
    summary_key = f"{filter_key}_{len(filtered_df)}"

    if st.session_state.get("pv_summary_key") != summary_key:
        st.session_state.summary_df = calculate_summary_metrics_cached(
            filtered_df, inverter_col, pv_voltage_cols, pv_current_cols
        )
        st.session_state.pv_summary_key = summary_key

    summary_df = st.session_state.summary_df

    # ==========================================
    # CREATE THREE TABS INSIDE PV STRING DETAILS
    # (st.tabs only renders plain text, so Font Awesome markup can't be
    # used for these labels - kept as clean text instead of emoji)
    # ==========================================
    tab1, tab2, tab3, tab4 = st.tabs([
        "INV-PV Details",
        "Individual String Details",
        "Failed Inverters",
        "Low Performance Strings"
    ])

    # ==========================================
    # TAB 1: INV-PV DETAILS (DEFAULT)
    # ==========================================
    with tab1:
        # INVERTER SUMMARY METRICS
        st.markdown('<h3><i class="fas fa-chart-simple"></i> Inverter Summary</h3>', unsafe_allow_html=True)

        total_inverters = len(summary_df)
        total_working = summary_df["Working Strings"].sum()
        total_strings_all = summary_df["Total Strings"].sum()
        total_failed = summary_df["Failed Strings"].sum()
        overall_availability = (total_working / total_strings_all * 100) if total_strings_all > 0 else 0
        avail_color = "#10b981" if overall_availability >= 90 else "#34d399" if overall_availability >= 70 else "#fbbf24" if overall_availability >= 50 else "#ef4444"

        render_kpi_cards([
            {"label": "TOTAL INVERTERS", "value": f"{total_inverters:,}", "icon": "fas fa-microchip", "color": "#818cf8", "sub": "Unique Inverter IDs"},
            {"label": "TOTAL STRINGS", "value": f"{int(total_strings_all):,}", "icon": "fas fa-plug", "color": "#a78bfa"},
            {"label": "WORKING STRINGS", "value": f"{int(total_working):,}", "icon": "fas fa-circle-check", "color": "#10b981"},
            {"label": "FAILED STRINGS", "value": f"{int(total_failed):,}", "icon": "fas fa-circle-xmark", "color": "#ef4444"},
            {"label": "AVAILABILITY", "value": f"{overall_availability:.1f}%", "icon": "fas fa-gauge-high", "color": avail_color},
        ])

        # BLOCK-WISE KPI CARDS - scoped to whatever Plot/Block/SACU/Inverter
        # filters are currently selected above, so picking a Plot narrows
        # these cards down to that plot's blocks automatically.
        st.markdown("---")
        block_summary = calculate_block_summary_cached(filtered_df, inverter_col)
        if not block_summary.empty:
            plot_context = selected_plot if selected_plot != "All" else "All Plots"
            display_block_metrics(block_summary, plot_context)

        # INVERTER-WISE SUMMARY TABLE
        st.markdown("---")
        st.markdown('<h3><i class="fas fa-table-list"></i> Inverter-wise Summary</h3>', unsafe_allow_html=True)

        styled_summary = get_styled_summary(summary_df)
        st.dataframe(styled_summary, use_container_width=True)
        st.download_button(
            label="Download Summary (Excel, color-coded)",
            data=create_colored_excel_download({"Inverter Summary": summary_df}),
            file_name=f"inverter_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_inverter_summary_xlsx",
            on_click=_log_download, args=("inverter_summary_xlsx",),
        )

        # DETAILED PV STRING DATA
        st.markdown("---")
        st.markdown('<h3><i class="fas fa-bolt"></i> Detailed PV String Data</h3>', unsafe_allow_html=True)
        st.caption("Green = Good (>5A), Yellow = Fair (1.5-5A), Orange = Poor (0.5-1.5A), Red = Critical (<0.5A)")

        # Build display columns efficiently
        display_cols = [inverter_col, "Plot", "Block", "SACU"]
        base_cols = ["Total Active Strings", "Working String Count", "Failed String Count", "Availability (%)", "Failure Percentage (%)"]
        for col in base_cols:
            if col in filtered_df.columns:
                display_cols.append(col)

        additional_cols = ["Grid", "E-Daily(KWH)", "Active Power", "Reactive Power", "VAB", "VBC", "VCA", "IA", "IB", "IC"]
        for col in additional_cols:
            if col in filtered_df.columns:
                display_cols.append(col)

        pv_columns_to_show = []
        if show_voltage:
            pv_columns_to_show.extend(sorted(pv_voltage_cols))
        if show_current:
            pv_columns_to_show.extend(sorted(pv_current_cols))
        display_cols.extend(pv_columns_to_show)

        display_df = filtered_df[display_cols].copy()

        # Rename columns efficiently
        rename_map = {inverter_col: "Inverter ID"}
        rename_pairs = {
            "E-Daily(KWH)": "Energy (KWh)", "Active Power": "Active Power (KW)",
            "Reactive Power": "Reactive Power (KVAR)", "Total Active Strings": "Total Strings",
            "Working String Count": "Working", "Failed String Count": "Failed",
            "Failure Percentage (%)": "Failure %", "VAB": "VAB (V)", "VBC": "VBC (V)",
            "VCA": "VCA (V)", "IA": "IA (A)", "IB": "IB (A)", "IC": "IC (A)",
        }
        for k, v in rename_pairs.items():
            if k in display_df.columns:
                rename_map[k] = v
        display_df = display_df.rename(columns=rename_map)

        pv_current_cols_display = [col for col in pv_current_cols if col in display_df.columns]
        pv_voltage_cols_display = [col for col in pv_voltage_cols if col in display_df.columns]

        def apply_detailed_styling(df_display):
            styled = df_display.style
            for col in pv_columns_to_show:
                if col in df_display.columns:
                    non_null = df_display[col].notna().sum()
                    if non_null > 0:
                        working_count = (df_display[col] > WORKING_CURRENT_THRESHOLD).sum()
                        working_pct = (working_count / non_null) * 100
                        color = get_column_header_color(working_pct)
                        styled = styled.set_table_styles(
                            [{'selector': f'th.col{df_display.columns.get_loc(col)}',
                              'props': [('background-color', color), ('color', 'white'), ('font-weight', 'bold')]}],
                            overwrite=False
                        )
            for col in pv_current_cols_display:
                if col in df_display.columns:
                    styled = styled.map(
                        lambda x: f'background-color: {get_string_health_color(x)}; color: white; font-weight: bold;'
                        if pd.notna(x) and isinstance(x, (int, float)) else '', subset=[col]
                    )
            if "Availability (%)" in df_display.columns:
                styled = styled.map(color_availability, subset=['Availability (%)'])
            if "Failed" in df_display.columns:
                styled = styled.map(color_failed_strings, subset=['Failed'])
            styled = styled.set_table_styles([
                {'selector': 'thead th', 'props': [('position', 'sticky'), ('top', '0'), ('z-index', '999')]},
                {'selector': 'td', 'props': [('padding', '2px 4px'), ('font-size', '12px')]},
                {'selector': 'th', 'props': [('padding', '4px 8px'), ('font-size', '11px')]}
            ], overwrite=False)
            return styled

        if not display_df.empty:
            try:
                styled_df = apply_detailed_styling(display_df)
                st.dataframe(styled_df, use_container_width=True, height=400)
            except Exception as e:
                st.warning(f"Styling error: {str(e)}. Showing unstyled data.")
                st.dataframe(display_df, use_container_width=True, height=400)

    # ==========================================
    # TAB 2: INDIVIDUAL STRING DETAILS
    # ==========================================
    with tab2:
        st.markdown('<h3><i class="fas fa-bolt"></i> Individual String Details</h3>', unsafe_allow_html=True)
        st.caption("Detailed view of individual string performance with filters")

        # ==========================================
        # INDIVIDUAL STRING FILTERS
        # ==========================================
        filter_col1, filter_col2, filter_col3 = st.columns([2, 2, 2])

        with filter_col1:
            # Select inverter
            inverter_list = get_filter_options_cached(filtered_df, inverter_col)
            selected_inverter_detail = st.selectbox(
                "Select Inverter",
                inverter_list,
                key="detail_inverter_select"
            )

        with filter_col2:
            # String selection or filter
            string_options = ["All Strings", "Working Only", "Failed Only", "Critical (<0.5A)"]
            string_filter = st.selectbox("String Status Filter", string_options, key="string_status_filter")

        with filter_col3:
            # Grid filter for detail view
            if "Grid" in filtered_df.columns:
                grid_options_detail = ["All"] + get_filter_options_cached(filtered_df, "Grid")
                grid_filter_detail = st.selectbox("Filter by Grid", grid_options_detail, key="detail_grid_filter")
            else:
                grid_filter_detail = "All"

        # Get the selected inverter data
        if selected_inverter_detail:
            # Apply grid filter to inverter data
            inverter_data_df = filtered_df.copy()
            if grid_filter_detail != "All" and "Grid" in inverter_data_df.columns:
                inverter_data_df = inverter_data_df[inverter_data_df["Grid"] == grid_filter_detail]

            # Get the specific inverter
            if selected_inverter_detail in inverter_data_df[inverter_col].values:
                inverter_data = inverter_data_df[inverter_data_df[inverter_col] == selected_inverter_detail].iloc[0]

                # ==========================================
                # INVERTER DETAILS HEADER
                # ==========================================
                col1, col2, col3, col4, col5, col6 = st.columns(6)
                with col1:
                    st.metric("Inverter", inverter_data[inverter_col])
                with col2:
                    st.metric("Plot", inverter_data.get("Plot", "N/A"))
                with col3:
                    st.metric("Block", inverter_data.get("Block", "N/A"))
                with col4:
                    st.metric("SACU", inverter_data.get("SACU", "N/A"))
                with col5:
                    if "Availability (%)" in inverter_data and pd.notna(inverter_data["Availability (%)"]):
                        availability = inverter_data["Availability (%)"]
                    else:
                        working, total = 0, 0
                        for col in pv_current_cols:
                            if col in inverter_data and pd.notna(inverter_data[col]):
                                total += 1
                                if inverter_data[col] > WORKING_CURRENT_THRESHOLD:
                                    working += 1
                        availability = (working / total * 100) if total > 0 else 0
                    st.metric("Availability", f"{availability:.1f}%")
                with col6:
                    if "Failed String Count" in inverter_data and pd.notna(inverter_data["Failed String Count"]):
                        st.metric("Failed Strings", int(inverter_data["Failed String Count"]))

                # Grid info
                if "Grid" in inverter_data:
                    st.info(f"**Grid:** {inverter_data['Grid']}")

                # ==========================================
                # INDIVIDUAL STRING STATUS - ORDERED BY PV NUMBER
                # ==========================================
                st.markdown('<h4><i class="fas fa-list-check"></i> PV String Status</h4>', unsafe_allow_html=True)
                st.caption(
                    f"Green = Working (>{WORKING_CURRENT_THRESHOLD}A) | Yellow = Low Performance "
                    f"({LOW_PERFORMANCE_DROP_PCT*100:.0f}% below average) | Red = Failed (<={WORKING_CURRENT_THRESHOLD}A) | "
                    f"Blinking Red = Negative Values (<= {NEGATIVE_VALUE_THRESHOLD}A)")

                # Sort PV current columns by number with proper parsing
                def get_pv_number(col):
                    try:
                        if 'PV-I' in col:
                            return int(col.replace('PV-I', ''))
                        elif 'PV' in col and 'I' not in col:
                            return int(col.replace('PV', ''))
                        elif '-' in col:
                            return int(col.split('-')[1])
                        else:
                            import re
                            numbers = re.findall(r'\d+', col)
                            if numbers:
                                return int(numbers[0])
                            return None
                    except (ValueError, IndexError, AttributeError):
                        return None

                pv_current_cols_sorted = []
                for col in pv_current_cols:
                    num = get_pv_number(col)
                    if num is not None:
                        pv_current_cols_sorted.append((col, num))

                pv_current_cols_sorted = sorted(pv_current_cols_sorted, key=lambda x: x[1])

                working_values = []
                all_string_data = {}

                for col, pv_num in pv_current_cols_sorted:
                    if col in inverter_data:
                        value = inverter_data[col]
                        if pd.notna(value):
                            status = "Working" if value > WORKING_CURRENT_THRESHOLD else "Failed"
                            is_negative = value <= NEGATIVE_VALUE_THRESHOLD
                            all_string_data[col] = {
                                "value": value,
                                "status": status,
                                "is_negative": is_negative,
                                "pv_num": pv_num
                            }
                            if status == "Working":
                                working_values.append(value)

                avg_working = sum(working_values) / len(working_values) if working_values else 0
                low_performance_threshold = avg_working * (1 - LOW_PERFORMANCE_DROP_PCT) if avg_working > 0 else 0

                string_data = {}
                for col, data in all_string_data.items():
                    value = data["value"]
                    status = data["status"]
                    is_negative = data["is_negative"]
                    pv_num = data["pv_num"]

                    performance = "normal"
                    if is_negative:
                        performance = "negative"
                    elif status == "Failed":
                        performance = "failed"
                    elif status == "Working" and avg_working > 0 and value < low_performance_threshold:
                        performance = "low"
                    else:
                        performance = "normal"

                    show = True
                    if string_filter == "Working Only" and status != "Working":
                        show = False
                    elif string_filter == "Failed Only" and status != "Failed":
                        show = False
                    elif string_filter == "Critical (<0.5A)" and not (status == "Failed" or value < 0.5):
                        show = False

                    if show:
                        string_data[col] = {
                            "value": value,
                            "status": status,
                            "is_negative": is_negative,
                            "performance": performance,
                            "pv_num": pv_num
                        }

                if string_data:
                    cols_per_row = 8
                    string_list = list(string_data.items())

                    st.markdown("""
                    <style>
                    @keyframes blink {
                        0% { opacity: 1; }
                        50% { opacity: 0.3; }
                        100% { opacity: 1; }
                    }
                    .blinking {
                        animation: blink 1s infinite;
                        border: 2px solid #ff0000 !important;
                    }
                    </style>
                    """, unsafe_allow_html=True)

                    for i in range(0, len(string_list), cols_per_row):
                        string_cols = st.columns(cols_per_row)
                        for idx, (col, data) in enumerate(string_list[i:i + cols_per_row]):
                            value = data["value"]
                            status = data["status"]
                            performance = data["performance"]
                            is_negative = data["is_negative"]
                            pv_num = data["pv_num"]

                            if is_negative:
                                color = "#ef4444"
                                extra_class = "blinking"
                                status_text = '<i class="fas fa-triangle-exclamation"></i> NEGATIVE!'
                            elif performance == "failed":
                                color = "#ef4444"
                                extra_class = ""
                                status_text = "Failed"
                            elif performance == "low":
                                color = "#fbbf24"
                                extra_class = ""
                                status_text = '<i class="fas fa-triangle-exclamation"></i> Low Perf'
                            else:
                                color = "#10b981"
                                extra_class = ""
                                status_text = "Working"

                            pv_display = f"PV-{pv_num}"

                            with string_cols[idx]:
                                st.markdown(f"""
                                <div class='{extra_class}' style='background-color: {color}; padding: 8px; border-radius: 5px; text-align: center; color: white; margin: 2px;'>
                                    <div style='font-size: 10px; font-weight: bold;'>{pv_display}</div>
                                    <div style='font-size: 14px; font-weight: bold;'>{value:.1f}A</div>
                                    <div style='font-size: 9px;'>{status_text}</div>
                                </div>
                                """, unsafe_allow_html=True)

                    total_filtered = len(string_data)
                    working_filtered = sum(1 for d in string_data.values() if d["status"] == "Working")
                    failed_filtered = sum(1 for d in string_data.values() if d["performance"] == "failed")
                    low_performance = sum(1 for d in string_data.values() if d["performance"] == "low")
                    negative_count = sum(1 for d in string_data.values() if d["is_negative"])

                    st.markdown("---")
                    col1, col2, col3, col4, col5 = st.columns(5)
                    col1.metric("Total Strings", total_filtered)
                    col2.metric("Working", working_filtered)
                    col3.metric("Failed", failed_filtered)
                    col4.metric("Low Performance", low_performance)
                    col5.metric("Negative", negative_count)

                    if working_values:
                        st.info(
                            f"Average Working Current: **{avg_working:.2f}A** | Low Performance Threshold: **{low_performance_threshold:.2f}A** ({LOW_PERFORMANCE_DROP_PCT*100:.0f}% below average)")

                    st.markdown('<h4><i class="fas fa-list"></i> Performance Summary</h4>', unsafe_allow_html=True)

                    perf_summary = []
                    for col, data in string_data.items():
                        perf_summary.append({
                            "String": col,
                            "PV Number": data["pv_num"],
                            "Current (A)": round(data["value"], 2),
                            "Status": data["status"],
                            "Performance": "Normal" if data["performance"] == "normal" else
                            "Low" if data["performance"] == "low" else
                            "Negative" if data["is_negative"] else "Failed"
                        })

                    if perf_summary:
                        perf_df = pd.DataFrame(perf_summary)

                        def color_performance(val):
                            if val == "Normal":
                                return 'background-color: #10b981; color: white; font-weight: bold'
                            elif val == "Low":
                                return 'background-color: #fbbf24; color: black; font-weight: bold'
                            elif val == "Negative":
                                return 'background-color: #ef4444; color: white; font-weight: bold; animation: blink 1s infinite;'
                            elif val == "Failed":
                                return 'background-color: #dc2626; color: white; font-weight: bold'
                            return ''

                        styled_perf = perf_df.style.map(color_performance, subset=['Performance'])
                        st.dataframe(styled_perf, use_container_width=True)

                    if low_performance > 0 or negative_count > 0:
                        st.markdown("---")
                        st.markdown('<h3><i class="fas fa-triangle-exclamation"></i> Performance Alerts</h3>', unsafe_allow_html=True)

                        alert_cols = st.columns(2)

                        with alert_cols[0]:
                            if low_performance > 0:
                                st.warning(
                                    f"**{low_performance}** string(s) are performing below the threshold ({LOW_PERFORMANCE_DROP_PCT*100:.0f}% below average)")
                                low_strings = [f"PV-{data['pv_num']} ({data['value']:.1f}A)" for col, data in
                                               string_data.items() if data["performance"] == "low"]
                                if low_strings:
                                    st.write("**Low Performance Strings:**")
                                    st.write(", ".join(low_strings))

                        with alert_cols[1]:
                            if negative_count > 0:
                                st.error(
                                    f"**{negative_count}** string(s) have NEGATIVE values! Immediate attention required!")
                                neg_strings = [f"PV-{data['pv_num']} ({data['value']:.1f}A)" for col, data in
                                               string_data.items() if data["is_negative"]]
                                if neg_strings:
                                    st.write("**Negative Value Strings:**")
                                    st.write(", ".join(neg_strings))
                else:
                    st.info("No strings match the current filter criteria")

    # ==========================================
    # TAB 3: FAILED INVERTERS ONLY  (new module)
    # Shows just the inverters with at least one failed string, with their
    # full PV string breakdown, so a field engineer can jump straight to
    # what needs attention without scanning the whole dataset.
    # ==========================================
    with tab3:
        st.markdown('<h3><i class="fas fa-triangle-exclamation" style="color:#ef4444;"></i> Failed Inverters Only</h3>', unsafe_allow_html=True)
        st.caption("Inverters with one or more failed PV strings, scoped to the filters above")

        failed_summary_df = summary_df[summary_df["Failed Strings"] > 0].copy()

        if failed_summary_df.empty:
            st.success("No failed inverters for the current filter selection.")
        else:
            failed_summary_df = failed_summary_df.sort_values("Failed Strings", ascending=False)

            fcol1, fcol2, fcol3 = st.columns(3)
            fcol1.metric("Failed Inverters", len(failed_summary_df))
            fcol2.metric("Total Failed Strings", int(failed_summary_df["Failed Strings"].sum()))
            fcol3.metric("Avg Availability", f"{failed_summary_df['Availability (%)'].mean():.1f}%")

            st.markdown("---")
            st.markdown('<h4><i class="fas fa-table-list"></i> Failed Inverter Summary</h4>', unsafe_allow_html=True)
            styled_failed_summary = get_styled_summary(failed_summary_df)
            st.dataframe(styled_failed_summary, use_container_width=True)

            st.markdown("---")
            st.markdown('<h4><i class="fas fa-bolt"></i> PV String Details for Failed Inverters</h4>', unsafe_allow_html=True)
            st.caption("Only the failed inverters' individual PV-string currents are shown below")

            failed_inverter_ids = failed_summary_df["Inverter ID"].tolist()
            failed_detail_df = filtered_df[filtered_df[inverter_col].isin(failed_inverter_ids)].copy()

            detail_display_cols = [inverter_col, "Plot", "Block", "SACU", "Grid"]
            detail_display_cols = [c for c in detail_display_cols if c == inverter_col or c in failed_detail_df.columns]
            for col in ["Total Active Strings", "Working String Count", "Failed String Count",
                        "Availability (%)", "Failure Percentage (%)"]:
                if col in failed_detail_df.columns:
                    detail_display_cols.append(col)
            detail_display_cols.extend([c for c in sorted(pv_current_cols) if c in failed_detail_df.columns])

            failed_detail_display = failed_detail_df[detail_display_cols].rename(columns={inverter_col: "Inverter ID"})
            failed_detail_display = failed_detail_display.rename(columns={
                "Total Active Strings": "Total Strings", "Working String Count": "Working",
                "Failed String Count": "Failed", "Failure Percentage (%)": "Failure %",
            })

            def highlight_failed_current(x):
                if pd.notna(x) and isinstance(x, (int, float)):
                    return f'background-color: {get_string_health_color(x)}; color: white; font-weight: bold;'
                return ''

            try:
                styled_failed_detail = failed_detail_display.style
                for col in pv_current_cols:
                    if col in failed_detail_display.columns:
                        styled_failed_detail = styled_failed_detail.map(highlight_failed_current, subset=[col])
                if "Availability (%)" in failed_detail_display.columns:
                    styled_failed_detail = styled_failed_detail.map(color_availability, subset=['Availability (%)'])
                if "Failed" in failed_detail_display.columns:
                    styled_failed_detail = styled_failed_detail.map(color_failed_strings, subset=['Failed'])
                st.dataframe(styled_failed_detail, use_container_width=True, height=400)
            except Exception as e:
                st.warning(f"Styling error: {str(e)}. Showing unstyled data.")
                st.dataframe(failed_detail_display, use_container_width=True, height=400)

            fdl_col1, fdl_col2 = st.columns(2)
            with fdl_col1:
                csv_failed = failed_detail_display.to_csv(index=False)
                st.download_button(
                    label="Download Failed Inverters (CSV)",
                    data=csv_failed,
                    file_name=f"failed_inverters_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    key="download_failed_inverters_csv",
                    on_click=_log_download, args=("failed_inverters_csv",),
                    use_container_width=True,
                )
            with fdl_col2:
                excel_failed = create_colored_excel_download({"Failed Inverters": failed_detail_display})
                st.download_button(
                    label="Download Failed Inverters (Excel, color-coded)",
                    data=excel_failed,
                    file_name=f"failed_inverters_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_failed_inverters_xlsx",
                    on_click=_log_download, args=("failed_inverters_xlsx",),
                    use_container_width=True,
                )

    # ==========================================
    # TAB 4: LOW PERFORMANCE STRINGS (yellow) - across every inverter in
    # the current filter selection, not just one inverter at a time.
    # ==========================================
    with tab4:
        st.markdown('<h3><i class="fas fa-triangle-exclamation" style="color:#fbbf24;"></i> Low Performance Strings</h3>', unsafe_allow_html=True)
        st.caption(
            f"Strings that are working but sitting more than {LOW_PERFORMANCE_DROP_PCT*100:.0f}% below their own inverter's average "
            "working current are flagged yellow, scoped to the filters above."
        )

        low_perf_df = find_low_performance_strings_cached(filtered_df, inverter_col, pv_current_cols)

        if low_perf_df.empty:
            st.info(f"No low performance PV strings found based on the defined {LOW_PERFORMANCE_DROP_PCT*100:.0f}% below average threshold.")
        else:
            lp_col1, lp_col2, lp_col3 = st.columns(3)
            lp_col1.metric("Low Performance Strings", len(low_perf_df))
            lp_col2.metric("Inverters Affected", low_perf_df["Inverter Name"].nunique())
            lp_col3.metric("Avg Current (A)", f"{low_perf_df['PV Value'].mean():.2f}")

            def color_low_perf(val):
                return 'background-color: #fbbf24; color: black; font-weight: bold;'

            threshold_col = f"Low Performance Threshold ({LOW_PERFORMANCE_DROP_PCT*100:.0f}% below avg)"
            styled_low_perf = low_perf_df.style.map(color_low_perf, subset=["PV Value", threshold_col])
            st.dataframe(styled_low_perf, use_container_width=True, height=420)

            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    label="Download Low Performance Strings (CSV)",
                    data=low_perf_df.to_csv(index=False),
                    file_name=f"low_performance_strings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    key="download_low_perf_csv",
                    on_click=_log_download, args=("low_performance_strings_csv",),
                    use_container_width=True,
                )
            with dl_col2:
                st.download_button(
                    label="Download Low Performance Strings (Excel, color-coded)",
                    data=create_colored_excel_download({"Low Performance": low_perf_df}),
                    file_name=f"low_performance_strings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_low_perf_xlsx",
                    on_click=_log_download, args=("low_performance_strings_xlsx",),
                    use_container_width=True,
                )

def get_styled_summary(summary_df):
    """Apply styling to summary dataframe without caching"""
    styled = summary_df.style.map(color_availability, subset=['Availability (%)'])
    styled = styled.map(color_health_status, subset=['Health Status'])
    styled = styled.map(color_failed_strings, subset=['Failed Strings'])
    styled = styled.format({
        'Availability (%)': '{:.1f}%', 'Avg PV Voltage (V)': '{:.1f}', 'Avg PV Current (A)': '{:.2f}'
    })
    return styled

# ==========================================
# DASHBOARD FUNCTIONS WITH OPTIMIZED CACHING
# ==========================================
@st.cache_data(ttl=300)
def calculate_plot_summary_cached(df, inverter_col):
    """Cache plot summary calculations"""
    if df.empty:
        return pd.DataFrame()

    # Match the PV String Details summary and Data Table: each processed row is
    # one physical inverter, even when the displayed inverter ID repeats.
    inverter_counts = df.groupby("Plot").size().reset_index(name="Total_Inverters")

    plot_summary = df.groupby("Plot", as_index=False).agg(
        Total_Active_Strings=("Total Active Strings", "sum"),
        Total_Working_Strings=("Working String Count", "sum"),
        Total_Failed_Strings=("Failed String Count", "sum")
    )
    plot_summary = plot_summary.merge(inverter_counts, on="Plot", how="left")
    plot_summary["Total_Inverters"] = plot_summary["Total_Inverters"].fillna(0).astype(int)
    plot_summary["Availability (%)"] = ((plot_summary["Total_Working_Strings"] / plot_summary["Total_Active_Strings"]) * 100).fillna(0).round(2)
    plot_summary["Failure Percentage (%)"] = ((plot_summary["Total_Failed_Strings"] / plot_summary["Total_Active_Strings"]) * 100).fillna(0).round(2)
    plot_summary["Health Status"] = plot_summary["Availability (%)"].apply(
        lambda x: "Excellent" if x >= 90 else "Good" if x >= 70 else "Fair" if x >= 50 else "Poor"
    )
    block_count = df.groupby("Plot")["Block"].nunique().reset_index(name="Total_Blocks")
    plot_summary = plot_summary.merge(block_count, on="Plot", how="left")
    return plot_summary

@st.cache_data(ttl=300)
def calculate_block_summary_cached(df, inverter_col):
    """
    Cache block-wise summary calculations (grouped by Plot + Block).
    Inverter counts use nunique() on the Inverter ID column so duplicate
    rows for the same physical inverter are only ever counted once.
    """
    if df is None or df.empty or "Block" not in df.columns:
        return pd.DataFrame()

    group_cols = ["Plot", "Block"] if "Plot" in df.columns else ["Block"]

    if inverter_col and inverter_col in df.columns:
        inverter_counts = df.groupby(group_cols)[inverter_col].nunique().reset_index(name="Total_Inverters")
    else:
        inverter_counts = df.groupby(group_cols).size().reset_index(name="Total_Inverters")

    block_summary = df.groupby(group_cols, as_index=False).agg(
        Total_Active_Strings=("Total Active Strings", "sum"),
        Total_Working_Strings=("Working String Count", "sum"),
        Total_Failed_Strings=("Failed String Count", "sum"),
    )
    block_summary = block_summary.merge(inverter_counts, on=group_cols, how="left")
    block_summary["Total_Inverters"] = block_summary["Total_Inverters"].fillna(0).astype(int)
    block_summary["Availability (%)"] = ((block_summary["Total_Working_Strings"] / block_summary["Total_Active_Strings"]) * 100).fillna(0).round(2)
    block_summary["Failure Percentage (%)"] = ((block_summary["Total_Failed_Strings"] / block_summary["Total_Active_Strings"]) * 100).fillna(0).round(2)
    block_summary["Health Status"] = block_summary["Availability (%)"].apply(
        lambda x: "Excellent" if x >= 90 else "Good" if x >= 70 else "Fair" if x >= 50 else "Poor"
    )
    if "Plot" in block_summary.columns:
        block_summary["Block Label"] = block_summary["Plot"].astype(str) + " · " + block_summary["Block"].astype(str)
    else:
        block_summary["Block Label"] = block_summary["Block"].astype(str)

    block_summary = block_summary.sort_values("Block Label").reset_index(drop=True)
    return block_summary


def display_block_metrics(block_summary, plot_context="All Plots"):
    """Render Block-wise Performance Overview KPI cards (purple/indigo theme,
    to visually pair with - but stay distinct from - the Plot-wise cards)."""
    st.markdown(
        f'<i class="fas fa-layer-group"></i> Block-wise Performance Overview '
        f'<span style="color:#94a3b8; font-size:0.85rem; font-weight:400;">— {plot_context}</span>',
        unsafe_allow_html=True,
    )

    records = block_summary.to_dict("records")
    cards_per_row = 4
    for row_start in range(0, len(records), cards_per_row):
        row_records = records[row_start:row_start + cards_per_row]
        cols = st.columns(len(row_records))
        for col, row in zip(cols, row_records):
            avail = row["Availability (%)"]
            if avail >= 90:
                status_color, status_icon, status_text = "#10b981", '<i class="fas fa-circle-check"></i>', "Excellent"
            elif avail >= 70:
                status_color, status_icon, status_text = "#34d399", '<i class="fas fa-circle-check"></i>', "Good"
            elif avail >= 50:
                status_color, status_icon, status_text = "#fbbf24", '<i class="fas fa-circle-exclamation"></i>', "Fair"
            else:
                status_color, status_icon, status_text = "#ef4444", '<i class="fas fa-circle-xmark"></i>', "Poor"

            with col:
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #1e1b4b 0%, #0f172a 100%); border: 2px solid {status_color}; border-radius: 12px; padding: 14px; margin: 5px 0;'>
                    <div style='display: flex; justify-content: space-between; align-items: center;'>
                        <h4 style='margin: 0; color: #f1f5f9; font-size: 14px;'><i class="fas fa-layer-group" style="color:#a78bfa; margin-right:6px;"></i>{row['Block Label']}</h4>
                        <span style='font-size: 18px; color: {status_color};'>{status_icon}</span>
                    </div>
                    <div style='margin-top: 8px;'>
                        <div style='display: flex; justify-content: space-between;'>
                            <span style='color: #94a3b8; font-size: 11px;'>Status</span>
                            <span style='color: {status_color}; font-weight: bold; font-size: 13px;'>{status_text}</span>
                        </div>
                        <div style='display: flex; justify-content: space-between; margin-top: 4px;'>
                            <span style='color: #94a3b8; font-size: 11px;'>Inverters (unique)</span>
                            <span style='color: #f1f5f9; font-weight: bold;'>{int(row['Total_Inverters']):,}</span>
                        </div>
                        <div style='display: flex; justify-content: space-between;'>
                            <span style='color: #94a3b8; font-size: 11px;'>Total Strings</span>
                            <span style='color: #f1f5f9; font-weight: bold;'>{int(row['Total_Active_Strings']):,}</span>
                        </div>
                        <div style='display: flex; justify-content: space-between;'>
                            <span style='color: #94a3b8; font-size: 11px;'><i class="fas fa-circle-check" style="color:#10b981;"></i> Working</span>
                            <span style='color: #10b981; font-weight: bold;'>{int(row['Total_Working_Strings']):,}</span>
                        </div>
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px;'>
                            <span style='color: #94a3b8; font-size: 11px;'><i class="fas fa-circle-xmark" style="color:#ef4444;"></i> Failed</span>
                            <span style='color: #ef4444; font-weight: bold;'>{int(row['Total_Failed_Strings']):,}</span>
                        </div>
                        <div style='background-color: #1e293b; height: 7px; border-radius: 4px; overflow: hidden;'>
                            <div style='background: linear-gradient(90deg, {status_color}, {status_color}88); width: {avail}%; height: 100%;'></div>
                        </div>
                        <div style='display: flex; justify-content: space-between; margin-top: 4px;'>
                            <span style='color: #94a3b8; font-size: 10px;'>Availability</span>
                            <span style='color: {status_color}; font-weight: bold; font-size: 14px;'>{avail:.1f}%</span>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)


@st.cache_data(ttl=300)
def create_plot_charts_cached(plot_summary):
    """Cache chart generation"""
    charts = {}

    fig_bar = px.bar(
        plot_summary, x="Plot", y=["Total_Working_Strings", "Total_Failed_Strings"], barmode="stack",
        title="Plot-wise String Status",
        labels={"value": "Number of Strings", "Plot": "Plot", "variable": "Status"},
        color_discrete_map={"Total_Working_Strings": "#10b981", "Total_Failed_Strings": "#ef4444"}, text_auto=True
    )
    fig_bar.update_layout(height=450, hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font=dict(size=12))
    fig_bar.update_traces(textfont_size=12, textposition="inside", insidetextanchor="middle")
    fig_bar.update_yaxes(tickformat=",.0f")
    charts["bar"] = fig_bar

    total_working = plot_summary["Total_Working_Strings"].sum()
    total_failed = plot_summary["Total_Failed_Strings"].sum()
    fig_donut = go.Figure(data=[go.Pie(
        labels=["Working Strings", "Failed Strings"], values=[total_working, total_failed], hole=0.6,
        marker_colors=["#10b981", "#ef4444"], textinfo="label+percent", textposition="auto", pull=[0.05, 0]
    )])
    fig_donut.update_layout(height=400, title="Overall String Health",
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font=dict(size=12),
        annotations=[dict(text=f"<b>{total_working + total_failed:,}</b><br>Total Strings", x=0.5, y=0.5, font_size=16, showarrow=False)])
    charts["donut"] = fig_donut

    fig_treemap = px.treemap(
        plot_summary, path=["Plot"], values="Total_Active_Strings", color="Availability (%)",
        color_continuous_scale=[[0, "#ef4444"], [0.3, "#f59e0b"], [0.5, "#fbbf24"], [0.7, "#34d399"], [1, "#10b981"]],
        range_color=[0, 100], title="String Distribution by Plot",
        hover_data={"Total_Active_Strings": True, "Total_Working_Strings": True, "Total_Failed_Strings": True,
                    "Total_Inverters": True, "Total_Blocks": True, "Availability (%)": ":.1f%"}
    )
    fig_treemap.update_traces(textinfo="label+value", textfont_size=14, marker=dict(cornerradius=4),
        hovertemplate='<b>%{label}</b><br>Total Strings: %{value:,.0f}<br>Availability: %{color:,.1f}%<extra></extra>')
    fig_treemap.update_layout(height=400, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(size=12), coloraxis_showscale=False)
    charts["treemap"] = fig_treemap

    return charts

def display_plot_metrics(plot_summary):
    st.markdown('<i class="fas fa-chart-pie"></i> Plot-wise Performance Overview', unsafe_allow_html=True)
    cols = st.columns(min(6, len(plot_summary)))

    for idx, (_, row) in enumerate(plot_summary.iterrows()):
        if idx >= 6:
            break
        col_idx = idx % 6
        with cols[col_idx]:
            avail = row["Availability (%)"]
            if avail >= 90:
                status_color, status_icon, status_text = "#10b981", '<i class="fas fa-circle-check"></i>', "Excellent"
            elif avail >= 70:
                status_color, status_icon, status_text = "#34d399", '<i class="fas fa-circle-check"></i>', "Good"
            elif avail >= 50:
                status_color, status_icon, status_text = "#fbbf24", '<i class="fas fa-circle-exclamation"></i>', "Fair"
            else:
                status_color, status_icon, status_text = "#ef4444", '<i class="fas fa-circle-xmark"></i>', "Poor"

            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%); border: 2px solid {status_color}; border-radius: 12px; padding: 15px; margin: 5px 0;'>
                <div style='display: flex; justify-content: space-between; align-items: center;'>
                    <h3 style='margin: 0; color: #f1f5f9;'>{row['Plot']}</h3>
                    <span style='font-size: 22px; color: {status_color};'>{status_icon}</span>
                </div>
                <div style='margin-top: 8px;'>
                    <div style='display: flex; justify-content: space-between;'>
                        <span style='color: #94a3b8; font-size: 12px;'>Status</span>
                        <span style='color: {status_color}; font-weight: bold; font-size: 14px;'>{status_text}</span>
                    </div>
                    <div style='display: flex; justify-content: space-between; margin-top: 4px;'>
                        <span style='color: #94a3b8; font-size: 12px;'>Inverters</span>
                        <span style='color: #f1f5f9; font-weight: bold;'>{int(row['Total_Inverters']):,}</span>
                    </div>
                    <div style='display: flex; justify-content: space-between;'>
                        <span style='color: #94a3b8; font-size: 12px;'>Total Strings</span>
                        <span style='color: #f1f5f9; font-weight: bold;'>{int(row['Total_Active_Strings']):,}</span>
                    </div>
                    <div style='display: flex; justify-content: space-between;'>
                        <span style='color: #94a3b8; font-size: 12px;'><i class="fas fa-circle-check" style="color: #10b981;"></i> Working</span>
                        <span style='color: #10b981; font-weight: bold;'>{int(row['Total_Working_Strings']):,}</span>
                    </div>
                    <div style='display: flex; justify-content: space-between; margin-bottom: 8px;'>
                        <span style='color: #94a3b8; font-size: 12px;'><i class="fas fa-circle-xmark" style="color: #ef4444;"></i> Failed</span>
                        <span style='color: #ef4444; font-weight: bold;'>{int(row['Total_Failed_Strings']):,}</span>
                    </div>
                    <div style='background-color: #1e293b; height: 8px; border-radius: 4px; overflow: hidden;'>
                        <div style='background: linear-gradient(90deg, {status_color}, {status_color}88); width: {avail}%; height: 100%;'></div>
                    </div>
                    <div style='display: flex; justify-content: space-between; margin-top: 4px;'>
                        <span style='color: #94a3b8; font-size: 11px;'>Availability</span>
                        <span style='color: {status_color}; font-weight: bold; font-size: 16px;'>{avail:.1f}%</span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

def render_kpi_cards(metrics):
    """Render a row of attractive gradient KPI cards. `metrics` is a list of
    dicts: {label, value, icon (Font Awesome class), color (hex), sub (optional)}."""
    cols = st.columns(len(metrics))
    for col, m in zip(cols, metrics):
        color = m["color"]
        sub_html = (
            f"<div style='color:#64748b; font-size:11px; margin-top:2px;'>{m['sub']}</div>"
            if m.get("sub") else ""
        )
        with col:
            st.markdown(f"""
            <div style='background: linear-gradient(150deg, #0f172a 0%, #111827 100%);
                        border: 1px solid {color}55; border-left: 4px solid {color};
                        border-radius: 12px; padding: 14px 16px; margin: 4px 0;
                        box-shadow: 0 8px 20px rgba(2,6,23,0.25);'>
                <div style='display:flex; align-items:center; justify-content:space-between;'>
                    <span style='color:#94a3b8; font-size:12px; font-weight:600; letter-spacing:.02em;'>{m['label']}</span>
                    <span style='width:30px; height:30px; border-radius:9px; background:{color}22;
                                 display:flex; align-items:center; justify-content:center; color:{color}; font-size:14px;'>
                        <i class="{m['icon']}"></i>
                    </span>
                </div>
                <div style='color:#f8fafc; font-size:1.55rem; font-weight:800; margin-top:6px; line-height:1;'>{m['value']}</div>
                {sub_html}
            </div>
            """, unsafe_allow_html=True)


def main_dashboard_tab(df, sheet_df=None, sheet_name="Sheet1", snapshot_date=None):
    st.markdown('<h1><i class="fas fa-sun" style="color:#fbbf24;"></i> Solar PV String Performance Dashboard</h1>', unsafe_allow_html=True)
    st.caption("Internal beta release - data processing, history, and comparison features are under testing.")

    # ---- Persistent duplicate-Inverter-ID banner ----
    # Surfaces here (not just as a one-time upload toast) so it stays
    # visible for anyone viewing this snapshot's dashboard later.
    duplicate_warnings = get_duplicate_inverter_warnings_for_snapshot(snapshot_date)
    if duplicate_warnings:
        dup_lines = []
        total_dup_count = 0
        for sh_name, dup_ids in duplicate_warnings.items():
            total_dup_count += len(dup_ids)
            dup_lines.append(
                f"<div style='margin-top:4px;'><b>{sh_name}:</b> "
                f"<span style='color:#fecaca;'>{', '.join(dup_ids)}</span></div>"
            )
        st.markdown(f"""
        <div style='background: linear-gradient(120deg, rgba(127,29,29,0.35) 0%, rgba(15,23,42,0.9) 100%);
                    border: 1px solid rgba(248,113,113,0.4); border-left: 4px solid #ef4444;
                    border-radius: 12px; padding: 12px 16px; margin-bottom: 14px;'>
            <div style='display:flex; align-items:center; gap:8px;'>
                <i class="fas fa-triangle-exclamation" style="color:#f87171; font-size:16px;"></i>
                <span style='color:#fecaca; font-weight:700; font-size:0.95rem;'>
                    Duplicate Inverter IDs detected in this snapshot ({total_dup_count} total)
                </span>
            </div>
            <div style='color:#e2e8f0; font-size:0.82rem; margin-top:6px;'>
                These were de-duplicated (first occurrence kept) when calculating the metrics below.
                {''.join(dup_lines)}
            </div>
        </div>
        """, unsafe_allow_html=True)

    inverter_col = get_inverter_column_cached(df)

    plot_summary = calculate_plot_summary_cached(df, inverter_col)

    st.markdown("### <i class='fas fa-chart-line'></i> Key Performance Indicators", unsafe_allow_html=True)

    # Inverter count reflects UNIQUE Inverter IDs only, consistent with the
    # de-duplication applied at upload time.
    total_inverters = int(df[inverter_col].nunique()) if inverter_col and inverter_col in df.columns else int(len(df))
    total_strings = int(df["Total Active Strings"].sum()) if "Total Active Strings" in df.columns else 0
    working_strings = int(df["Working String Count"].sum()) if "Working String Count" in df.columns else 0
    failed_strings = int(df["Failed String Count"].sum()) if "Failed String Count" in df.columns else 0
    overall_availability = round((working_strings / total_strings) * 100, 2) if total_strings > 0 else 0.0
    num_plots = plot_summary["Plot"].nunique() if not plot_summary.empty else 0
    avail_color = "#10b981" if overall_availability >= 90 else "#34d399" if overall_availability >= 70 else "#fbbf24" if overall_availability >= 50 else "#ef4444"

    render_kpi_cards([
        {"label": "TOTAL PLOTS", "value": f"{num_plots:,}", "icon": "fas fa-map-location-dot", "color": "#38bdf8"},
        {"label": "TOTAL INVERTERS", "value": f"{total_inverters:,}", "icon": "fas fa-microchip", "color": "#818cf8", "sub": "Unique Inverter IDs"},
        {"label": "TOTAL STRINGS", "value": f"{total_strings:,}", "icon": "fas fa-plug", "color": "#a78bfa"},
        {"label": "WORKING", "value": f"{working_strings:,}", "icon": "fas fa-circle-check", "color": "#10b981"},
        {"label": "FAILED", "value": f"{failed_strings:,}", "icon": "fas fa-circle-xmark", "color": "#ef4444"},
        {"label": "AVAILABILITY", "value": f"{overall_availability:.1f}%", "icon": "fas fa-gauge-high", "color": avail_color},
    ])

    st.markdown("---")
    if not plot_summary.empty:
        display_plot_metrics(plot_summary)
        st.markdown("---")

    st.subheader("Plot-wise Visualization Dashboard")
    st.caption("Understanding your PV plant performance at a glance")
    charts = create_plot_charts_cached(plot_summary)

    col1, col2 = st.columns([2, 1])
    with col1:
        st.plotly_chart(charts["bar"], use_container_width=True, key="plot_bar")
    with col2:
        st.plotly_chart(charts["donut"], use_container_width=True, key="overall_donut")

    st.plotly_chart(charts["treemap"], use_container_width=True, key="plot_treemap")
    st.markdown("---")
    st.markdown('<i class="fas fa-table"></i> Detailed Plot Summary', unsafe_allow_html=True)

    if not plot_summary.empty:
        display_plot_df = plot_summary.copy()
        display_plot_df = display_plot_df[[
            "Plot", "Total_Blocks", "Total_Inverters", "Total_Active_Strings",
            "Total_Working_Strings", "Total_Failed_Strings", "Availability (%)",
            "Failure Percentage (%)", "Health Status"
        ]]

        def color_health_status(val):
            if "Excellent" in str(val): return 'background-color: #10b981; color: white; font-weight: bold; border-radius: 4px;'
            elif "Good" in str(val): return 'background-color: #34d399; color: white; font-weight: bold; border-radius: 4px;'
            elif "Fair" in str(val): return 'background-color: #fbbf24; color: black; font-weight: bold; border-radius: 4px;'
            elif "Poor" in str(val): return 'background-color: #ef4444; color: white; font-weight: bold; border-radius: 4px;'
            return ''

        styled_plot_df = display_plot_df.style.map(color_health_status, subset=['Health Status'])
        styled_plot_df = styled_plot_df.format({
            'Total_Active_Strings': '{:,.0f}', 'Total_Working_Strings': '{:,.0f}',
            'Total_Failed_Strings': '{:,.0f}', 'Total_Inverters': '{:,.0f}',
            'Total_Blocks': '{:,.0f}', 'Availability (%)': '{:.1f}%', 'Failure Percentage (%)': '{:.1f}%'
        })

        def availability_bar(val):
            if isinstance(val, (int, float)):
                color = "#10b981" if val >= 90 else "#34d399" if val >= 70 else "#fbbf24" if val >= 50 else "#ef4444"
                return f'background: linear-gradient(90deg, {color} {val}%, transparent {val}%); font-weight: bold; padding: 4px 8px; border-radius: 4px;'
            return ''

        styled_plot_df = styled_plot_df.map(availability_bar, subset=['Availability (%)'])

        st.dataframe(styled_plot_df, use_container_width=True, column_config={
            "Plot": "Plot",
            "Total_Blocks": "Blocks",
            "Total_Inverters": "Inverters",
            "Total_Active_Strings": "Total Strings",
            "Total_Working_Strings": "Working",
            "Total_Failed_Strings": "Failed",
            "Availability (%)": "Availability",
            "Failure Percentage (%)": "Failure %",
            "Health Status": "Health"
        })

        col1, col2 = st.columns(2)
        with col1:
            csv = plot_summary.to_csv(index=False)
            st.download_button(
                label='Download Plot Summary (CSV)', data=csv,
                file_name=f"plot_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv", use_container_width=True,
                on_click=_log_download, args=("plot_summary_csv",),
            )
        with col2:
            best_plot = plot_summary.loc[plot_summary["Availability (%)"].idxmax()]
            worst_plot = plot_summary.loc[plot_summary["Availability (%)"].idxmin()]
            st.info(f"""
            **Insights:**
            - Best performing plot: **{best_plot['Plot']}** ({best_plot['Availability (%)']:.1f}% availability)
            - Needs attention: **{worst_plot['Plot']}** ({worst_plot['Availability (%)']:.1f}% availability)
            - Total working strings: **{working_strings:,}** out of **{total_strings:,}**
            """)
    else:
        st.warning("No plot summary available.")

    display_negative_pv_values(sheet_df if sheet_df is not None else df, sheet_name=sheet_name)

# ==========================================
# 8. AUDIT LOGGING HELPERS FOR DOWNLOADS
# ==========================================
def _log_download(report_name):
    """on_click callback for download buttons - records who downloaded what and when."""
    user = get_current_user()
    if not user:
        return
    storage1.log_audit_event(user["username"], user["role"], "download_report", {"report": report_name})

# ==========================================
# 8b. HEADER CALENDAR - VIEW ANY PREPROCESSED SNAPSHOT DATE
# ==========================================
def render_header_calendar():
    """
    Renders a date-picker in the dashboard header letting any user browse
    the already-preprocessed SCADA data for any past snapshot date, without
    needing a fresh upload. Returns (selected_date, is_latest).
    """
    available_dates = storage1.get_available_snapshot_dates()
    latest_upload = storage1.get_latest_upload()

    if not available_dates or not latest_upload:
        return None, True

    default_date = latest_upload["snapshot_date"]
    date_options = list(reversed(available_dates))  # newest first
    default_index = date_options.index(default_date) if default_date in date_options else 0

    st.markdown(
        f"""
        <div class="calendar-banner">
            <div class="calendar-icon-wrap"><i class="fas fa-calendar-days"></i></div>
            <div>
                <strong>Snapshot Date</strong>
                <div class="calendar-subtitle">Browse any previously preprocessed snapshot without re-uploading.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    cal_col1, cal_col2 = st.columns([1, 3])
    with cal_col1:
        selected_date = st.selectbox(
            "Snapshot Date",
            options=date_options,
            index=default_index,
            key="header_snapshot_date",
            label_visibility="collapsed",
        )
    with cal_col2:
        if selected_date == default_date:
            st.caption(f"Viewing the latest uploaded snapshot ({selected_date}).")
        else:
            st.caption(f"Viewing a previously preprocessed snapshot from **{selected_date}**. Switch back to {default_date} for the latest data.")

    return selected_date, (selected_date == default_date)

# ==========================================
# 9. MAIN APP WITH OPTIMIZED STATE MANAGEMENT
# ==========================================
def main():
    storage1.init_default_users()
    storage1.cleanup_expired_sessions()

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    # ---- Restore session across a page refresh -----------------------
    # st.session_state is tied to the browser websocket connection and is
    # wiped out on a hard refresh. A session token is kept in the URL
    # (st.query_params) so a refresh doesn't force a fresh login; the token
    # itself is only valid until the end of the calendar day it was issued.
    if not st.session_state.authenticated:
        sid = st.query_params.get("sid")
        if sid:
            remembered_username = storage1.validate_session(sid)
            if remembered_username:
                users = storage1.load_users()
                user_data = users.get(remembered_username)
                if user_data:
                    st.session_state.user = {
                        "username": remembered_username,
                        "role": user_data["role"],
                        "full_name": user_data.get("full_name", remembered_username),
                        "assigned_plots": user_data.get("assigned_plots", []),
                    }
                    st.session_state.authenticated = True
                    st.session_state.session_token = sid
            else:
                # Stale/expired token still sitting in the URL - drop it.
                try:
                    del st.query_params["sid"]
                except Exception:
                    pass

    # ---------------- LOGIN ----------------
    if not st.session_state.authenticated:
        st.markdown(
            '<h1 style="text-align:center;"><i class="fas fa-sun" style="color:#fbbf24;"></i> Solar PV String Analytics</h1>',
            unsafe_allow_html=True,
        )

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown(
                '<div class="login-card">'
                '<div class="login-title"><i class="fas fa-right-to-bracket"></i> Welcome back</div>'
                '<div class="login-subtitle">Login to access the dashboard</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            with st.form(key="login_form", border=False):
                username = st.text_input("Username", key="login_username")
                password = st.text_input("Password", type="password", key="login_password")
                login_clicked = st.form_submit_button("🔐 Login", use_container_width=True, type="primary")

            if login_clicked:
                user_data = storage1.authenticate_user(username, password)
                if user_data:
                    st.session_state.user = {
                        "username": username,
                        "role": user_data["role"],
                        "full_name": user_data.get("full_name", username),
                        "assigned_plots": user_data.get("assigned_plots", []),
                    }
                    st.session_state.authenticated = True
                    token = storage1.create_session(username)
                    st.session_state.session_token = token
                    st.query_params["sid"] = token
                    storage1.log_audit_event(username, user_data["role"], "login", {})
                    st.rerun()
                else:
                    st.error("Invalid username or password")
        return

    # ---------------- AUTHENTICATED ----------------
    current_user = get_current_user()
    if not current_user:
        st.error("User not found")
        return

    role = current_user["role"]
    full_name = current_user.get("full_name", current_user["username"])
    avatar_initial = (full_name or current_user["username"])[0].upper()

    # ---- Maintenance mode gate (super admin can toggle; everyone else is blocked while it's on) ----
    maintenance = storage1.get_maintenance_status()
    is_super = storage1.is_super_admin(current_user["username"])
    if maintenance.get("enabled") and not is_super:
        st.markdown(
            f"""
    <div style="display:flex; align-items:center; justify-content:center; min-height:65vh;">
      <div style="max-width:480px; text-align:center; background: rgba(15,23,42,0.75);
                  border: 1px solid rgba(251,191,36,0.35); border-radius:20px; padding:36px 28px;">
        <div style="font-size:2.6rem;">🚧</div>
        <div style="font-size:1.4rem; font-weight:700; color:#f8fafc; margin-top:8px; letter-spacing:0.5px;">
            System Maintenance
        </div>
        <div style="font-size:1.05rem; color:#cbd5e1; margin-top:8px; line-height:1.5;">
            {maintenance.get("message") or storage1.DEFAULT_MAINTENANCE_MESSAGE}
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True,
        )
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("Logout", use_container_width=True, key="maintenance_logout"):
                storage1.log_audit_event(current_user["username"], role, "logout", {})
                storage1.invalidate_session(st.session_state.get("session_token"))
                try:
                    del st.query_params["sid"]
                except Exception:
                    pass
                st.session_state.authenticated = False
                st.session_state.user = None
                st.session_state.session_token = None
                st.rerun()
        st.stop()

    analysis_allowed = role in ("admin", "manager") or is_super

    def _dashboard_page():
        # ---- Header: greeting + role badge, redesigned as a proper header bar ----
        st.markdown(f"""
        <div class="app-header">
            <div class="app-header-left">
                <div class="app-header-avatar">{avatar_initial}</div>
                <div>
                    <p class="app-header-greeting">👋 Hi, {full_name}</p>
                    <div class="app-header-sub">Solar PV String Analytics · welcome back</div>
                </div>
            </div>
            <div class="app-header-right">
                <span class="user-badge-{role}">{ROLE_BADGES.get(role, role)}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ---- Sidebar ----
        st.sidebar.markdown('<h2><i class="fas fa-bolt" style="color:#38bdf8;"></i> PV String Template</h2>', unsafe_allow_html=True)

        with st.sidebar.expander("👤 My Profile", expanded=False):
            st.write(f"**Full Name:** {full_name}")
            st.write(f"**Username:** {current_user['username']}")
            st.write(f"**Role:** {ROLE_BADGES.get(role, role)}", unsafe_allow_html=True)
            st.write(f"**Assigned Plots:** {', '.join(current_user.get('assigned_plots', [])) or 'None'}")

            st.markdown("---")
            st.write("**Change My Password**")
            self_new_pw = st.text_input("New Password", type="password", key="self_pw_1")
            self_new_pw_confirm = st.text_input("Confirm New Password", type="password", key="self_pw_2")
            if st.button("Update Password", key="self_pw_update_btn"):
                if not self_new_pw or self_new_pw != self_new_pw_confirm:
                    st.error("Passwords don't match or are empty.")
                else:
                    ok, msg = storage1.reset_password(current_user["username"], self_new_pw)
                    if ok:
                        storage1.log_audit_event(current_user["username"], role, "password_reset_self", {})
                        st.success(msg)
                    else:
                        st.error(msg)

        st.sidebar.markdown(f"**User:** {current_user['username']} ({ROLE_BADGES.get(role, role)})", unsafe_allow_html=True)

        if st.sidebar.button(" Logout", use_container_width=True):
            storage1.log_audit_event(current_user["username"], role, "logout", {})
            storage1.invalidate_session(st.session_state.get("session_token"))
            try:
                del st.query_params["sid"]
            except Exception:
                pass
            st.session_state.authenticated = False
            st.session_state.user = None
            st.session_state.session_token = None
            st.rerun()

        st.sidebar.markdown("---")

        # ---- Maintenance mode control (super admin only) ----
        if is_super:
            if maintenance.get("enabled"):
                st.warning(
                    f"🚧 Maintenance mode is **ON** (enabled by {maintenance.get('enabled_by', 'unknown')}). "
                    "Every other user is currently blocked from the app."
                )
            with st.sidebar.expander("🛠️ Maintenance Mode (Super Admin)", expanded=False):
                st.caption("While enabled, every user except super admins sees a maintenance popup and can't use the app.")
                maint_message = st.text_input(
                    "Popup message", value=maintenance.get("message") or storage1.DEFAULT_MAINTENANCE_MESSAGE,
                    key="maintenance_message_input",
                )
                m_col1, m_col2 = st.columns(2)
                with m_col1:
                    if not maintenance.get("enabled"):
                        if st.button("🚧 Enable Maintenance Mode", use_container_width=True, key="enable_maintenance_btn"):
                            storage1.set_maintenance_mode(True, current_user["username"], maint_message)
                            storage1.log_audit_event(current_user["username"], role, "maintenance_mode_enabled", {"message": maint_message})
                            st.rerun()
                with m_col2:
                    if maintenance.get("enabled"):
                        if st.button("✅ Disable Maintenance Mode", use_container_width=True, key="disable_maintenance_btn"):
                            storage1.set_maintenance_mode(False, current_user["username"])
                            storage1.log_audit_event(current_user["username"], role, "maintenance_mode_disabled", {})
                            st.rerun()

        st.sidebar.markdown("---")

        # ---- File upload (admin only) ----
        st.sidebar.subheader("📁 File Management")

        if is_admin():
            latest_upload = storage1.get_latest_upload()
            if latest_upload:
                st.sidebar.info(f"Current file: {latest_upload['original_filename']}\n"
                                 f"Snapshot date: {latest_upload['snapshot_date']}\n"
                                 f"Uploaded: {latest_upload['upload_timestamp']}")

            snapshot_date = st.sidebar.date_input("Snapshot date for this upload", value=datetime.now().date())
            uploaded_file = st.sidebar.file_uploader("Upload new SCADA Report (.xlsx)", type=["xlsx"])
            if uploaded_file:
                file_bytes = uploaded_file.getvalue()
                file_hash = hashlib.md5(file_bytes).hexdigest()
                # The file_uploader widget keeps holding its last value even
                # after st.rerun(), so without this guard the same file kept
                # getting reprocessed (and re-logged to the audit log) on every
                # single rerun - looping "file_uploaded" events. Only process a
                # given file once per browser session unless it actually changes.
                upload_signature = f"{snapshot_date}:{file_hash}"
                already_processed = st.session_state.get("last_processed_upload_signature") == upload_signature
                if not already_processed:
                    ok, msg = process_and_save_upload(
                        file_bytes, uploaded_file.name, snapshot_date,
                        current_user["username"], role,
                    )
                    if ok:
                        st.session_state.last_processed_upload_signature = upload_signature
                        st.session_state.header_snapshot_date = str(snapshot_date)
                        st.sidebar.success(msg)
                        st.rerun()
                    else:
                        st.sidebar.error(msg)
        else:
            latest_upload = storage1.get_latest_upload()
            if latest_upload:
                st.sidebar.info(f"Current file: {latest_upload['original_filename']}")
                st.sidebar.caption(f"Snapshot date: {latest_upload['snapshot_date']}")
            else:
                st.sidebar.warning("No file available. Please contact admin.")

        st.sidebar.markdown("---")

        # ---- Load current data ----
        latest_upload = storage1.get_latest_upload()
        if not latest_upload:
            st.info("No SCADA file available. Please contact admin to upload one.")
            return

        # ---- Header calendar: browse any previously preprocessed snapshot date ----
        selected_snapshot_date, using_latest = render_header_calendar()

        if selected_snapshot_date and not using_latest:
            processed_dataframes, snapshot_entry = storage1.get_processed_dataframes_for_date(selected_snapshot_date)
            if not processed_dataframes:
                st.warning(f"No preprocessed data could be loaded for {selected_snapshot_date}. Showing the latest snapshot instead.")
                processed_dataframes = None
        else:
            processed_dataframes = None

        if processed_dataframes is None:
            # Only the processed CSVs are kept on disk (storage optimization -
            # the original .xlsx is never persisted), so the "latest" snapshot
            # is loaded the same way any other snapshot date is: straight from
            # its already-preprocessed CSVs, with no re-parsing needed.
            processed_dataframes, _ = storage1.get_processed_dataframes_for_date(latest_upload["snapshot_date"])
            if not processed_dataframes:
                st.error("Could not load the latest snapshot's preprocessed data from backend storage.")
                return

        if role == "engineer":
            allowed_plots = current_user.get("assigned_plots", [])
            if allowed_plots:
                st.sidebar.markdown("---")
                st.sidebar.subheader("Assigned Plots")
                st.sidebar.write(", ".join(allowed_plots))

        sheet_selection = st.sidebar.selectbox("Select Sheet", list(processed_dataframes.keys()))
        df_selected = processed_dataframes[sheet_selection].copy()

        if role == "engineer":
            allowed_plots = current_user.get("assigned_plots", [])
            if allowed_plots:
                df_selected = df_selected[df_selected["Plot"].isin(allowed_plots)]

        st.sidebar.markdown("---")
        st.sidebar.subheader("🔎 Filters")

        # Initialize sidebar filters in session state
        if "sidebar_filters" not in st.session_state:
            st.session_state.sidebar_filters = {
                "plot": "All",
                "block": "All",
                "sacu": "All"
            }

        plots = ["All"] + sorted_filter_options(df_selected["Plot"])
        selected_plot = st.sidebar.selectbox(
            "Plot",
            plots,
            index=0 if st.session_state.sidebar_filters["plot"] == "All" else plots.index(st.session_state.sidebar_filters["plot"]) if st.session_state.sidebar_filters["plot"] in plots else 0
        )
        st.session_state.sidebar_filters["plot"] = selected_plot

        filtered_df = df_selected.copy()
        if selected_plot != "All":
            filtered_df = filtered_df[filtered_df["Plot"] == selected_plot]

        blocks = ["All"] + sorted_filter_options(filtered_df["Block"])
        selected_block = st.sidebar.selectbox(
            "Block",
            blocks,
            index=0 if st.session_state.sidebar_filters["block"] == "All" else blocks.index(st.session_state.sidebar_filters["block"]) if st.session_state.sidebar_filters["block"] in blocks else 0
        )
        st.session_state.sidebar_filters["block"] = selected_block

        if selected_block != "All":
            filtered_df = filtered_df[filtered_df["Block"] == selected_block]

        sacus = ["All"] + sorted_filter_options(filtered_df["SACU"])
        selected_sacu = st.sidebar.selectbox(
            "SACU",
            sacus,
            index=0 if st.session_state.sidebar_filters["sacu"] == "All" else sacus.index(st.session_state.sidebar_filters["sacu"]) if st.session_state.sidebar_filters["sacu"] in sacus else 0
        )
        st.session_state.sidebar_filters["sacu"] = selected_sacu

        if selected_sacu != "All":
            filtered_df = filtered_df[filtered_df["SACU"] == selected_sacu]

        # ---- User management (admin / manager) ----
        user_management_ui()

        inverter_col = get_inverter_column_cached(filtered_df)

        tab1, tab2, tab3, tab4, tab5 = st.tabs(
        [
            "Dashboard",
            "PV String Details",
            "Data Table",
            "Restore & TAT",
            "Audit Log",
        ]
    )

        with tab1:
            if not filtered_df.empty:
                main_dashboard_tab(
                    filtered_df, sheet_df=df_selected, sheet_name=sheet_selection,
                    snapshot_date=selected_snapshot_date or latest_upload["snapshot_date"],
                )
            else:
                st.warning("No data available with current filters and permissions")

        with tab2:
            if not filtered_df.empty:
                create_pv_string_tab(filtered_df)
            else:
                st.warning("No data available for PV string analysis")

        with tab3:
            st.subheader("Inverter Data Table")
            if not filtered_df.empty:
                display_df = filtered_df.copy()
                if inverter_col and inverter_col != "Inverter ID":
                    display_df = display_df.rename(columns={inverter_col: "Inverter ID"})

                _, table_pv_current_cols = get_pv_string_columns_cached(filtered_df)
                pv_cols_in_display = [c for c in table_pv_current_cols if c in display_df.columns]
                column_config = {
                    "Availability (%)": st.column_config.ProgressColumn("Availability (%)", min_value=0, max_value=100, format="%.2f%%"),
                    "Failure Percentage (%)": st.column_config.NumberColumn("Failure Percentage (%)", format="%.2f%%"),
                }

                if pv_cols_in_display:
                    # Multi-tier Amp-based coloring (not just red/green) - same
                    # scale used elsewhere in the app: >5A / >3A / >1.5A / >0.5A / else.
                    def _pv_cell_color(x):
                        if pd.notna(x) and isinstance(x, (int, float)):
                            return f'background-color: {get_string_health_color(x)}; color: white; font-weight: bold;'
                        return ''

                    styled_display = display_df.style
                    for col in pv_cols_in_display:
                        styled_display = styled_display.map(_pv_cell_color, subset=[col])
                    st.caption("PV-I current coloring: Green >5A, Light Green >3A, Yellow >1.5A, Orange >0.5A, Red <=0.5A.")
                    st.dataframe(styled_display, use_container_width=True, column_config=column_config)
                else:
                    st.dataframe(display_df, use_container_width=True, column_config=column_config)

                download_bytes = create_colored_excel_download({sheet_selection: filtered_df})
                st.download_button(
                    label="Download Filtered Excel (color-coded)", data=download_bytes,
                    file_name=f"processed_{sheet_selection}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=_log_download, args=(f"filtered_excel:{sheet_selection}",),
                )
            else:
                st.info("No data available")

        with tab4:
            if not filtered_df.empty:
                restore.get_restore_tab(
                    processed_dataframes, filtered_df,
                    sheet_name=sheet_selection, user_role=role, username=current_user["username"],
                    upload_handler=lambda file_bytes, filename, snap_date: process_and_save_upload(
                        file_bytes, filename, snap_date, current_user["username"], role
                    ) if is_admin() else (False, "Only admins can upload snapshots."),
                    snapshot_date=selected_snapshot_date or latest_upload["snapshot_date"],
                )
            else:
                st.warning("No data available for TAT analysis")

        with tab5:
            audit_log_tab()

    pages = [
        st.Page(_dashboard_page, title="Dashboard", icon=":material/dashboard:", url_path="dashboard", default=True),
    ]
    if analysis_allowed:
        pages.append(
            st.Page(lambda: analysis.render_analysis_page(current_user), title="Analysis", icon=":material/query_stats:", url_path="analysis")
        )

    pg = st.navigation(pages, position="sidebar")
    pg.run()

if __name__ == "__main__":
    main()
