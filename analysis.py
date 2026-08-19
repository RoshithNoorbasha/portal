"""
PV Plant String-Level Monitoring Dashboard
===========================================
A Streamlit re-implementation of the Colab SCADA processing pipeline:
  1. Per-day Excel ingestion (separate Day1 / Day2 / Day3 uploads) & per-inverter
     string metrics, with duplicate-inverter-row handling.
  2. Cross-day PV string fault detection (Newly Failed / Re-Failed), negative-value
     detection, and low-performing-string detection.
  3. A polished, cached, interactive dashboard + downloadable Excel report.

Design goals:
  - Nothing is re-parsed / re-computed on every widget interaction.
    Every expensive step is wrapped in st.cache_data, keyed off the raw
    file bytes + the active configuration, so Streamlit reruns are cheap.
  - Session state holds the *results* of the pipeline so tab switches,
    filter changes, etc. never touch the Excel parser again.
  - The pipeline re-runs automatically whenever an uploaded file or a
    setting changes (no manual "Process" click required) so the dashboard
    always reflects the latest SCADA logs.
"""

import io
import json
import hashlib
import re
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

import storage1  # for the admin/manager/super-admin permission check


def render_analysis_page(current_user):
    """
    Renders the multi-day SCADA string fault analysis page.
    Restricted to admin / manager / super-admin roles - the caller (app.py)
    is expected to only add this page to navigation for those roles, but we
    re-check here too as a safety net in case this is ever called directly.
    """
    role = (current_user or {}).get("role")
    is_authorized = role in ("admin", "manager") or storage1.is_super_admin((current_user or {}).get("username"))
    if not is_authorized:
        st.error("You don't have permission to view this page. Contact an admin or manager for access.")
        st.stop()

    # Accent colors — used for UI charts/badges AND for the (light-background)
    # Excel report, so these stay as vivid, print-friendly hues.
    COLORS = {
        "primary_dark": "#1A365D",
        "primary_blue": "#2B6CB0",
        "primary_light": "#EBF4FF",
        "success_green": "#38A169",
        "success_light": "#F0FFF4",
        "danger_red": "#E53E3E",
        "danger_light": "#FFF5F5",
        "warning_orange": "#ED8936",
        "warning_light": "#FFFAF0",
        "info_purple": "#805AD5",
        "info_light": "#FAF5FF",
        "dark_grey": "#2D3748",
        "medium_grey": "#718096",
        "light_grey": "#E2E8F0",
        "white": "#FFFFFF",
    }

    # Dark-mode surface palette — used only for the Streamlit UI chrome.
    DARK = {
        "bg": "#0B1220",
        "sidebar_from": "#0B1220",
        "sidebar_to": "#0F1729",
        "surface": "#141B2D",
        "surface_alt": "#1B2436",
        "border": "#26314A",
        "text": "#E6EDF3",
        "text_muted": "#93A1B5",
    }

    CUSTOM_CSS = f"""
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.2/css/all.min.css">
    <style>
        :root {{
            --bg: {DARK['bg']};
            --surface: {DARK['surface']};
            --surface-alt: {DARK['surface_alt']};
            --border: {DARK['border']};
            --text: {DARK['text']};
            --text-muted: {DARK['text_muted']};
        }}

        .stApp {{ background-color: var(--bg); }}
        #MainMenu, footer {{ visibility: hidden; }}
        body, .stMarkdown, .stCaption, p, span, label {{ color: var(--text); }}

        .hero-banner {{
            background: linear-gradient(120deg, #0B1830 0%, {COLORS['primary_blue']} 160%);
            padding: 1.6rem 2rem;
            border-radius: 16px;
            color: #F5F9FF;
            margin-bottom: 1.4rem;
            border: 1px solid var(--border);
            box-shadow: 0 8px 28px rgba(0,0,0,0.45);
        }}
        .hero-banner h1 {{ margin: 0; font-size: 1.9rem; font-weight: 800; color: #F5F9FF; }}
        .hero-banner h1 i {{ color: {COLORS['warning_orange']}; margin-right: 0.55rem; }}
        .hero-banner p {{ margin: 0.3rem 0 0 0; opacity: 0.85; font-size: 0.95rem; }}

        .sidebar-brand {{
            font-size: 1.25rem; font-weight: 800; color: #F5F9FF; margin-bottom: 0.1rem;
        }}
        .sidebar-brand i {{ color: {COLORS['warning_orange']}; margin-right: 0.5rem; }}

        .meta-caption {{
            color: var(--text-muted);
            font-size: 0.85rem;
            margin: -0.4rem 0 0.9rem 0;
        }}
        .meta-caption i {{ margin-right: 0.35rem; color: {COLORS['primary_blue']}; }}

        .live-pill {{
            display: inline-flex; align-items: center; gap: 0.35rem;
            background: rgba(56,161,105,0.15); color: {COLORS['success_green']};
            padding: 0.2rem 0.7rem; border-radius: 999px; font-size: 0.75rem; font-weight: 700;
            margin-left: 0.5rem;
        }}
        .live-pill i {{ font-size: 0.6rem; }}

        div[data-testid="stMetric"] {{
            background: var(--surface);
            border-radius: 14px;
            padding: 1rem 1rem 0.6rem 1rem;
            box-shadow: 0 2px 10px rgba(0,0,0,0.35);
            border: 1px solid var(--border);
        }}
        div[data-testid="stMetricLabel"] {{
            font-weight: 700;
            letter-spacing: 0.02em;
            color: var(--text-muted) !important;
            text-transform: uppercase;
            font-size: 0.72rem;
        }}
        div[data-testid="stMetricValue"] {{ color: var(--text) !important; }}

        .section-title {{
            font-size: 1.05rem;
            font-weight: 800;
            color: var(--text);
            margin: 0.4rem 0 0.6rem 0;
            padding-bottom: 0.4rem;
            border-bottom: 3px solid {COLORS['primary_blue']};
            display: inline-block;
        }}
        .section-title i {{ color: {COLORS['primary_blue']}; margin-right: 0.55rem; }}

        .priority-card {{
            background: var(--surface);
            border-left: 6px solid {COLORS['danger_red']};
            border-radius: 10px;
            padding: 0.7rem 1rem;
            margin-bottom: 0.6rem;
            border: 1px solid var(--border);
            box-shadow: 0 2px 10px rgba(0,0,0,0.3);
        }}
        .priority-card .label {{ font-size: 0.72rem; font-weight: 700; color: var(--text-muted); text-transform: uppercase; }}
        .priority-card .label i {{ margin-right: 0.35rem; color: {COLORS['danger_red']}; }}
        .priority-card .value {{ font-size: 1.15rem; font-weight: 800; color: {COLORS['danger_red']}; }}
        .priority-card .note {{ font-size: 0.78rem; color: var(--text-muted); font-style: italic; }}

        .status-pill {{
            display: inline-block; padding: 0.15rem 0.7rem; border-radius: 999px;
            font-size: 0.75rem; font-weight: 700;
        }}
        .pill-green {{ background: rgba(56,161,105,0.15); color: {COLORS['success_green']}; }}
        .pill-red {{ background: rgba(229,62,62,0.15); color: {COLORS['danger_red']}; }}
        .pill-orange {{ background: rgba(237,137,54,0.15); color: {COLORS['warning_orange']}; }}
        .pill-purple {{ background: rgba(128,90,213,0.15); color: {COLORS['info_purple']}; }}

        section[data-testid="stSidebar"] {{
            background: linear-gradient(180deg, {DARK['sidebar_from']} 0%, {DARK['sidebar_to']} 100%);
            border-right: 1px solid var(--border);
        }}
        section[data-testid="stSidebar"] * {{ color: #EDF2F7 !important; }}
        section[data-testid="stSidebar"] .stButton>button {{
            background: {COLORS['primary_blue']}; color: white !important; border: none;
            font-weight: 700; border-radius: 8px;
        }}
        section[data-testid="stSidebar"] div[data-testid="stExpander"] {{
            background: var(--surface-alt); border-radius: 10px; border: 1px solid var(--border);
        }}

        div[data-testid="stDataFrame"], div[data-testid="stTable"] {{
            border: 1px solid var(--border); border-radius: 10px; overflow: hidden;
        }}

        button[data-baseweb="tab"] {{ font-weight: 600; }}
    </style>
    """
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    # ============================================================================
    # DEFAULT CONFIGURATION
    # ============================================================================

    DEFAULT_TOTAL_ACTIVE_STRINGS = 19
    DEFAULT_OVERRIDES_ROWS = [
        {"Plot": "P2", "Block": "IB1", "Active Strings": 18},
        {"Plot": "P2", "Block": "IB3", "Active Strings": 17},
        {"Plot": "P2", "Block": "IB4", "Active Strings": 18},
        {"Plot": "P2", "Block": "IB5", "Active Strings": 18},
        {"Plot": "P6", "Block": "IB1", "Active Strings": 18},
        {"Plot": "P6", "Block": "IB2", "Active Strings": 18},
        {"Plot": "P6", "Block": "IB3", "Active Strings": 18},
        {"Plot": "P6", "Block": "IB5", "Active Strings": 18},
        {"Plot": "P6", "Block": "IB6", "Active Strings": 18},
        {"Plot": "P6", "Block": "IB7", "Active Strings": 18},
    ]
    DEFAULT_THRESHOLD = 0.5
    # An inverter ID that appears this many times (or more) within a single day's
    # sheet is treated as a data-integrity problem and fully excluded from
    # calculations & the dashboard. Below this, duplicate rows are still
    # considered "present in SCADA" and kept in the calculations, but flagged.
    DEFAULT_DUPLICATE_EXCLUDE_THRESHOLD = 4
    # A working string (current > threshold) whose reading falls below this
    # fraction of its own inverter's median working-string current is flagged
    # as "low performing" (e.g. 0.7 = below 70% of its siblings).
    DEFAULT_LOW_PERF_RATIO = 0.70

    PV_CURRENT_COLUMNS = [f"PV-I{i}" for i in range(1, 29)]
    INVERTER_ID_CANDIDATES = [
        "Inverter ID", "Inverter_ID", "Inverter", "ID",
        "Device Name", "String Inverter", "Inverters",
    ]
    DAY_LABELS = ["Day1", "Day2", "Day3"]

    # Sheet keys + display labels offered on the Export tab, and used by
    # generate_excel_report's `included_sheets` filter (requirement #3).
    EXPORT_SHEET_OPTIONS = [
        ("dashboard", "Dashboard (KPIs + Plot-wise status)"),
        ("trend", "3-Day Trend"),
        ("plot_summary", "Plot Summary"),
        ("block_summary", "Block Summary"),
        ("inverter_matrix", "Inverter Matrix"),
        ("failed", "Failed Strings"),
        ("refailed", "Re-Failed Strings"),
        ("negative", "Negative Values"),
        ("low_perf", "Low Performance"),
        ("duplicates", "Duplicate Inverters"),
    ]

    # ============================================================================
    # SESSION STATE
    # ============================================================================

    def init_state():
        defaults = {
            "processed_days": {},          # day_key -> DataFrame (calc-ready, duplicates resolved)
            "duplicate_rows_by_day": {},   # day_key -> DataFrame of flagged duplicate rows (kept in calc)
            "excluded_rows_by_day": {},    # day_key -> DataFrame of excluded rows (>= threshold duplicates)
            "file_names_by_day": {},       # day_key -> filename
            "day_keys": [],
            "multiday_df": None,
            "fault_df": None,
            "date_mapping": {},
            "last_signature": None,
            "excel_report_bytes": None,
            "excel_report_name": None,
        }
        for k, v in defaults.items():
            if k not in st.session_state:
                st.session_state[k] = v

    init_state()

    # ============================================================================
    # CORE PARSING / METRIC LOGIC
    # ============================================================================

    def normalize_text(value):
        if pd.isna(value):
            return ""
        return str(value).strip().upper()


    def map_inverter_to_sacu(inverter_id_str):
        if not isinstance(inverter_id_str, str):
            return "Invalid Inverter ID"
        match = re.search(r"-(\d[\.\-]\d)-", inverter_id_str)
        if match:
            sacu_identifier = match.group(1)
            try:
                first_digit_str = sacu_identifier.split(".")[0] if "." in sacu_identifier else sacu_identifier.split("-")[0]
                first_digit = int(first_digit_str)
                if first_digit in [1, 2]:
                    return "SACU-1"
                elif first_digit in [3, 4]:
                    return "SACU-2"
            except ValueError:
                pass
        return "Unknown SACU"


    def extract_plot(inverter_id_str):
        if isinstance(inverter_id_str, str):
            parts = inverter_id_str.split("-")
            if parts:
                return parts[0].strip()
        return "Unknown Plot"


    def extract_block(inverter_id_str):
        if isinstance(inverter_id_str, str):
            parts = inverter_id_str.split("-")
            if len(parts) > 1:
                return parts[1].strip()
        return "Unknown Block"


    def find_header_row_index(file_bytes, sheet_name, possible_header_columns, max_rows_to_check=100):
        try:
            temp_df = pd.read_excel(
                io.BytesIO(file_bytes), sheet_name=sheet_name, header=None,
                nrows=max_rows_to_check, engine="openpyxl",
            )
        except Exception:
            return None
        possible_headers_lower = [str(c).strip().lower() for c in possible_header_columns]
        for i, row in temp_df.iterrows():
            row_values_lower = [str(v).strip().lower() for v in row.dropna()]
            if any(c in row_values_lower for c in possible_headers_lower):
                return i
        return None


    def get_available_pv_columns(df):
        normalized_map = {str(c).strip().upper(): c for c in df.columns}
        return [normalized_map[c.upper()] for c in PV_CURRENT_COLUMNS if c.upper() in normalized_map]


    def get_total_active_strings(plot, block, overrides, default_active):
        plot_key, block_key = normalize_text(plot), normalize_text(block)
        return overrides.get(plot_key, {}).get(block_key, default_active)


    def calculate_working_string_count(row, pv_columns, threshold):
        count = 0
        for col in pv_columns:
            value = pd.to_numeric(row.get(col), errors="coerce")
            if pd.notna(value) and value > threshold:
                count += 1
        return count


    def apply_string_metrics(df, overrides, default_active, threshold, plot_col="Plot", block_col="Block"):
        """Compute Total Active / Working / Failed string counts per row.

        BUGFIX: the raw "working" count is a straight tally of PV-I columns whose
        current exceeds the threshold. When a sheet has more PV-I columns than an
        inverter actually has active strings (common on partially-wired blocks),
        stray positive noise on the unused columns could previously push the raw
        working count *above* the block's Total Active Strings. That silently
        masked real failures, because Failed = Total Active - Working was then
        clipped to 0 instead of reflecting the true state. Working Strings is now
        capped at Total Active Strings before Failed Strings is derived.
        """
        pv_columns = get_available_pv_columns(df)
        if not pv_columns:
            df["Total Active Strings"] = 0
            df["Working String Count"] = 0
            df["Failed String Count"] = 0
            df["Availability (%)"] = 0.0
            df["Failure Percentage (%)"] = 0.0
            return df

        df["Total Active Strings"] = df.apply(
            lambda r: get_total_active_strings(r.get(plot_col), r.get(block_col), overrides, default_active), axis=1
        )
        raw_working = df.apply(lambda r: calculate_working_string_count(r, pv_columns, threshold), axis=1)
        # Cap working strings at the total active strings for that block (fix #1).
        df["Working String Count"] = pd.concat([raw_working, df["Total Active Strings"]], axis=1).min(axis=1)
        df["Failed String Count"] = (df["Total Active Strings"] - df["Working String Count"]).clip(lower=0)
        df["Availability (%)"] = ((df["Working String Count"] / df["Total Active Strings"]) * 100).fillna(0).round(2)
        df["Failure Percentage (%)"] = ((df["Failed String Count"] / df["Total Active Strings"]) * 100).fillna(0).round(2)
        return df


    @st.cache_data(show_spinner=False)
    def process_scada_excel_bytes(file_bytes: bytes, file_name: str, overrides_json: str, default_active: int, threshold: float):
        """Parse one uploaded workbook into per-sheet, metric-enriched DataFrames.
        Cached on (file bytes, filename, config) - identical uploads/config never re-parse."""
        overrides = json.loads(overrides_json)
        processed = {}
        missing_log = {}

        excel_file = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")

        for sheet_name in excel_file.sheet_names:
            header_row_index = find_header_row_index(file_bytes, sheet_name, INVERTER_ID_CANDIDATES)
            if header_row_index is None:
                continue
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, skiprows=header_row_index, header=0, engine="openpyxl")
            except Exception:
                continue

            df.dropna(how="all", inplace=True)
            df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed:", case=False, regex=True)]

            actual_inverter_col = None
            cols_lower_map = {str(c).strip().lower(): c for c in df.columns}
            for cand in INVERTER_ID_CANDIDATES:
                if cand in df.columns:
                    actual_inverter_col = cand
                    break
                if cand.strip().lower() in cols_lower_map:
                    actual_inverter_col = cols_lower_map[cand.strip().lower()]
                    break

            if not actual_inverter_col:
                missing_log[sheet_name] = INVERTER_ID_CANDIDATES
                continue

            # Standardise the inverter-id column name for reliable downstream joins
            if actual_inverter_col != "String Inverter":
                df.rename(columns={actual_inverter_col: "String Inverter"}, inplace=True)
            actual_inverter_col = "String Inverter"

            df[actual_inverter_col] = df[actual_inverter_col].astype(str).str.strip()
            df["Plot"] = df[actual_inverter_col].apply(extract_plot)
            df["Block"] = df[actual_inverter_col].apply(extract_block)
            df.loc[(df["Plot"] == "P6") & (df["Block"] == "IB09"), "Block"] = "IB9"
            df["SACU"] = df[actual_inverter_col].apply(map_inverter_to_sacu)

            pv_other_cols_to_drop = [c for c in df.columns if re.match(r"^PV\d{1,2}$", str(c)) and c not in PV_CURRENT_COLUMNS]
            if pv_other_cols_to_drop:
                df.drop(columns=pv_other_cols_to_drop, inplace=True)

            df = apply_string_metrics(df, overrides, default_active, threshold, plot_col="Plot", block_col="Block")

            preferred = ["Plot", "Block", actual_inverter_col, "SACU", "Total Active Strings",
                         "Working String Count", "Failed String Count", "Availability (%)", "Failure Percentage (%)"]
            remaining = [c for c in df.columns if c not in preferred]
            df = df[[c for c in preferred if c in df.columns] + remaining]

            processed[sheet_name] = df

        return processed, missing_log


    def extract_date_from_filename(filename: str):
        date_match = re.search(r"(\d{1,2}-\d{1,2}-\d{4})", filename) or re.search(r"(\d{1,2}-\d{1,2}-\d{2})", filename)
        if not date_match:
            return None
        date_str = date_match.group(1)
        try:
            year_part = date_str.split("-")[2]
            fmt = "%d-%m-%Y" if len(year_part) == 4 else "%d-%m-%y"
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            return None


    # ============================================================================
    # DUPLICATE INVERTER HANDLING  (requirement #3)
    # ============================================================================

    def handle_duplicate_inverters(df, id_col="String Inverter", exclude_threshold=4):
        """Resolve duplicate inverter-ID rows within a single day's sheet.

        Rules:
          - An inverter ID that occurs >= exclude_threshold times in the raw SCADA
            sheet is treated as corrupted/unreliable data and is fully EXCLUDED
            from calculations and the dashboard.
          - An inverter ID that occurs more than once but fewer than
            exclude_threshold times is still "present in SCADA" and IS included
            in calculations (its first occurrence is used so string counts are
            not double-counted), but every such ID is surfaced separately so it
            can be reviewed as a duplicate inverter ID on the dashboard.
          - Non-duplicated rows pass through unchanged.

        Returns (df_calc, df_duplicate_rows, df_excluded_rows, duplicate_ids, excluded_ids)
        """
        if df.empty or id_col not in df.columns:
            return df.copy(), df.iloc[0:0].copy(), df.iloc[0:0].copy(), [], []

        work = df.copy()
        work[id_col] = work[id_col].astype(str).str.strip()
        counts = work[id_col].value_counts()

        excluded_ids = sorted(counts[counts >= exclude_threshold].index.tolist())
        duplicate_ids = sorted(counts[(counts > 1) & (counts < exclude_threshold)].index.tolist())

        df_excluded = work[work[id_col].isin(excluded_ids)].copy()
        df_remaining = work[~work[id_col].isin(excluded_ids)].copy()

        # SCADA-present duplicates (<threshold) stay in the calculation set, but
        # only their first reading is counted per inverter to avoid inflating
        # working/failed totals from the same physical inverter.
        df_calc = df_remaining.drop_duplicates(subset=[id_col], keep="first").reset_index(drop=True)
        df_duplicate_rows = df_remaining[df_remaining[id_col].isin(duplicate_ids)].reset_index(drop=True)

        return df_calc, df_duplicate_rows, df_excluded, duplicate_ids, excluded_ids


    @st.cache_data(show_spinner=False)
    def build_multiday_comparison(day_frames: dict, day_keys: tuple):
        """Melt each day's PV-I columns to long format and outer-merge across days."""
        long_frames = []
        for day_key in day_keys:
            df_day = day_frames[day_key]
            pv_cols_exist = [c for c in PV_CURRENT_COLUMNS if c in df_day.columns]
            if not pv_cols_exist:
                continue
            id_vars = [c for c in ["String Inverter", "Plot", "Block", "Grid"] if c in df_day.columns]
            df_long = df_day.melt(id_vars=id_vars, value_vars=pv_cols_exist, var_name="MPPT PV No", value_name=f"PV_Value_{day_key}")
            long_frames.append(df_long)

        if not long_frames:
            return pd.DataFrame()

        common_keys = [c for c in ["String Inverter", "Plot", "Block", "Grid", "MPPT PV No"] if c in long_frames[0].columns]
        merged = long_frames[0]
        for frame in long_frames[1:]:
            merged = pd.merge(merged, frame, on=common_keys, how="outer")
        return merged


    @st.cache_data(show_spinner=False)
    def compute_fault_events(df_multiday: pd.DataFrame, day_keys: tuple, threshold: float):
        """Newly Failed / Re-Failed / Current Failure detection across the last 2-3 days."""
        id_cols = [c for c in ["Plot", "Block", "Grid", "String Inverter", "MPPT PV No"] if c in df_multiday.columns]
        comparison_day_keys = list(day_keys[-3:])
        number_of_days = len(comparison_day_keys)
        fault_details = []

        for _, row in df_multiday.iterrows():
            current_row = {c: row[c] for c in id_cols}
            day_values = []
            for day_key in comparison_day_keys:
                pv_col = f"PV_Value_{day_key}"
                day_values.append(pd.to_numeric(row[pv_col], errors="coerce") if pv_col in row.index else None)

            latest_value = day_values[-1]
            if pd.isna(latest_value) or latest_value > threshold:
                continue

            if number_of_days == 2:
                previous_value = day_values[-2]
                if pd.isna(previous_value):
                    continue
                if previous_value > threshold and latest_value <= threshold:
                    fault_type = "NEWLY FAILED"
                else:
                    continue
            else:
                day1_value, day2_value, day3_value = day_values[0], day_values[1], day_values[2]
                if pd.isna(day3_value) or pd.isna(day2_value):
                    continue
                if day2_value > threshold and day3_value <= threshold:
                    if pd.notna(day1_value) and day1_value > threshold:
                        fault_type = "NEWLY FAILED"
                    elif pd.notna(day1_value) and day1_value <= threshold:
                        fault_type = "RE-FAILED"
                    else:
                        fault_type = "CURRENT FAILURE"
                else:
                    continue

            for i, pv_value in enumerate(day_values, start=1):
                current_row[f"Day {i}"] = pv_value
            current_row["Fault Type"] = fault_type
            fault_details.append(current_row)

        if not fault_details:
            return pd.DataFrame(columns=id_cols + [f"Day {i}" for i in range(1, number_of_days + 1)] + ["Fault Type"])

        day_columns = [f"Day {i}" for i in range(1, number_of_days + 1)]
        final_columns = [c for c in id_cols + day_columns + ["Fault Type"] if c in pd.DataFrame(fault_details).columns]
        df_out = pd.DataFrame(fault_details)[final_columns]
        sort_cols = [c for c in ["Plot", "Block", "Grid", "String Inverter", "MPPT PV No"] if c in df_out.columns]
        return df_out.sort_values(by=sort_cols).reset_index(drop=True)


    # ============================================================================
    # NEGATIVE VALUES  (requirement #4)
    # ============================================================================

    def compute_negative_events(df_multiday: pd.DataFrame, day_keys: list):
        """Return (current_negative_df, continuous_negative_df).

        current_negative_df: every string whose LATEST reading is negative.
        continuous_negative_df: subset of the above that was also negative on the
        immediately preceding day(s), i.e. a persisting fault rather than a
        one-off reading.
        """
        empty_cols = ["Plot", "Block", "String Inverter", "MPPT PV No", "Latest Value (A)", "Continuous Days Negative"]
        if df_multiday is None or df_multiday.empty or not day_keys:
            return pd.DataFrame(columns=empty_cols), pd.DataFrame(columns=empty_cols)

        comparison_day_keys = list(day_keys[-3:])
        latest_key = comparison_day_keys[-1]
        latest_col = f"PV_Value_{latest_key}"
        if latest_col not in df_multiday.columns:
            return pd.DataFrame(columns=empty_cols), pd.DataFrame(columns=empty_cols)

        id_cols = [c for c in ["Plot", "Block", "String Inverter", "MPPT PV No"] if c in df_multiday.columns]

        work = df_multiday.copy()
        work[latest_col] = pd.to_numeric(work[latest_col], errors="coerce")
        current_neg = work[work[latest_col] < 0].copy()
        if current_neg.empty:
            return pd.DataFrame(columns=empty_cols), pd.DataFrame(columns=empty_cols)

        prior_keys = list(reversed(comparison_day_keys[:-1]))
        streak = pd.Series(1, index=current_neg.index)
        still_negative = pd.Series(True, index=current_neg.index)
        for dk in prior_keys:
            col = f"PV_Value_{dk}"
            if col not in current_neg.columns:
                break
            vals = pd.to_numeric(current_neg[col], errors="coerce")
            is_neg = vals < 0
            streak = streak + (still_negative & is_neg).astype(int)
            still_negative = still_negative & is_neg

        current_neg["Continuous Days Negative"] = streak
        current_neg = current_neg.rename(columns={latest_col: "Latest Value (A)"})
        current_neg["Latest Value (A)"] = current_neg["Latest Value (A)"].round(2)
        keep_cols = id_cols + ["Latest Value (A)", "Continuous Days Negative"]
        current_neg = current_neg[[c for c in keep_cols if c in current_neg.columns]]
        current_neg = current_neg.sort_values("Continuous Days Negative", ascending=False).reset_index(drop=True)
        continuous_neg = current_neg[current_neg["Continuous Days Negative"] > 1].reset_index(drop=True)
        return current_neg, continuous_neg


    # ============================================================================
    # LOW PERFORMANCE STRINGS  (requirement #6)
    # ============================================================================

    def compute_low_performance_strings(df: pd.DataFrame, threshold: float, ratio: float):
        """Flag working strings that under-perform their own inverter's peers.

        A string counts as "working" once its current exceeds the fault
        threshold, but two working strings can still differ a lot (e.g. one
        string reading ~6A while the rest of the same inverter reads ~10A). That
        kind of gap indicates soiling, shading, or a partial fault that simple
        working/failed counting misses. For every inverter we take the MEDIAN
        current among its currently-working strings as the local baseline, then
        flag any working string whose current falls below `ratio` (e.g. 70%) of
        that baseline. Using the per-inverter median (rather than a plant-wide
        constant) means the check adapts automatically to weather, time of day,
        and inverter sizing.
        """
        pv_columns = get_available_pv_columns(df)
        cols_out = ["Plot", "Block", "String Inverter", "String", "Current (A)",
                    "Inverter Median (A)", "Performance vs Median (%)"]
        if not pv_columns or df.empty:
            return pd.DataFrame(columns=cols_out)

        records = []
        for _, row in df.iterrows():
            working_vals = {}
            for col in pv_columns:
                v = pd.to_numeric(row.get(col), errors="coerce")
                if pd.notna(v) and v > threshold:
                    working_vals[col] = v
            if len(working_vals) < 2:
                continue
            median_val = float(np.median(list(working_vals.values())))
            if median_val <= 0:
                continue
            cutoff = median_val * ratio
            for col, v in working_vals.items():
                if v < cutoff:
                    records.append({
                        "Plot": row.get("Plot"),
                        "Block": row.get("Block"),
                        "String Inverter": row.get("String Inverter"),
                        "String": col,
                        "Current (A)": round(float(v), 2),
                        "Inverter Median (A)": round(median_val, 2),
                        "Performance vs Median (%)": round(float(v) / median_val * 100, 1),
                    })

        if not records:
            return pd.DataFrame(columns=cols_out)
        out = pd.DataFrame(records)[cols_out]
        return out.sort_values("Performance vs Median (%)").reset_index(drop=True)


    @st.cache_data(show_spinner=False)
    def compute_summaries(latest_df: pd.DataFrame, fault_df: pd.DataFrame):
        """Build the inverter/plot/block roll-ups and headline KPIs used across the dashboard."""
        status_cols = [c for c in ["String Inverter", "Plot", "Block", "Total Active Strings",
                                    "Working String Count", "Failed String Count"] if c in latest_df.columns]
        inverter_status = latest_df[status_cols].dropna(subset=["String Inverter"]).copy()
        inverter_status["String Inverter"] = inverter_status["String Inverter"].astype(str).str.strip()
        # processed_days is already de-duplicated by handle_duplicate_inverters, this
        # is a defensive no-op safety net in case an already-clean frame is passed in.
        inverter_status = inverter_status.drop_duplicates(subset=["String Inverter"], keep="first").reset_index(drop=True)
        for col in ["Total Active Strings", "Working String Count", "Failed String Count"]:
            inverter_status[col] = pd.to_numeric(inverter_status[col], errors="coerce").fillna(0).astype(int)

        df_newly_failed = fault_df[fault_df["Fault Type"].astype(str).str.upper().str.strip() == "NEWLY FAILED"].copy() if "Fault Type" in fault_df.columns else pd.DataFrame()
        df_refailed = fault_df[fault_df["Fault Type"].astype(str).str.upper().str.strip() == "RE-FAILED"].copy() if "Fault Type" in fault_df.columns else pd.DataFrame()

        total_inverters = int(inverter_status["String Inverter"].nunique())
        total_active_strings = int(inverter_status["Total Active Strings"].sum())
        working_strings = int(inverter_status["Working String Count"].sum())
        failed_strings = int(inverter_status["Failed String Count"].sum())
        newly_failed_strings = len(df_newly_failed)
        refailed_strings = len(df_refailed)
        availability = round((working_strings / total_active_strings * 100), 2) if total_active_strings else 0.0
        failure_pct = round((failed_strings / total_active_strings * 100), 2) if total_active_strings else 0.0

        def _group_summary(group_cols):
            base = inverter_status.groupby(group_cols, dropna=False).agg(
                Inverters=("String Inverter", "nunique"),
                **{"Total Active Strings": ("Total Active Strings", "sum"),
                   "Working Strings": ("Working String Count", "sum"),
                   "Failed Strings": ("Failed String Count", "sum")}
            ).reset_index()

            nf = (df_newly_failed.groupby(group_cols, dropna=False).size().reset_index(name="Newly Failed Strings")
                  if not df_newly_failed.empty else pd.DataFrame(columns=group_cols + ["Newly Failed Strings"]))
            rf = (df_refailed.groupby(group_cols, dropna=False).size().reset_index(name="Re-Failed Strings")
                  if not df_refailed.empty else pd.DataFrame(columns=group_cols + ["Re-Failed Strings"]))

            out = base.merge(nf, on=group_cols, how="left").merge(rf, on=group_cols, how="left")
            for col in ["Newly Failed Strings", "Re-Failed Strings"]:
                out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype(int)
            out["Availability %"] = (out["Working Strings"] / out["Total Active Strings"].replace(0, pd.NA) * 100).fillna(0).round(2)
            out["Failure %"] = (out["Failed Strings"] / out["Total Active Strings"].replace(0, pd.NA) * 100).fillna(0).round(2)
            for col in ["Inverters", "Total Active Strings", "Working Strings", "Failed Strings"]:
                out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype(int)
            return out.sort_values(["Failed Strings", "Newly Failed Strings"], ascending=[False, False]).reset_index(drop=True)

        plot_summary = _group_summary(["Plot"]) if "Plot" in inverter_status.columns else pd.DataFrame()
        block_summary = _group_summary(["Plot", "Block"]) if {"Plot", "Block"}.issubset(inverter_status.columns) else pd.DataFrame()

        kpis = {
            "total_inverters": total_inverters, "total_active_strings": total_active_strings,
            "working_strings": working_strings, "failed_strings": failed_strings,
            "newly_failed_strings": newly_failed_strings, "refailed_strings": refailed_strings,
            "availability": availability, "failure_pct": failure_pct,
            "highest_failed_plot": (plot_summary.iloc[0]["Plot"] if not plot_summary.empty else "N/A"),
            "highest_failed_plot_count": (int(plot_summary.iloc[0]["Failed Strings"]) if not plot_summary.empty else 0),
            "highest_failed_block_plot": (block_summary.iloc[0]["Plot"] if not block_summary.empty else "N/A"),
            "highest_failed_block": (block_summary.iloc[0]["Block"] if not block_summary.empty else "N/A"),
            "highest_failed_block_count": (int(block_summary.iloc[0]["Failed Strings"]) if not block_summary.empty else 0),
        }

        return inverter_status, plot_summary, block_summary, df_newly_failed, df_refailed, kpis


    def compute_daily_totals(processed_days: dict, day_keys: list, date_mapping: dict):
        """Roll each day's already-deduplicated sheet up into one summary row,
        so the dashboard can show a 3-day (or 1-3 day) trend at a glance."""
        rows = []
        for dk in day_keys:
            df_day = processed_days.get(dk)
            if df_day is None or df_day.empty:
                continue
            total_active = int(pd.to_numeric(df_day.get("Total Active Strings"), errors="coerce").fillna(0).sum())
            working = int(pd.to_numeric(df_day.get("Working String Count"), errors="coerce").fillna(0).sum())
            failed = int(pd.to_numeric(df_day.get("Failed String Count"), errors="coerce").fillna(0).sum())
            inverters = int(df_day["String Inverter"].nunique()) if "String Inverter" in df_day.columns else 0
            availability = round((working / total_active * 100), 2) if total_active else 0.0
            failure_pct = round((failed / total_active * 100), 2) if total_active else 0.0
            date_raw = date_mapping.get(dk, dk)
            try:
                date_display = pd.to_datetime(date_raw).strftime("%d %b %Y")
            except Exception:
                date_display = str(date_raw)
            rows.append({
                "Day": dk, "Date": date_display, "Inverters": inverters,
                "Total Active Strings": total_active, "Working Strings": working,
                "Failed Strings": failed, "Availability %": availability, "Failure %": failure_pct,
            })
        return pd.DataFrame(rows)


    def build_inverter_matrix(processed_days: dict, comparison_day_keys: list, date_mapping: dict):
        """Per-inverter, per-day Working String Count grid (requirement: 'Inverter Matrix').

        One row per inverter, one column per comparison day holding that day's Working
        String Count, plus the inverter's Total Active Strings so each day-cell can be
        colored: green when the inverter is fully working that day, amber for a small
        shortfall, red for a larger one - so a viewer can see a string count go
        19 -> 18 -> 19 (drop, then recovery) at a glance across the day columns.

        Returns (matrix_df, day_col_labels) where day_col_labels are the actual column
        names used for each day (formatted with the date, e.g. "Day1\\n(12 Aug)").
        """
        day_col_labels = []
        frames = []
        for dk in comparison_day_keys:
            df_day = processed_days.get(dk)
            if df_day is None or df_day.empty or "String Inverter" not in df_day.columns:
                continue
            date_raw = date_mapping.get(dk, dk)
            try:
                date_disp = pd.to_datetime(date_raw).strftime("%d %b")
            except Exception:
                date_disp = str(date_raw)
            col_label = f"{dk} ({date_disp})"
            day_col_labels.append(col_label)
            cols = [c for c in ["String Inverter", "Plot", "Block", "Working String Count"] if c in df_day.columns]
            sub = df_day[cols].rename(columns={"Working String Count": col_label})
            frames.append(sub)

        if not frames:
            return pd.DataFrame(), []

        result = frames[0]
        for sub in frames[1:]:
            join_cols = [c for c in ["String Inverter", "Plot", "Block"] if c in result.columns and c in sub.columns]
            result = result.merge(sub, on=join_cols, how="outer")

        latest_key = comparison_day_keys[-1]
        latest_df = processed_days.get(latest_key)
        if latest_df is not None and "Total Active Strings" in latest_df.columns:
            totals = latest_df[["String Inverter", "Total Active Strings"]].drop_duplicates("String Inverter")
            result = result.merge(totals, on="String Inverter", how="left")
        else:
            result["Total Active Strings"] = pd.NA

        for c in day_col_labels + ["Total Active Strings"]:
            if c in result.columns:
                result[c] = pd.to_numeric(result[c], errors="coerce")

        if len(day_col_labels) >= 2:
            result["Change"] = result[day_col_labels[-1]] - result[day_col_labels[0]]

        ordered = [c for c in ["Plot", "Block", "String Inverter", "Total Active Strings"] if c in result.columns]
        ordered += day_col_labels
        if "Change" in result.columns:
            ordered.append("Change")
        result = result[ordered]
        sort_cols = [c for c in ["Plot", "Block", "String Inverter"] if c in result.columns]
        result = result.sort_values(sort_cols).reset_index(drop=True)
        return result, day_col_labels


    # ============================================================================
    # EXCEL REPORT GENERATION  (condensed port of the "Professional Edition" report)
    # ============================================================================

    def _fill(hexcode):
        return PatternFill("solid", fgColor=hexcode.lstrip("#"))


    def _thin_border():
        side = Side(style="thin", color=COLORS["light_grey"].lstrip("#"))
        return Border(left=side, right=side, top=side, bottom=side)


    def _style_title(ws, title, end_col, subtitle=None):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        c = ws.cell(1, 1, title)
        c.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        c.fill = _fill(COLORS["primary_dark"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
            c2 = ws.cell(2, 1, subtitle)
            c2.font = Font(name="Calibri", size=11, color=COLORS["medium_grey"].lstrip("#"))
            c2.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 22


    def _style_header(ws, row, start_col, end_col, color):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
            cell.fill = _fill(color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color=COLORS["light_grey"].lstrip("#")),
                right=Side(style="thin", color=COLORS["light_grey"].lstrip("#")),
                top=Side(style="medium", color=color.lstrip("#")),
                bottom=Side(style="medium", color=color.lstrip("#")),
            )
        ws.row_dimensions[row].height = 30


    def _auto_width(ws, max_width=32):
        for col_cells in ws.columns:
            length = 0
            letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    length = max(length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(max(length + 3, 12), max_width)


    def _write_table(ws, df, start_row, header_color, highlight_map=None, zebra=True):
        """Write a DataFrame as a styled table starting at start_row (1-indexed header row).
        Rows alternate a very light tint (zebra striping) for readability, except in cells
        that already carry a semantic highlight (working/failed/etc)."""
        highlight_map = highlight_map or {}
        zebra_fill = _fill("F7FAFC")
        for col_num, col_name in enumerate(df.columns, start=1):
            ws.cell(start_row, col_num, col_name)
        _style_header(ws, start_row, 1, len(df.columns), header_color)

        for i, row in enumerate(df.itertuples(index=False, name=None)):
            r_off = start_row + 1 + i
            is_even = (i % 2 == 1)
            for col_num, value in enumerate(row, start=1):
                cell = ws.cell(r_off, col_num, value)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _thin_border()
                col_name = df.columns[col_num - 1]
                if col_name in highlight_map:
                    fill_c, font_c = highlight_map[col_name]
                    cell.fill = _fill(fill_c)
                    cell.font = Font(name="Calibri", size=10, bold=True, color=font_c.lstrip("#"))
                elif zebra and is_even:
                    cell.fill = zebra_fill
            ws.row_dimensions[r_off].height = 20
        return start_row + len(df)


    def generate_excel_report(kpis, plot_summary, block_summary, df_newly_failed, df_refailed,
                               inverter_status, comparison_day_keys, date_mapping, threshold,
                               df_negative_current=None, df_low_perf=None, duplicate_ids_by_day=None,
                               excluded_ids_by_day=None, daily_totals=None, inverter_matrix_df=None,
                               matrix_day_cols=None, included_sheets=None):
        """Build the multi-sheet Excel report. `included_sheets` is a set of sheet keys
        to include (see EXPORT_SHEET_OPTIONS below); None means "include everything"."""
        all_keys = {k for k, _ in EXPORT_SHEET_OPTIONS}
        included_sheets = all_keys if included_sheets is None else set(included_sheets)

        wb = Workbook()
        wb.remove(wb.active)
        number_of_days = len(comparison_day_keys)
        latest_key = comparison_day_keys[-1] if comparison_day_keys else None
        latest_date_raw = date_mapping.get(latest_key, latest_key) if latest_key else "N/A"
        try:
            latest_date_display = pd.to_datetime(latest_date_raw).strftime("%d-%b-%Y")
        except Exception:
            latest_date_display = str(latest_date_raw)

        # --- Dashboard sheet -----------------------------------------------
        if "dashboard" in included_sheets:
            ws = wb.create_sheet("Dashboard")
            _style_title(ws, "PNP PLANT OPERATIONAL DASHBOARD", 8,
                         f"Latest Date: {latest_date_display} | {number_of_days}-Day Reference | Threshold <= {threshold} A")

            kpi_cards = [
                ("TOTAL INVERTERS", kpis["total_inverters"], COLORS["primary_blue"]),
                ("ACTIVE STRINGS", kpis["total_active_strings"], COLORS["info_purple"]),
                ("WORKING STRINGS", kpis["working_strings"], COLORS["success_green"]),
                ("FAILED/PENDING STRINGS", kpis["failed_strings"], COLORS["danger_red"]),
                ("NEWLY FAILED", kpis["newly_failed_strings"], COLORS["warning_orange"]),
                ("RE-FAILED", kpis["refailed_strings"], COLORS["danger_red"]),
                ("AVAILABILITY", f"{kpis['availability']:.2f}%", COLORS["success_green"]),
                ("FAILURE RATE", f"{kpis['failure_pct']:.2f}%", COLORS["danger_red"]),
            ]
            for idx, (label, value, color) in enumerate(kpi_cards):
                row = 4 + (idx // 4) * 3
                col = 1 + (idx % 4) * 2
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
                lc = ws.cell(row, col, label)
                lc.font = Font(size=9, bold=True, color="FFFFFF")
                lc.fill = _fill(color)
                lc.alignment = Alignment(horizontal="center")
                vc = ws.cell(row + 1, col, value)
                vc.font = Font(size=16, bold=True, color=color.lstrip("#"))
                vc.alignment = Alignment(horizontal="center")

        highlight = {
            "Working Strings": (COLORS["success_light"], COLORS["success_green"]),
            "Failed/Pending Strings": (COLORS["danger_light"], COLORS["danger_red"]),
            "Newly Failed Strings": (COLORS["warning_light"], COLORS["warning_orange"]),
            "Re-Failed Strings": (COLORS["info_light"], COLORS["info_purple"]),
        }
        if "dashboard" in included_sheets:
            ws.cell(11, 1, "PLOT-WISE CURRENT STATUS").font = Font(size=13, bold=True, color="FFFFFF")
            ws.cell(11, 1).fill = _fill(COLORS["primary_blue"])
            ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=9)
            if not plot_summary.empty:
                _write_table(ws, plot_summary, 12, COLORS["dark_grey"], highlight)
            _auto_width(ws)
            ws.freeze_panes = "A13"

        # --- 3-Day Trend sheet --------------------------------------------
        if "trend" in included_sheets:
            ws_trend = wb.create_sheet("3-Day Trend")
            trend_cols = ["Day", "Date", "Inverters", "Working Strings", "Failed Strings", "Availability %"]
            if isinstance(daily_totals, pd.DataFrame):
                trend_df = daily_totals.copy()
            else:
                trend_df = pd.DataFrame(columns=trend_cols)
            if not trend_df.empty:
                trend_df = trend_df[[c for c in trend_cols if c in trend_df.columns]].copy()
                if "Availability %" in trend_df.columns:
                    trend_df["Availability %"] = pd.to_numeric(trend_df["Availability %"], errors="coerce").round(2)
            _style_title(
                ws_trend,
                "3-DAY STRING AVAILABILITY TREND",
                max(len(trend_df.columns), 1) if not trend_df.empty else len(trend_cols),
                f"Latest Date: {latest_date_display}",
            )
            if not trend_df.empty:
                trend_highlight = {
                    "Working Strings": (COLORS["success_light"], COLORS["success_green"]),
                    "Failed Strings": (COLORS["danger_light"], COLORS["danger_red"]),
                    "Availability %": (COLORS["info_light"], COLORS["info_purple"]),
                }
                last_row = _write_table(ws_trend, trend_df, 4, COLORS["info_purple"], trend_highlight)
                ws_trend.auto_filter.ref = f"A4:{get_column_letter(len(trend_df.columns))}{last_row}"
            _auto_width(ws_trend)
            ws_trend.freeze_panes = "A5"

        # --- Plot Summary sheet ---------------------------------------------
        if "plot_summary" in included_sheets:
            ws_plot = wb.create_sheet("Plot Summary")
            _style_title(ws_plot, "PLOT-WISE STRING PERFORMANCE", max(len(plot_summary.columns), 1),
                         f"Latest Date: {latest_date_display} | Threshold <= {threshold} A")
            if not plot_summary.empty:
                last_row = _write_table(ws_plot, plot_summary, 4, COLORS["primary_blue"], highlight)
                ws_plot.auto_filter.ref = f"A4:{get_column_letter(len(plot_summary.columns))}{last_row}"
                ws_plot.conditional_formatting.add(
                    f"H5:H{last_row}", CellIsRule(operator="lessThan", formula=["90"], fill=_fill(COLORS["warning_light"]))
                )
            _auto_width(ws_plot)
            ws_plot.freeze_panes = "A5"

        # --- Block Summary sheet --------------------------------------------
        if "block_summary" in included_sheets:
            ws_block = wb.create_sheet("Block Summary")
            _style_title(ws_block, "BLOCK-WISE STRING PERFORMANCE", max(len(block_summary.columns), 1),
                         f"Latest Date: {latest_date_display} | Threshold <= {threshold} A")
            if not block_summary.empty:
                last_row = _write_table(ws_block, block_summary, 4, COLORS["info_purple"], highlight)
                ws_block.auto_filter.ref = f"A4:{get_column_letter(len(block_summary.columns))}{last_row}"
            _auto_width(ws_block)
            ws_block.freeze_panes = "A5"

        # --- Inverter Matrix sheet -------------------------------------------
        if "inverter_matrix" in included_sheets and isinstance(inverter_matrix_df, pd.DataFrame) and not inverter_matrix_df.empty:
            ws_matrix = wb.create_sheet("Inverter Matrix")
            _style_title(ws_matrix, "INVERTER MATRIX — WORKING STRINGS BY DAY", max(len(inverter_matrix_df.columns), 1),
                         f"{number_of_days}-Day Reference | Green = fully working, Amber = small shortfall, Red = larger shortfall")
            matrix_highlight = {}
            if "Change" in inverter_matrix_df.columns:
                matrix_highlight["Change"] = (COLORS["info_light"], COLORS["info_purple"])
            last_row = _write_table(ws_matrix, inverter_matrix_df, 4, COLORS["dark_grey"], matrix_highlight, zebra=False)
            # Color each day cell by that row's working/total ratio, matching the dashboard's coding.
            if matrix_day_cols and "Total Active Strings" in inverter_matrix_df.columns:
                for r_off, (_, row) in enumerate(inverter_matrix_df.iterrows(), start=5):
                    total = row.get("Total Active Strings")
                    for day_col in matrix_day_cols:
                        if day_col not in inverter_matrix_df.columns:
                            continue
                        c_idx = list(inverter_matrix_df.columns).index(day_col) + 1
                        val = row.get(day_col)
                        cell = ws_matrix.cell(r_off, c_idx)
                        if pd.isna(val) or pd.isna(total) or total == 0:
                            continue
                        ratio = val / total
                        if ratio >= 1:
                            cell.fill = _fill(COLORS["success_light"])
                            cell.font = Font(color=COLORS["success_green"].lstrip("#"), bold=True)
                        elif ratio >= 0.9:
                            cell.fill = _fill(COLORS["warning_light"])
                            cell.font = Font(color=COLORS["warning_orange"].lstrip("#"), bold=True)
                        else:
                            cell.fill = _fill(COLORS["danger_light"])
                            cell.font = Font(color=COLORS["danger_red"].lstrip("#"), bold=True)
            ws_matrix.auto_filter.ref = f"A4:{get_column_letter(len(inverter_matrix_df.columns))}{last_row}"
            _auto_width(ws_matrix)
            ws_matrix.freeze_panes = "D5"

        # --- Failed (Newly Failed) sheet ----------------------------------------------
        fault_status = inverter_status[["String Inverter", "Working String Count", "Failed String Count"]].rename(
            columns={"Working String Count": "Current Working Strings Count", "Failed String Count": "Current Failed Strings"}
        )
        if "failed" in included_sheets:
            ws_fault = wb.create_sheet("Failed Strings")
            # df_newly_failed / df_refailed may already carry these columns (the main app
            # pre-merges them for the on-screen tabs) - only merge here if they're missing,
            # otherwise pandas would suffix the duplicate column names and break lookups below.
            already_has_status_cols = (not df_newly_failed.empty) and "Current Working Strings Count" in df_newly_failed.columns
            df_fault_export = (
                df_newly_failed if already_has_status_cols
                else (df_newly_failed.merge(fault_status, on="String Inverter", how="left") if not df_newly_failed.empty else df_newly_failed)
            )
            _style_title(ws_fault, "NEWLY FAILED PV STRINGS", max(len(df_fault_export.columns), 1) if not df_fault_export.empty else 8,
                         f"Threshold <= {threshold} A")
            if not df_fault_export.empty:
                for col in ["Current Working Strings Count", "Current Failed Strings"]:
                    df_fault_export[col] = pd.to_numeric(df_fault_export[col], errors="coerce").fillna(0).astype(int)
                fault_highlight = dict(highlight)
                fault_highlight["Current Working Strings Count"] = (COLORS["success_light"], COLORS["success_green"])
                fault_highlight["Current Failed Strings"] = (COLORS["danger_light"], COLORS["danger_red"])
                fault_highlight["Fault Type"] = (COLORS["warning_light"], COLORS["warning_orange"])
                last_row = _write_table(ws_fault, df_fault_export, 4, COLORS["danger_red"], fault_highlight)
                ws_fault.auto_filter.ref = f"A4:{get_column_letter(len(df_fault_export.columns))}{last_row}"
            _auto_width(ws_fault)
            ws_fault.freeze_panes = "A5"

        # --- Re-Failed sheet ---------------------------------------------------
        if "refailed" in included_sheets:
            ws_re = wb.create_sheet("Re-Failed Strings")
            already_has_status_cols_re = (not df_refailed.empty) and "Current Working Strings Count" in df_refailed.columns
            df_re_export = (
                df_refailed if already_has_status_cols_re
                else (df_refailed.merge(fault_status, on="String Inverter", how="left") if not df_refailed.empty else df_refailed)
            )
            _style_title(ws_re, "RE-FAILED PV STRINGS", max(len(df_re_export.columns), 1) if not df_re_export.empty else 8,
                         f"Latest Date: {latest_date_display}")
            if not df_re_export.empty:
                for col in ["Current Working Strings Count", "Current Failed Strings"]:
                    df_re_export[col] = pd.to_numeric(df_re_export[col], errors="coerce").fillna(0).astype(int)
                last_row = _write_table(ws_re, df_re_export, 4, COLORS["danger_red"], highlight)
                ws_re.auto_filter.ref = f"A4:{get_column_letter(len(df_re_export.columns))}{last_row}"
            _auto_width(ws_re)
            ws_re.freeze_panes = "A5"

        # --- Negative Values sheet ----------------------------------------------
        if "negative" in included_sheets and df_negative_current is not None:
            ws_neg = wb.create_sheet("Negative Values")
            _style_title(ws_neg, "NEGATIVE STRING READINGS", max(len(df_negative_current.columns), 1) if not df_negative_current.empty else 8,
                         f"Latest Date: {latest_date_display}")
            if not df_negative_current.empty:
                neg_highlight = {"Continuous Days Negative": (COLORS["warning_light"], COLORS["warning_orange"]),
                                  "Latest Value (A)": (COLORS["danger_light"], COLORS["danger_red"])}
                last_row = _write_table(ws_neg, df_negative_current, 4, COLORS["danger_red"], neg_highlight)
                ws_neg.auto_filter.ref = f"A4:{get_column_letter(len(df_negative_current.columns))}{last_row}"
            _auto_width(ws_neg)
            ws_neg.freeze_panes = "A5"

        # --- Low Performance sheet ----------------------------------------------
        if "low_perf" in included_sheets and df_low_perf is not None:
            ws_lp = wb.create_sheet("Low Performance")
            _style_title(ws_lp, "LOW PERFORMING STRINGS", max(len(df_low_perf.columns), 1) if not df_low_perf.empty else 8,
                         f"Latest Date: {latest_date_display}")
            if not df_low_perf.empty:
                lp_highlight = {"Performance vs Median (%)": (COLORS["warning_light"], COLORS["warning_orange"])}
                last_row = _write_table(ws_lp, df_low_perf, 4, COLORS["warning_orange"], lp_highlight)
                ws_lp.auto_filter.ref = f"A4:{get_column_letter(len(df_low_perf.columns))}{last_row}"
            _auto_width(ws_lp)
            ws_lp.freeze_panes = "A5"

        # --- Duplicate / Excluded Inverters sheet --------------------------------
        if "duplicates" in included_sheets and (duplicate_ids_by_day or excluded_ids_by_day):
            ws_dup = wb.create_sheet("Duplicate Inverters")
            rows = []
            for dk, ids in (duplicate_ids_by_day or {}).items():
                for i in ids:
                    rows.append({"Day": dk, "Inverter ID": i, "Status": "Duplicate (included in calc)"})
            for dk, ids in (excluded_ids_by_day or {}).items():
                for i in ids:
                    rows.append({"Day": dk, "Inverter ID": i, "Status": "Excluded (>= threshold duplicates)"})
            df_dup = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Day", "Inverter ID", "Status"])
            _style_title(ws_dup, "DUPLICATE / EXCLUDED INVERTER IDs", max(len(df_dup.columns), 1))
            dup_highlight = {"Status": (COLORS["warning_light"], COLORS["warning_orange"])}
            last_row = _write_table(ws_dup, df_dup, 4, COLORS["info_purple"], dup_highlight)
            ws_dup.auto_filter.ref = f"A4:{get_column_letter(len(df_dup.columns))}{last_row}"
            _auto_width(ws_dup)
            ws_dup.freeze_panes = "A5"

        if len(wb.sheetnames) == 0:
            # Every sheet was deselected - still return a valid, openable workbook.
            ws_empty = wb.create_sheet("Report")
            _style_title(ws_empty, "NO SHEETS SELECTED", 4, "Choose at least one report on the Export tab and regenerate.")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


    # ============================================================================
    # SIDEBAR - UPLOAD (BY DAY) + CONFIG
    # ============================================================================

    with st.sidebar:
        st.markdown('<div class="sidebar-brand"><i class="fa-solid fa-solar-panel"></i>PV Plant Monitor</div>', unsafe_allow_html=True)
        st.caption("Multi-day SCADA string fault analysis")
        st.markdown("---")

        st.markdown("**Data source**")
        data_source_mode = st.radio(
            "Data source",
            options=["Saved snapshots", "Upload files"],
            index=0,
            key="analysis_data_source_mode",
            label_visibility="collapsed",
            help="Saved snapshots reuse the data already preprocessed on the Dashboard tab - no re-upload needed. Upload files lets you analyze workbooks directly with fully custom settings below.",
        )
        st.caption("Each day is processed independently, then compared to detect faults.")

        day1_file = day2_file = day3_file = None
        saved_dates_by_day = {}

        if data_source_mode == "Saved snapshots":
            available_dates = sorted(storage1.get_available_snapshot_dates())
            if not available_dates:
                st.warning("No preprocessed snapshots are available yet. Upload a file on the Dashboard tab first, or switch to \"Upload files\" here.")
            else:
                selected_dates = st.multiselect(
                    "Select up to 3 snapshot dates to compare",
                    options=available_dates,
                    default=available_dates[-1:],
                    max_selections=3,
                    key="analysis_saved_dates",
                    help="Pick 1-3 dates. They're automatically ordered oldest \u2192 latest regardless of pick order.",
                )
                selected_dates = sorted(selected_dates)
                for label, date_str in zip(DAY_LABELS, selected_dates):
                    saved_dates_by_day[label] = date_str
                if selected_dates:
                    st.caption("Using: " + " \u2192 ".join(selected_dates))
        else:
            day1_file = st.file_uploader("Day 1 (oldest)", type=["xlsx"], accept_multiple_files=False, key="day1_upload")
            day2_file = st.file_uploader("Day 2", type=["xlsx"], accept_multiple_files=False, key="day2_upload")
            day3_file = st.file_uploader("Day 3 (latest, optional)", type=["xlsx"], accept_multiple_files=False, key="day3_upload")

        with st.expander("Analysis Settings", icon=":material/tune:", expanded=False):
            threshold = st.number_input("Working current threshold (A)", min_value=0.0, max_value=5.0,
                                         value=DEFAULT_THRESHOLD, step=0.1,
                                         help="Used for fault comparison across days in both data source modes.")
            default_active = st.number_input("Default active strings / block", min_value=1, max_value=100,
                                              value=DEFAULT_TOTAL_ACTIVE_STRINGS, step=1,
                                              disabled=(data_source_mode == "Saved snapshots"))
            dup_exclude_threshold = st.number_input(
                "Exclude inverter if duplicate rows >=", min_value=2, max_value=20,
                value=DEFAULT_DUPLICATE_EXCLUDE_THRESHOLD, step=1,
                help="Inverter IDs repeated at least this many times in one day's sheet are treated as bad data and dropped. Fewer repeats are kept in calculations but flagged as duplicates.",
            )
            low_perf_ratio_pct = st.slider(
                "Low-performance cutoff (% of inverter median)", min_value=30, max_value=95,
                value=int(DEFAULT_LOW_PERF_RATIO * 100), step=5,
                help="A working string below this % of its own inverter's median current is flagged as low performing.",
            )
            if data_source_mode == "Saved snapshots":
                st.caption("Active-string overrides are already baked into saved snapshots from the Dashboard. Duplicate-ID detection above still runs on snapshot data. Switch to \"Upload files\" to control overrides here.")
            low_perf_ratio = low_perf_ratio_pct / 100.0

        with st.expander("Active-String Overrides", icon=":material/table_rows:", expanded=False):
            st.caption("Blocks with fewer active strings than the default")
            if data_source_mode == "Saved snapshots":
                st.caption("Not used in Saved snapshots mode - already applied on the Dashboard.")
            overrides_df = st.data_editor(
                pd.DataFrame(DEFAULT_OVERRIDES_ROWS), num_rows="dynamic", use_container_width=True,
                key="overrides_editor",
                disabled=(data_source_mode == "Saved snapshots"),
            )

        st.markdown("---")
        process_clicked = st.button("Reprocess Now", icon=":material/refresh:", type="primary", use_container_width=True)
        st.caption("The dashboard already refreshes automatically whenever a file or setting changes — use this only to force a manual re-run.")

    # Build overrides dict {PLOT: {BLOCK: count}}
    overrides_dict = {}
    for _, r in overrides_df.dropna(subset=["Plot", "Block"]).iterrows():
        p, b = normalize_text(r["Plot"]), normalize_text(r["Block"])
        try:
            overrides_dict.setdefault(p, {})[b] = int(r["Active Strings"])
        except (ValueError, TypeError):
            continue
    overrides_json = json.dumps(overrides_dict, sort_keys=True)

    # ============================================================================
    # PIPELINE — runs automatically whenever a file or setting changes
    # ============================================================================

    # ============================================================================
    # PIPELINE — runs automatically whenever a file/date-selection or setting changes
    # ============================================================================

    if data_source_mode == "Saved snapshots":
        day_files = [(label, date_str) for label, date_str in saved_dates_by_day.items()]
    else:
        day_files = []
        for label, f in zip(DAY_LABELS, [day1_file, day2_file, day3_file]):
            if f is not None:
                day_files.append((label, f))

    def files_signature(mode, day_files, overrides_json, threshold, default_active, dup_exclude_threshold, low_perf_ratio):
        h = hashlib.sha256()
        h.update(mode.encode())
        for day_key, payload in day_files:
            h.update(day_key.encode())
            if mode == "Saved snapshots":
                # payload is a snapshot date string; also fold in the underlying
                # upload's hash/timestamp so re-uploading data for that date busts the cache.
                entry = storage1.get_upload_for_date(payload)
                h.update(payload.encode())
                if entry:
                    h.update(str(entry.get("file_hash", "")).encode())
                    h.update(str(entry.get("upload_timestamp", "")).encode())
            else:
                h.update(payload.name.encode())
                h.update(payload.getvalue())
        h.update(overrides_json.encode())
        h.update(str(threshold).encode())
        h.update(str(default_active).encode())
        h.update(str(dup_exclude_threshold).encode())
        h.update(str(low_perf_ratio).encode())
        return h.hexdigest()


    if day_files:
        sig = files_signature(data_source_mode, day_files, overrides_json, threshold, default_active, dup_exclude_threshold, low_perf_ratio)
        needs_run = process_clicked or (st.session_state["last_signature"] != sig)

        if needs_run:
            spinner_text = "Loading saved snapshots..." if data_source_mode == "Saved snapshots" else "Parsing workbooks & computing string metrics..."
            with st.spinner(spinner_text):
                processed_days, file_names_by_day, missing_any = {}, {}, []
                duplicate_rows_by_day, excluded_rows_by_day = {}, {}

                if data_source_mode == "Saved snapshots":
                    for day_key, snapshot_date in day_files:
                        dfs, entry = storage1.get_processed_dataframes_for_date(snapshot_date)
                        if not dfs:
                            missing_any.append(snapshot_date)
                            continue
                        target_sheet = "Sheet1" if "Sheet1" in dfs else next(iter(dfs))
                        raw_df = dfs[target_sheet]

                        # Snapshots are pre-processed by the Dashboard pipeline, but that
                        # pipeline doesn't currently exclude/flag duplicate inverter ID rows,
                        # so duplicates can still be present here. Run the same detection
                        # used for direct uploads instead of assuming it's already clean.
                        df_calc, df_dup_rows, df_excluded, dup_ids, excluded_ids = handle_duplicate_inverters(
                            raw_df, exclude_threshold=int(dup_exclude_threshold)
                        )
                        processed_days[day_key] = df_calc
                        duplicate_rows_by_day[day_key] = df_dup_rows
                        excluded_rows_by_day[day_key] = df_excluded
                        file_names_by_day[day_key] = (entry or {}).get("original_filename", snapshot_date)

                    day_keys = [dk for dk, _ in day_files if dk in processed_days]
                    date_mapping = {dk: date_str for dk, date_str in day_files if dk in processed_days}
                else:
                    for day_key, f in day_files:
                        sheets, missing_log = process_scada_excel_bytes(f.getvalue(), f.name, overrides_json, int(default_active), float(threshold))
                        target_sheet = "Sheet1" if "Sheet1" in sheets else (next(iter(sheets)) if sheets else None)
                        if not target_sheet:
                            missing_any.append(f.name)
                            continue

                        raw_df = sheets[target_sheet]
                        df_calc, df_dup_rows, df_excluded, dup_ids, excluded_ids = handle_duplicate_inverters(
                            raw_df, exclude_threshold=int(dup_exclude_threshold)
                        )
                        processed_days[day_key] = df_calc
                        duplicate_rows_by_day[day_key] = df_dup_rows
                        excluded_rows_by_day[day_key] = df_excluded
                        file_names_by_day[day_key] = f.name

                    day_keys = [dk for dk, _ in day_files if dk in processed_days]
                    date_mapping = {dk: (extract_date_from_filename(file_names_by_day[dk]) or dk) for dk in day_keys}

                st.session_state["processed_days"] = processed_days
                st.session_state["duplicate_rows_by_day"] = duplicate_rows_by_day
                st.session_state["excluded_rows_by_day"] = excluded_rows_by_day
                st.session_state["file_names_by_day"] = file_names_by_day
                st.session_state["day_keys"] = day_keys
                st.session_state["date_mapping"] = date_mapping
                st.session_state["last_signature"] = sig
                st.session_state["excel_report_bytes"] = None

                if len(day_keys) >= 2:
                    multiday_df = build_multiday_comparison({k: v for k, v in processed_days.items()}, tuple(day_keys))
                    fault_df = compute_fault_events(multiday_df, tuple(day_keys), float(threshold))
                    st.session_state["multiday_df"] = multiday_df
                    st.session_state["fault_df"] = fault_df
                else:
                    st.session_state["multiday_df"] = None
                    st.session_state["fault_df"] = None

                if missing_any:
                    if data_source_mode == "Saved snapshots":
                        st.warning(f"No preprocessed data could be loaded for: {', '.join(missing_any)}")
                    else:
                        st.warning(f"Could not detect a valid header/inverter column in: {', '.join(missing_any)}")
    else:
        st.session_state["last_signature"] = None


    # ============================================================================
    # MAIN AREA
    # ============================================================================

    st.markdown(
        """
        <div class="hero-banner">
            <h1><i class="fa-solid fa-solar-panel"></i>PV Plant String-Level Monitoring
            <span class="live-pill"><i class="fa-solid fa-circle"></i>LIVE</span></h1>
            <p>Day-by-day SCADA upload &bull; inverter &amp; string availability &bull; automated fault, negative-value &amp; low-performance detection</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    day_keys = st.session_state["day_keys"]
    processed_days = st.session_state["processed_days"]
    duplicate_rows_by_day = st.session_state["duplicate_rows_by_day"]
    excluded_rows_by_day = st.session_state["excluded_rows_by_day"]

    if not day_keys or len(day_keys) < 2:
        st.info(
            "Upload **Day 1** and **Day 2** (Day 3 optional) dated SCADA Excel exports in the sidebar. "
            "The dashboard builds itself automatically as soon as two days are available — no button needed. "
            "At least two days are required to detect newly-failed strings.",
            icon=":material/upload_file:",
        )
        st.stop()

    multiday_df = st.session_state["multiday_df"]
    fault_df = st.session_state["fault_df"]
    date_mapping = st.session_state["date_mapping"]
    comparison_day_keys = day_keys[-3:]
    latest_key = comparison_day_keys[-1]
    latest_df = processed_days[latest_key]

    inverter_status, plot_summary, block_summary, df_newly_failed, df_refailed, kpis = compute_summaries(latest_df, fault_df)

    # Attach "Current Working Strings Count" to Failed / Re-Failed tables for cross-verification (req #5)
    fault_status_lookup = inverter_status[["String Inverter", "Working String Count", "Failed String Count"]].rename(
        columns={"Working String Count": "Current Working Strings Count", "Failed String Count": "Current Failed Strings"}
    )
    df_newly_failed_ui = (df_newly_failed.merge(fault_status_lookup, on="String Inverter", how="left")
                           if not df_newly_failed.empty else df_newly_failed)
    df_refailed_ui = (df_refailed.merge(fault_status_lookup, on="String Inverter", how="left")
                       if not df_refailed.empty else df_refailed)

    # Negative values (req #4)
    df_negative_current, df_negative_continuous = compute_negative_events(multiday_df, day_keys)

    # Low performance strings (req #6)
    df_low_perf = compute_low_performance_strings(latest_df, float(threshold), float(low_perf_ratio))

    # 3-day (or however many days are loaded) roll-up used for the Dashboard trend section + Excel export
    daily_totals = compute_daily_totals(processed_days, comparison_day_keys, date_mapping)

    # Per-inverter, per-day Working String Count grid (Inverter Matrix tab)
    inverter_matrix_df, matrix_day_cols = build_inverter_matrix(processed_days, comparison_day_keys, date_mapping)

    # Duplicate inverter roll-up across all uploaded days (req #3)
    duplicate_ids_by_day = {dk: sorted(df["String Inverter"].astype(str).str.strip().unique().tolist())
                             for dk, df in duplicate_rows_by_day.items() if not df.empty}
    excluded_ids_by_day = {dk: sorted(df["String Inverter"].astype(str).str.strip().unique().tolist())
                            for dk, df in excluded_rows_by_day.items() if not df.empty}
    total_duplicate_ids = len({i for ids in duplicate_ids_by_day.values() for i in ids})
    total_excluded_ids = len({i for ids in excluded_ids_by_day.values() for i in ids})

    try:
        latest_date_display = pd.to_datetime(date_mapping.get(latest_key, latest_key)).strftime("%d %b %Y")
    except Exception:
        latest_date_display = str(date_mapping.get(latest_key, latest_key))

    st.markdown(
        f"""<div class="meta-caption"><i class="fa-regular fa-calendar"></i>Latest data:
        <strong>{latest_date_display}</strong> &bull; {len(comparison_day_keys)}-day comparison window &bull;
        Fault threshold &le; {threshold} A &bull; Files: {', '.join(st.session_state['file_names_by_day'].values())}</div>""",
        unsafe_allow_html=True,
    )

    if total_duplicate_ids or total_excluded_ids:
        st.markdown(
            f"""<div class="meta-caption"><i class="fa-solid fa-clone"></i>
            Duplicate inverter IDs kept in calculations: <strong>{total_duplicate_ids}</strong> &bull;
            Inverter IDs excluded (&ge;{int(dup_exclude_threshold)} duplicate rows): <strong>{total_excluded_ids}</strong></div>""",
            unsafe_allow_html=True,
        )

    day_tab_labels = [f":material/calendar_today: {dk}" for dk in comparison_day_keys]
    tab_labels = (
        [":material/dashboard: Dashboard", ":material/pin_drop: Plot Summary", ":material/grid_view: Block Summary",
         ":material/apps: Inverter Matrix"]
        + day_tab_labels
        + [":material/report: Failed Strings", ":material/history: Re-Failed Strings",
           ":material/trending_down: Negative Values", ":material/speed: Low Performance",
           ":material/file_download: Export"]
    )
    tabs = st.tabs(tab_labels)
    tab_dash, tab_plot, tab_block, tab_matrix = tabs[0], tabs[1], tabs[2], tabs[3]
    day_tabs = tabs[4:4 + len(comparison_day_keys)]
    tab_failed, tab_refailed, tab_negative, tab_lowperf, tab_export = tabs[4 + len(comparison_day_keys):]

    # ---------------------------------------------------------------- Dashboard
    with tab_dash:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Inverters", f"{kpis['total_inverters']:,}")
        c2.metric("Active Strings", f"{kpis['total_active_strings']:,}")
        c3.metric("Working Strings", f"{kpis['working_strings']:,}")
        c4.metric("Failed Strings", f"{kpis['failed_strings']:,}", delta=None)

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Newly Failed", f"{kpis['newly_failed_strings']:,}")
        c6.metric("Re-Failed", f"{kpis['refailed_strings']:,}")
        c7.metric("Availability", f"{kpis['availability']:.2f}%")
        c8.metric("Failure Rate", f"{kpis['failure_pct']:.2f}%")

        c9, c10, c11, c12 = st.columns(4)
        c9.metric("Negative Strings (current)", f"{len(df_negative_current):,}")
        c10.metric("Continuous Negative", f"{len(df_negative_continuous):,}")
        c11.metric("Low Performing Strings", f"{len(df_low_perf):,}")
        c12.metric("Duplicate Inverter IDs", f"{total_duplicate_ids:,}", help="Kept in calculations, flagged for review")

        st.markdown("<br>", unsafe_allow_html=True)

        # --------------------------------------------------------- 3-Day Summary
        st.markdown('<div class="section-title"><i class="fa-solid fa-calendar-week"></i>3-Day Summary</div>', unsafe_allow_html=True)
        if daily_totals.empty:
            st.info("No day-level data available yet.")
        else:
            trend_left, trend_right = st.columns([1.3, 1])

            with trend_left:
                trend_fig = go.Figure()
                trend_fig.add_trace(go.Scatter(
                    x=daily_totals["Day"], y=daily_totals["Working Strings"], name="Working Strings",
                    mode="lines+markers", line=dict(color=COLORS["success_green"], width=3),
                    marker=dict(size=9), fill="tozeroy", fillcolor="rgba(56,161,105,0.12)",
                ))
                trend_fig.add_trace(go.Scatter(
                    x=daily_totals["Day"], y=daily_totals["Failed Strings"], name="Failed Strings",
                    mode="lines+markers", line=dict(color=COLORS["danger_red"], width=3),
                    marker=dict(size=9), fill="tozeroy", fillcolor="rgba(229,62,62,0.12)",
                ))
                trend_fig.add_trace(go.Scatter(
                    x=daily_totals["Day"], y=daily_totals["Availability %"], name="Availability %",
                    mode="lines+markers", line=dict(color=COLORS["primary_blue"], width=2, dash="dot"),
                    marker=dict(size=7, symbol="diamond"), yaxis="y2",
                ))
                trend_fig.update_layout(
                    height=360, margin=dict(t=30, b=10, l=10, r=10),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                    xaxis=dict(title=None),
                    yaxis=dict(title="Strings"),
                    yaxis2=dict(title="Availability %", overlaying="y", side="right", range=[0, 100], showgrid=False),
                    hovermode="x unified",
                )
                st.plotly_chart(trend_fig, use_container_width=True)

            with trend_right:
                summary_view = daily_totals[["Day", "Date", "Inverters", "Working Strings", "Failed Strings", "Availability %"]]
                styled_daily = summary_view.style.background_gradient(subset=["Availability %"], cmap="RdYlGn", vmin=0, vmax=100) \
                    .format({"Availability %": "{:.2f}%"})
                st.dataframe(styled_daily, use_container_width=True, hide_index=True)
                if len(daily_totals) >= 2:
                    delta_working = int(daily_totals.iloc[-1]["Working Strings"] - daily_totals.iloc[0]["Working Strings"])
                    delta_avail = round(daily_totals.iloc[-1]["Availability %"] - daily_totals.iloc[0]["Availability %"], 2)
                    st.metric(
                        f"Change: {daily_totals.iloc[0]['Day']} \u2192 {daily_totals.iloc[-1]['Day']}",
                        f"{delta_working:+,} working strings", f"{delta_avail:+.2f} pts availability",
                    )

        left, right = st.columns([1.4, 1])

        with left:
            st.markdown('<div class="section-title"><i class="fa-solid fa-chart-column"></i>Failed Strings by Plot</div>', unsafe_allow_html=True)
            if not plot_summary.empty:
                fig = go.Figure()
                fig.add_bar(x=plot_summary["Plot"], y=plot_summary["Working Strings"], name="Working",
                            marker_color=COLORS["success_green"])
                fig.add_bar(x=plot_summary["Plot"], y=plot_summary["Failed Strings"], name="Failed",
                            marker_color=COLORS["danger_red"])
                fig.update_layout(barmode="stack", height=360, margin=dict(t=10, b=10, l=10, r=10),
                                   legend=dict(orientation="h", yanchor="bottom", y=1.02))
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No plot-level data available.")

        with right:
            st.markdown('<div class="section-title"><i class="fa-solid fa-chart-pie"></i>Plant Status Mix</div>', unsafe_allow_html=True)
            donut = go.Figure(data=[go.Pie(
                labels=["Working", "Failed"],
                values=[kpis["working_strings"], kpis["failed_strings"]],
                hole=0.55,
                marker_colors=[COLORS["success_green"], COLORS["danger_red"]],
            )])
            donut.update_layout(height=360, margin=dict(t=10, b=10, l=10, r=10), showlegend=True)
            st.plotly_chart(donut, use_container_width=True)

        st.markdown('<div class="section-title"><i class="fa-solid fa-triangle-exclamation"></i>Priority Attention</div>', unsafe_allow_html=True)
        p1, p2, p3, p4 = st.columns(4)
        for col, icon, label, value, note in [
            (p1, "fa-industry", "Highest Failed Plot", kpis["highest_failed_plot"], f"{kpis['highest_failed_plot_count']:,} failed strings"),
            (p2, "fa-cubes", "Highest Failed Block", f"{kpis['highest_failed_block_plot']} / {kpis['highest_failed_block']}", f"{kpis['highest_failed_block_count']:,} failed strings"),
            (p3, "fa-bolt", "Newly Failed Today", f"{kpis['newly_failed_strings']:,}", "Current faults — investigate"),
            (p4, "fa-percent", "Failure Rate", f"{kpis['failure_pct']:.2f}%", "Plant level"),
        ]:
            col.markdown(
                f'<div class="priority-card"><div class="label"><i class="fa-solid {icon}"></i>{label}</div>'
                f'<div class="value">{value}</div><div class="note">{note}</div></div>',
                unsafe_allow_html=True,
            )

        if duplicate_ids_by_day or excluded_ids_by_day:
            with st.expander("Duplicate Inverter Data Quality", icon=":material/content_copy:", expanded=False):
                st.caption(
                    f"Inverter IDs repeated 2–{int(dup_exclude_threshold) - 1} times are kept in calculations "
                    f"(first reading used) but flagged below. IDs repeated {int(dup_exclude_threshold)}+ times are "
                    "excluded from calculations entirely."
                )
                dcol1, dcol2 = st.columns(2)
                with dcol1:
                    st.markdown("**Duplicate — kept in calculations**")
                    for dk in comparison_day_keys:
                        ids = duplicate_ids_by_day.get(dk, [])
                        if ids:
                            st.markdown(f"*{dk}:* " + ", ".join(ids))
                    if not any(duplicate_ids_by_day.get(dk) for dk in comparison_day_keys):
                        st.caption("None")
                with dcol2:
                    st.markdown("**Excluded — too many duplicate rows**")
                    for dk in comparison_day_keys:
                        ids = excluded_ids_by_day.get(dk, [])
                        if ids:
                            st.markdown(f"*{dk}:* " + ", ".join(ids))
                    if not any(excluded_ids_by_day.get(dk) for dk in comparison_day_keys):
                        st.caption("None")

    # ------------------------------------------------------------- Plot Summary
    with tab_plot:
        st.markdown('<div class="section-title"><i class="fa-solid fa-location-dot"></i>Plot-wise String Performance</div>', unsafe_allow_html=True)
        if plot_summary.empty:
            st.info("No data.")
        else:
            styled = plot_summary.style.background_gradient(subset=["Availability %"], cmap="RdYlGn", vmin=0, vmax=100) \
                .background_gradient(subset=["Failed Strings"], cmap="Reds") \
                .format({"Availability %": "{:.2f}%", "Failure %": "{:.2f}%"})
            st.dataframe(styled, use_container_width=True, hide_index=True)
            fig = px.bar(plot_summary, x="Plot", y="Availability %", color="Availability %",
                         color_continuous_scale="RdYlGn", range_color=[0, 100], height=340)
            st.plotly_chart(fig, use_container_width=True)

    # ------------------------------------------------------------ Block Summary
    with tab_block:
        st.markdown('<div class="section-title"><i class="fa-solid fa-table-cells"></i>Block-wise String Performance</div>', unsafe_allow_html=True)
        if block_summary.empty:
            st.info("No data.")
        else:
            plots_available = ["All"] + sorted(block_summary["Plot"].astype(str).unique().tolist())
            chosen_plot = st.selectbox("Filter by Plot", plots_available, key="block_plot_filter")
            view = block_summary if chosen_plot == "All" else block_summary[block_summary["Plot"] == chosen_plot]
            styled = view.style.background_gradient(subset=["Availability %"], cmap="RdYlGn", vmin=0, vmax=100) \
                .background_gradient(subset=["Failed Strings"], cmap="Reds") \
                .format({"Availability %": "{:.2f}%", "Failure %": "{:.2f}%"})
            st.dataframe(styled, use_container_width=True, hide_index=True)

    # ------------------------------------------------------------- Inverter Matrix
    with tab_matrix:
        st.markdown('<div class="section-title"><i class="fa-solid fa-table-list"></i>Inverter Matrix — Working Strings by Day</div>', unsafe_allow_html=True)
        st.caption(
            "Each day's cell shows that inverter's Working String Count. "
            "Green = fully working that day (matches Total Active Strings), amber = a small "
            "shortfall, red = a larger shortfall — so a drop and later recovery (e.g. 19 → 18 → 19) "
            "is visible at a glance across the day columns."
        )
        if inverter_matrix_df.empty:
            st.info("No inverter-level data available yet.")
        else:
            im1, im2, im3, im4 = st.columns(4)
            im1.metric("Inverters", f"{inverter_matrix_df['String Inverter'].nunique():,}" if "String Inverter" in inverter_matrix_df else "0")
            if "Change" in inverter_matrix_df.columns:
                improved = int((inverter_matrix_df["Change"] > 0).sum())
                declined = int((inverter_matrix_df["Change"] < 0).sum())
                stable = int((inverter_matrix_df["Change"] == 0).sum())
                im2.metric("Improved", f"{improved:,}", help=f"{matrix_day_cols[0]} \u2192 {matrix_day_cols[-1]}")
                im3.metric("Declined", f"{declined:,}", help=f"{matrix_day_cols[0]} \u2192 {matrix_day_cols[-1]}")
                im4.metric("Stable", f"{stable:,}")
            st.markdown("<br>", unsafe_allow_html=True)

            im_c1, im_c2, im_c3 = st.columns(3)
            plots_im = ["All"] + sorted(inverter_matrix_df["Plot"].astype(str).unique().tolist()) if "Plot" in inverter_matrix_df else ["All"]
            chosen_plot_im = im_c1.selectbox("Plot", plots_im, key="matrix_plot_filter")
            search_im = im_c2.text_input("Search inverter ID", key="matrix_search")
            only_changed = im_c3.checkbox("Only show inverters with a change", value=False, key="matrix_only_changed")

            view = inverter_matrix_df.copy()
            if chosen_plot_im != "All":
                view = view[view["Plot"] == chosen_plot_im]
            if search_im and "String Inverter" in view.columns:
                view = view[view["String Inverter"].astype(str).str.contains(search_im, case=False, na=False)]
            if only_changed and "Change" in view.columns:
                view = view[view["Change"] != 0]

            def _matrix_row_style(row):
                styles = [""] * len(row)
                total = row.get("Total Active Strings")
                for i, col in enumerate(row.index):
                    if col in matrix_day_cols:
                        val = row[col]
                        if pd.isna(val) or pd.isna(total) or total == 0:
                            continue
                        ratio = val / total
                        if ratio >= 1:
                            styles[i] = f"background-color: {COLORS['success_light']}; color: {COLORS['success_green']}; font-weight: 700;"
                        elif ratio >= 0.9:
                            styles[i] = f"background-color: {COLORS['warning_light']}; color: {COLORS['warning_orange']}; font-weight: 700;"
                        else:
                            styles[i] = f"background-color: {COLORS['danger_light']}; color: {COLORS['danger_red']}; font-weight: 700;"
                    elif col == "Change" and pd.notna(row.get("Change")):
                        if row["Change"] > 0:
                            styles[i] = f"background-color: {COLORS['success_light']}; color: {COLORS['success_green']}; font-weight: 700;"
                        elif row["Change"] < 0:
                            styles[i] = f"background-color: {COLORS['danger_light']}; color: {COLORS['danger_red']}; font-weight: 700;"
                return styles

            styled_matrix = view.style.apply(_matrix_row_style, axis=1)
            st.dataframe(styled_matrix, use_container_width=True, hide_index=True)
            st.download_button(
                "Download CSV", view.to_csv(index=False).encode(),
                "inverter_matrix.csv", "text/csv", icon=":material/download:", key="dl_matrix",
            )

    # ------------------------------------------------------------------ Day tabs
    for dk, tab in zip(comparison_day_keys, day_tabs):
        with tab:
            st.markdown(
                f'<div class="section-title"><i class="fa-solid fa-database"></i>{dk} Raw Processed Data'
                f' <span class="meta-caption">({date_mapping.get(dk, dk)})</span></div>',
                unsafe_allow_html=True,
            )
            st.caption(f"Source file: {st.session_state['file_names_by_day'].get(dk, 'N/A')}")
            st.dataframe(processed_days[dk], use_container_width=True, hide_index=True)
            st.download_button(
                f"Download {dk} as CSV",
                processed_days[dk].to_csv(index=False).encode(),
                f"{dk}_processed.csv", "text/csv",
                icon=":material/download:",
                key=f"dl_{dk}",
            )
            dup_df = duplicate_rows_by_day.get(dk)
            exc_df = excluded_rows_by_day.get(dk)
            if dup_df is not None and not dup_df.empty:
                with st.expander(f"Duplicate inverter IDs on {dk} (kept in calculations)", icon=":material/content_copy:"):
                    st.dataframe(dup_df, use_container_width=True, hide_index=True)
            if exc_df is not None and not exc_df.empty:
                with st.expander(f"Excluded inverter IDs on {dk} (too many duplicate rows)", icon=":material/block:"):
                    st.dataframe(exc_df, use_container_width=True, hide_index=True)

    # ------------------------------------------------------------- Failed Strings
    with tab_failed:
        st.markdown('<div class="section-title"><i class="fa-solid fa-circle-exclamation"></i>Newly Failed PV Strings</div>', unsafe_allow_html=True)
        if df_newly_failed_ui.empty:
            st.success("No newly failed strings between the previous and latest reference day.", icon=":material/check_circle:")
        else:
            fc1, fc2 = st.columns(2)
            plots = ["All"] + sorted(df_newly_failed_ui["Plot"].astype(str).unique().tolist()) if "Plot" in df_newly_failed_ui else ["All"]
            chosen_plot = fc1.selectbox("Plot", plots, key="nf_plot")
            search = fc2.text_input("Search inverter ID", key="nf_search")
            view = df_newly_failed_ui.copy()
            if chosen_plot != "All":
                view = view[view["Plot"] == chosen_plot]
            if search and "String Inverter" in view.columns:
                view = view[view["String Inverter"].astype(str).str.contains(search, case=False, na=False)]
            st.caption("**Current Working Strings Count** shows the inverter's live working-string tally so a recovered string can be cross-verified at a glance.")
            st.dataframe(view, use_container_width=True, hide_index=True)
            st.download_button("Download CSV", view.to_csv(index=False).encode(), "failed_strings.csv", "text/csv", icon=":material/download:")

    # ---------------------------------------------------------------- Re-Failed Strings
    with tab_refailed:
        st.markdown('<div class="section-title"><i class="fa-solid fa-rotate"></i>Re-Failed PV Strings</div>', unsafe_allow_html=True)
        if df_refailed_ui.empty:
            st.success("No re-failed strings detected.", icon=":material/check_circle:")
        else:
            st.caption("**Current Working Strings Count** shows the inverter's live working-string tally so a recovered string can be cross-verified at a glance.")
            st.dataframe(df_refailed_ui, use_container_width=True, hide_index=True)
            st.download_button("Download CSV", df_refailed_ui.to_csv(index=False).encode(), "refailed_strings.csv", "text/csv", icon=":material/download:")

    # ---------------------------------------------------------------- Negative Values
    with tab_negative:
        st.markdown('<div class="section-title"><i class="fa-solid fa-arrow-trend-down"></i>Negative Value Strings</div>', unsafe_allow_html=True)
        ncol1, ncol2 = st.columns(2)
        ncol1.metric("Current Negative Strings", f"{len(df_negative_current):,}")
        ncol2.metric("Continuous (2+ days)", f"{len(df_negative_continuous):,}")

        st.markdown("**Current negative strings (latest day)**")
        if df_negative_current.empty:
            st.success("No negative string readings on the latest day.", icon=":material/check_circle:")
        else:
            st.dataframe(df_negative_current, use_container_width=True, hide_index=True)
            st.download_button("Download CSV", df_negative_current.to_csv(index=False).encode(),
                                "negative_strings_current.csv", "text/csv", icon=":material/download:", key="dl_neg_current")

        st.markdown("**Continuous negative strings (persisting from previous day(s))**")
        if df_negative_continuous.empty:
            st.success("No persisting negative-current strings.", icon=":material/check_circle:")
        else:
            st.warning(f"{len(df_negative_continuous)} string(s) have read negative for more than one day in a row — investigate wiring/sensor faults.", icon=":material/warning:")
            st.dataframe(df_negative_continuous, use_container_width=True, hide_index=True)
            st.download_button("Download CSV", df_negative_continuous.to_csv(index=False).encode(),
                                "negative_strings_continuous.csv", "text/csv", icon=":material/download:", key="dl_neg_continuous")

    # ---------------------------------------------------------------- Low Performance
    with tab_lowperf:
        st.markdown('<div class="section-title"><i class="fa-solid fa-gauge-high"></i>Low Performance Strings</div>', unsafe_allow_html=True)
        st.caption(
            f"A working string is flagged when its current is below **{int(low_perf_ratio * 100)}%** of its own "
            "inverter's median working-string current on the latest day (e.g. one string reading ~5–6A while its "
            "siblings read ~10A)."
        )
        if df_low_perf.empty:
            st.success("No underperforming strings detected relative to their inverter peers.", icon=":material/check_circle:")
        else:
            lp1, lp2 = st.columns(2)
            plots_lp = ["All"] + sorted(df_low_perf["Plot"].astype(str).unique().tolist()) if "Plot" in df_low_perf else ["All"]
            chosen_plot_lp = lp1.selectbox("Plot", plots_lp, key="lp_plot")
            search_lp = lp2.text_input("Search inverter ID", key="lp_search")
            view = df_low_perf.copy()
            if chosen_plot_lp != "All":
                view = view[view["Plot"] == chosen_plot_lp]
            if search_lp and "String Inverter" in view.columns:
                view = view[view["String Inverter"].astype(str).str.contains(search_lp, case=False, na=False)]
            styled_lp = view.style.background_gradient(subset=["Performance vs Median (%)"], cmap="RdYlGn", vmin=0, vmax=100)
            st.dataframe(styled_lp, use_container_width=True, hide_index=True)
            st.download_button("Download CSV", view.to_csv(index=False).encode(), "low_performance_strings.csv", "text/csv", icon=":material/download:")

    # ---------------------------------------------------------------- Export
    with tab_export:
        st.markdown('<div class="section-title"><i class="fa-solid fa-file-export"></i>Full Excel Report</div>', unsafe_allow_html=True)
        st.caption("Choose which reports to include, then generate. Everything is selected by default.")

        # Sheets that only make sense when their underlying data actually exists this run.
        availability_by_key = {
            "dashboard": True,
            "trend": not daily_totals.empty,
            "plot_summary": not plot_summary.empty,
            "block_summary": not block_summary.empty,
            "inverter_matrix": not inverter_matrix_df.empty,
            "failed": True,
            "refailed": True,
            "negative": True,
            "low_perf": True,
            "duplicates": bool(duplicate_ids_by_day or excluded_ids_by_day),
        }

        sel_col1, sel_col2, sel_col3 = st.columns([1, 1, 1])
        with sel_col1:
            if st.button("Select all", use_container_width=True, key="export_select_all"):
                for key, _ in EXPORT_SHEET_OPTIONS:
                    st.session_state[f"export_chk_{key}"] = availability_by_key.get(key, True)
        with sel_col2:
            if st.button("Select none", use_container_width=True, key="export_select_none"):
                for key, _ in EXPORT_SHEET_OPTIONS:
                    st.session_state[f"export_chk_{key}"] = False
        with sel_col3:
            if st.button("Failures-only", use_container_width=True, key="export_select_failures",
                         help="Just the Failed Strings, Re-Failed Strings, Negative Values and Low Performance sheets."):
                failures_only = {"failed", "refailed", "negative", "low_perf"}
                for key, _ in EXPORT_SHEET_OPTIONS:
                    st.session_state[f"export_chk_{key}"] = key in failures_only

        st.markdown("<br>", unsafe_allow_html=True)
        check_cols = st.columns(2)
        selected_sheets = set()
        for i, (key, label) in enumerate(EXPORT_SHEET_OPTIONS):
            available = availability_by_key.get(key, True)
            col = check_cols[i % 2]
            default_checked = st.session_state.get(f"export_chk_{key}", True) and available
            checked = col.checkbox(
                label if available else f"{label} (no data this run)",
                value=default_checked,
                key=f"export_chk_{key}",
                disabled=not available,
            )
            if checked and available:
                selected_sheets.add(key)

        st.markdown("<br>", unsafe_allow_html=True)
        if not selected_sheets:
            st.warning("Select at least one report to include in the export.", icon=":material/warning:")

        if st.button("Generate Excel Report", icon=":material/description:", type="primary", disabled=not selected_sheets):
            with st.spinner("Building workbook..."):
                report_bytes = generate_excel_report(
                    kpis, plot_summary, block_summary, df_newly_failed_ui, df_refailed_ui,
                    inverter_status, comparison_day_keys, date_mapping, threshold,
                    df_negative_current=df_negative_current, df_low_perf=df_low_perf,
                    duplicate_ids_by_day=duplicate_ids_by_day, excluded_ids_by_day=excluded_ids_by_day,
                    daily_totals=daily_totals, inverter_matrix_df=inverter_matrix_df,
                    matrix_day_cols=matrix_day_cols, included_sheets=selected_sheets,
                )
                st.session_state["excel_report_bytes"] = report_bytes
                st.session_state["excel_report_name"] = f"PV_Plant_Report_{latest_key}_{date_mapping.get(latest_key, '')}.xlsx"
            st.success(f"Report generated with {len(selected_sheets)} sheet(s) — download it below.", icon=":material/check_circle:")

        if st.session_state["excel_report_bytes"]:
            st.download_button(
                "Download Excel Report",
                st.session_state["excel_report_bytes"],
                st.session_state["excel_report_name"],
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                icon=":material/download:",
            )
