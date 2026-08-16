"""
restore.py
==========
Restore & TAT Analysis Module for PV SCADA Analytics.
Optimized with caching, session-backed pagination, and a trimmed set of
tabs focused on: registry health, a fast 3-day summary, a full string
history matrix, working hours reference, and calendar comparison.
"""

import json
import io
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
import hashlib
import re
import time
from typing import Optional, Tuple, Dict, Any, List

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

import storage1 as storage

# ==========================================
# CONFIGURATION - Increase pandas Styler limit
# ==========================================
pd.set_option("styler.render.max_elements", 1000000)  # Increase to 1 million cells

WORKING_HOURS_START = 6
WORKING_HOURS_END = 18
WORKING_HOURS_PER_DAY = WORKING_HOURS_END - WORKING_HOURS_START
WORKING_CURRENT_THRESHOLD = 0.5
DEFAULT_TOTAL_ACTIVE_STRINGS = 19

DATA_DIR = Path("data")
HISTORY_FILE = DATA_DIR / "string_history.json"
DATA_DIR.mkdir(exist_ok=True)

INVERTER_ID_COLS = [
    "Inverter ID", "Inverter_ID", "Inverter", "ID",
    "Device Name", "String Inverter", "Inverters",
]

GRID_COLS = ["Grid", "Grid No", "Grid Name", "Grid ID", "Grid_ID"]

# Active string overrides matching app.py
ACTIVE_STRING_OVERRIDES = {
    "P2": {"IB1": 18, "IB3": 17, "IB4": 18, "IB5": 18},
    "P6": {"IB1": 18, "IB2": 18, "IB3": 18, "IB5": 18, "IB6": 18, "IB7": 18},
}

# Cache TTL in seconds
CACHE_TTL = 300  # 5 minutes
CACHE_LONG_TTL = 3600  # 1 hour

DEFAULT_PAGE_SIZE = 25


def sorted_filter_options(series):
    """Return non-null unique values sorted safely for Streamlit filters."""
    values = [value for value in series.dropna().unique()]

    def natural_key(value):
        text = str(value).strip()
        return tuple(
            (0, int(part)) if part.isdigit() else (1, part.casefold())
            for part in re.split(r"(\d+)", text)
        )

    return sorted(values, key=natural_key)


# ==========================================
# PAGINATION HELPER (session-state backed)
# ==========================================
def paginate_dataframe(df, page_size=DEFAULT_PAGE_SIZE, key_prefix="pg"):
    """Slice a dataframe into pages using session_state, with Prev/Next
    controls and a direct page-number jump. Returns the current page slice."""
    total_rows = len(df)
    if total_rows == 0:
        st.caption("No rows to display.")
        return df

    total_pages = max(1, (total_rows - 1) // page_size + 1)
    page_key = f"{key_prefix}_page"

    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    else:
        current_page = st.session_state[page_key]
        if current_page < 1:
            st.session_state[page_key] = 1
        elif current_page > total_pages:
            st.session_state[page_key] = total_pages

    nav_col1, nav_col2, nav_col3 = st.columns([1, 2, 1])
    with nav_col1:
        if st.button("◀ Previous", key=f"{key_prefix}_prev_btn",
                     disabled=st.session_state[page_key] <= 1, width="stretch"):
            st.session_state[page_key] -= 1
            st.rerun()
    with nav_col3:
        if st.button("Next ▶", key=f"{key_prefix}_next_btn",
                     disabled=st.session_state[page_key] >= total_pages, width="stretch"):
            st.session_state[page_key] += 1
            st.rerun()
    with nav_col2:
        st.number_input(
            "Page", min_value=1, max_value=total_pages,
            key=page_key, step=1, label_visibility="collapsed",
        )

    page = st.session_state[page_key]
    start = (page - 1) * page_size
    end = start + page_size
    st.caption(f"Rows {start + 1}-{min(end, total_rows)} of {total_rows} - Page {page}/{total_pages}")
    return df.iloc[start:end]


# ==========================================
# GRID COLUMN CONFIG (mirrors app.py's Data Table tab styling)
# ==========================================
def build_grid_column_config(df: pd.DataFrame) -> Dict[str, Any]:
    """Build a Streamlit column_config dict for a dataframe's numeric
    columns - progress bars for Availability/Failure percentages, and
    formatted number columns for counts and hour figures. This mirrors
    the st.column_config styling app.py applies on its Data Table tab,
    so grids look and behave consistently across both modules."""
    if df is None or df.empty:
        return {}

    config: Dict[str, Any] = {}
    for col in df.columns:
        if not pd.api.types.is_numeric_dtype(df[col]):
            continue
        col_l = str(col).lower()
        if "availability" in col_l or ("percent" in col_l and "%" in str(col)):
            config[col] = st.column_config.ProgressColumn(
                str(col), min_value=0, max_value=100, format="%.2f%%"
            )
        elif "tat" in col_l or "hour" in col_l or "hrs" in col_l:
            config[col] = st.column_config.NumberColumn(str(col), format="%.1f")
        elif any(k in col_l for k in (
            "working", "failed", "restored", "count", "strings",
            "inverters", "total", "current failure",
        )):
            config[col] = st.column_config.NumberColumn(str(col), format="%d")
    return config


# ==========================================
# ROLE HELPERS
# ==========================================
def can_upload(user_role: str) -> bool:
    return str(user_role).strip().lower() == "admin"


# ==========================================
# STRING HISTORY MANAGEMENT (with caching)
# ==========================================
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def load_string_history_cached():
    """Cached version of load_string_history for better performance."""
    if HISTORY_FILE.exists():
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            data.setdefault("strings", {})
            data.setdefault("last_updated", None)
            return data
        except Exception:
            return {"strings": {}, "last_updated": None}
    return {"strings": {}, "last_updated": None}


def load_string_history():
    """Load string history with caching."""
    return load_string_history_cached()


def save_string_history(history):
    """Save string history and invalidate cache."""
    history.setdefault("strings", {})
    history["last_updated"] = datetime.now().isoformat()
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2)
    load_string_history_cached.clear()


def init_history():
    if not HISTORY_FILE.exists():
        save_string_history({"strings": {}, "last_updated": None})


def get_inverter_column(df: pd.DataFrame):
    if df is None or df.empty:
        return None
    df_columns_lower_map = {str(c).strip().lower(): c for c in df.columns}
    for col in INVERTER_ID_COLS:
        if col in df.columns:
            return col
        lower_col = col.strip().lower()
        if lower_col in df_columns_lower_map:
            return df_columns_lower_map[lower_col]
    return None


def get_grid_column(df: pd.DataFrame):
    if df is None or df.empty:
        return None
    df_columns_lower_map = {str(c).strip().lower(): c for c in df.columns}
    for col in GRID_COLS:
        if col in df.columns:
            return col
        lower_col = col.strip().lower()
        if lower_col in df_columns_lower_map:
            return df_columns_lower_map[lower_col]
    return None


def add_grid_from_snapshot(df: pd.DataFrame, snapshot_df: pd.DataFrame, target_inverter_col: str):
    """Fill or add Grid from the saved sheet for older history rows."""
    if df is None or df.empty or snapshot_df is None or snapshot_df.empty:
        return df
    source_inverter_col = get_inverter_column(snapshot_df)
    source_grid_col = get_grid_column(snapshot_df)
    if not source_inverter_col or not source_grid_col or target_inverter_col not in df.columns:
        return df

    grid_map = (
        snapshot_df[[source_inverter_col, source_grid_col]]
        .dropna(subset=[source_inverter_col])
        .drop_duplicates(subset=[source_inverter_col], keep="last")
        .assign(**{source_inverter_col: lambda d: d[source_inverter_col].astype(str)})
        .set_index(source_inverter_col)[source_grid_col]
    )

    enriched = df.copy()
    mapped_grid = enriched[target_inverter_col].astype(str).map(grid_map).fillna("")
    if "Grid" in enriched.columns:
        current_grid = enriched["Grid"].astype(str)
        enriched["Grid"] = current_grid.where(current_grid.str.strip().ne(""), mapped_grid)
    else:
        insert_at = 2 if "Block" in enriched.columns else min(2, len(enriched.columns))
        enriched.insert(insert_at, "Grid", mapped_grid)
    return enriched


def get_pv_current_columns(df: pd.DataFrame):
    """Return every PV-I current column present in the sheet, sorted by
    string number. Matches app.py's get_available_pv_columns_cached -
    no columns are dropped here; whether a given string counts as
    "active" for a given inverter is decided by get_total_active_strings()
    at the point of use, not by removing columns up front."""
    if df is None or df.empty:
        return []
    pv_cols = [c for c in df.columns if str(c).strip().upper().startswith("PV-I")]

    def sort_key(name):
        try:
            return int(str(name).replace("PV-I", "").strip())
        except Exception:
            return 9999

    return sorted(pv_cols, key=sort_key)


def get_total_active_strings(plot, block):
    """Get total active strings for a plot/block combination."""
    plot_key = str(plot).strip().upper() if plot else ""
    block_key = str(block).strip().upper() if block else ""
    if plot_key in ACTIVE_STRING_OVERRIDES and block_key in ACTIVE_STRING_OVERRIDES[plot_key]:
        return ACTIVE_STRING_OVERRIDES[plot_key][block_key]
    return DEFAULT_TOTAL_ACTIVE_STRINGS


def update_string_history(df, date_str):
    """Update per-string status history with a given day's data."""
    if df is None or df.empty:
        return

    history = load_string_history()
    inverter_col = get_inverter_column(df)
    if not inverter_col:
        return
    grid_col = get_grid_column(df)

    pv_current_cols = get_pv_current_columns(df)
    if not pv_current_cols:
        return

    df = df.copy()
    df["Total_Active"] = df.apply(
        lambda row: get_total_active_strings(row.get("Plot"), row.get("Block")), axis=1
    )

    for _, row in df.iterrows():
        inverter_id = str(row[inverter_col])
        total_active = int(row.get("Total_Active", DEFAULT_TOTAL_ACTIVE_STRINGS))
        grid_value = str(row.get(grid_col, "")) if grid_col else ""

        history["strings"].setdefault(inverter_id, {})
        history["strings"][inverter_id].setdefault("_metadata", {
            "plot": str(row.get("Plot", "")),
            "block": str(row.get("Block", "")),
            "sacu": str(row.get("SACU", "")),
            "grid": grid_value,
            "total_active": total_active,
        })
        history["strings"][inverter_id]["_metadata"].update({
            "plot": str(row.get("Plot", "")),
            "block": str(row.get("Block", "")),
            "sacu": str(row.get("SACU", "")),
            "grid": grid_value,
            "total_active": total_active,
        })

        for col in pv_current_cols:
            string_id = str(col).strip()
            match = re.search(r"(\d+)\s*$", string_id)
            string_num = int(match.group(1)) if match else None
            current_value = pd.to_numeric(row.get(col), errors="coerce")

            if string_num is not None and string_num > total_active:
                status = "open"
            elif pd.isna(current_value):
                status = "open"
            elif current_value > WORKING_CURRENT_THRESHOLD:
                status = "working"
            else:
                status = "failed"

            if string_id not in history["strings"][inverter_id]:
                history["strings"][inverter_id][string_id] = {
                    "status_history": [],
                    "current_status": "unknown",
                    "last_change": None,
                }

            status_history = history["strings"][inverter_id][string_id]["status_history"]
            last_record = status_history[-1] if status_history else None

            should_add = (
                last_record is None
                or last_record.get("status") != status
                or last_record.get("date") != date_str
            )

            if should_add:
                status_history.append({
                    "date": date_str,
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "status": status,
                    "value": float(current_value) if pd.notna(current_value) else 0,
                })
                history["strings"][inverter_id][string_id]["current_status"] = status
                history["strings"][inverter_id][string_id]["last_change"] = datetime.now().isoformat()

    save_string_history(history)


# ==========================================
# SNAPSHOT ACCESS (delegates to storage.py)
# ==========================================
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def get_available_snapshot_dates_cached():
    """Cached version of get_available_snapshot_dates."""
    return storage.get_available_snapshot_dates()


def get_available_snapshot_dates():
    """Get available snapshot dates with caching."""
    return get_available_snapshot_dates_cached()


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def load_snapshot_sheet_cached(snapshot_date, sheet_name):
    """Cached version of load_snapshot_sheet."""
    entry = storage.get_upload_for_date(snapshot_date)
    if not entry:
        return None
    return storage.load_sheet_csv(entry["upload_id"], sheet_name)


def load_snapshot_sheet(snapshot_date, sheet_name):
    """Load snapshot sheet with caching."""
    return load_snapshot_sheet_cached(snapshot_date, sheet_name)


def clear_snapshot_caches():
    """Clear cached snapshot reads after uploads, deletes, or full restores."""
    get_available_snapshot_dates_cached.clear()
    load_snapshot_sheet_cached.clear()


def get_snapshots_in_range(from_date, to_date):
    """Get all snapshot dates in the given range."""
    available = get_available_snapshot_dates()
    from_str, to_str = str(from_date), str(to_date)
    return [d for d in available if from_str <= d <= to_str]


# ==========================================
# COLOR EXCEL EXPORT FUNCTIONS
# ==========================================
def _status_fill_hex(status_value):
    """Get hex color for status values."""
    return {
        "Green - Working": "10B981",
        "Yellow - Low Performance": "FBBF24",
        "Red - Failed": "EF4444",
        "Blinking Red - Negative Value": "7F1D1D",
        "working": "10B981",
        "failed": "EF4444",
        "open": "94A3B8",
        "unknown": "64748B",
        "Excellent": "10B981",
        "Good": "34D399",
        "Fair": "FBBF24",
        "Poor": "EF4444",
    }.get(str(status_value))


def _get_value_color(value, min_val=0, max_val=100):
    """Get color based on value range."""
    try:
        v = float(value)
        if v >= 90:
            return "10B981"
        elif v >= 70:
            return "34D399"
        elif v >= 50:
            return "FBBF24"
        elif v >= 30:
            return "F59E0B"
        else:
            return "EF4444"
    except (TypeError, ValueError):
        return None


def _get_pv_value_color(value):
    """Get color for PV current values."""
    try:
        v = float(value)
        if v > 5.0:
            return "10B981"
        elif v > 3.0:
            return "34D399"
        elif v > 1.5:
            return "FBBF24"
        elif v > 0.5:
            return "F59E0B"
        else:
            return "EF4444"
    except (TypeError, ValueError):
        return None


def render_colored_excel_download(label, df, file_prefix, sheet_name="Sheet1", key=None, 
                                   status_cols=None, value_cols=None, pv_value_cols=None):
    """
    Export dataframe to Excel with color-coded cells.
    Colors are applied directly to the cells, not as extra columns.
    """
    from openpyxl.styles import PatternFill, Font

    if df.empty:
        st.caption("Nothing to export for the current selection.")
        return

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        col_idx_map = {col: idx + 1 for idx, col in enumerate(df.columns)}
        
        status_cols = status_cols or [c for c in df.columns if "Status" in c or "status" in c]
        value_cols = value_cols or [c for c in df.columns if "Value" in c or "Count" in c or "TAT" in c or "Failure" in c]
        pv_value_cols = pv_value_cols or [c for c in df.columns if c.startswith("PV-") or "Current" in c]

        for row_idx in range(2, ws.max_row + 1):
            # Color status columns
            for col_name in status_cols:
                if col_name in col_idx_map:
                    cell = ws.cell(row=row_idx, column=col_idx_map[col_name])
                    fill_hex = _status_fill_hex(cell.value)
                    if fill_hex:
                        cell.fill = PatternFill("solid", fgColor=fill_hex)
                        cell.font = Font(bold=True, color="FFFFFF")

            # Color PV value columns with gradient
            for col_name in pv_value_cols:
                if col_name in col_idx_map:
                    cell = ws.cell(row=row_idx, column=col_idx_map[col_name])
                    if cell.value is not None and cell.value != "":
                        try:
                            v = float(cell.value)
                            fill_hex = _get_pv_value_color(v)
                            if fill_hex:
                                cell.fill = PatternFill("solid", fgColor=fill_hex)
                                cell.font = Font(bold=True, color="FFFFFF")
                        except (TypeError, ValueError):
                            pass

            # Color general value columns
            for col_name in value_cols:
                if col_name in col_idx_map and col_name not in pv_value_cols:
                    cell = ws.cell(row=row_idx, column=col_idx_map[col_name])
                    if cell.value is not None and cell.value != "":
                        try:
                            v = float(cell.value)
                            fill_hex = _get_value_color(v)
                            if fill_hex:
                                cell.fill = PatternFill("solid", fgColor=fill_hex)
                                cell.font = Font(bold=True, color="FFFFFF")
                        except (TypeError, ValueError):
                            pass

    buffer.seek(0)
    st.download_button(
        label=label,
        data=buffer.getvalue(),
        file_name=f"{file_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
        width="stretch",
    )


# ==========================================
# FULL BACKUP & RESTORE FUNCTIONS
# ==========================================
def get_backup_manifest():
    """Build a simple manifest describing the files available for full backup."""
    data_dir = Path("data")
    total_files = 0
    total_size_bytes = 0
    file_types: Dict[str, int] = {}

    if data_dir.exists():
        for file_path in data_dir.rglob("*"):
            if file_path.is_file():
                total_files += 1
                total_size_bytes += file_path.stat().st_size
                ext = file_path.suffix.lower()
                if ext:
                    file_types[ext] = file_types.get(ext, 0) + 1
                else:
                    file_types["<no extension>"] = file_types.get("<no extension>", 0) + 1

    return {
        "total_files": total_files,
        "total_size_bytes": total_size_bytes,
        "total_size_kb": round(total_size_bytes / 1024, 1),
        "file_types": file_types,
    }


def export_full_backup_bytes():
    """Export all application data as a zip file."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        data_dir = Path("data")
        if data_dir.exists():
            for file_path in data_dir.rglob("*"):
                if file_path.is_file():
                    arcname = str(file_path.relative_to(Path(".")))
                    zipf.write(file_path, arcname)
    buffer.seek(0)
    return buffer.getvalue()


def import_full_backup_bytes(zip_bytes, overwrite=False):
    """Import application data from a zip file."""
    buffer = io.BytesIO(zip_bytes)
    restored_count = 0
    with zipfile.ZipFile(buffer, 'r') as zipf:
        for info in zipf.infolist():
            if info.filename.startswith("data/") and not info.is_dir():
                target_path = Path(info.filename)
                if target_path.exists() and not overwrite:
                    continue
                target_path.parent.mkdir(parents=True, exist_ok=True)
                with target_path.open('wb') as f:
                    f.write(zipf.read(info))
                restored_count += 1
    return restored_count


# ==========================================
# INVERTER HISTORY MATRIX - Date-wise Inverter String Counts
# ==========================================
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def build_inverter_date_matrix_cached(dates, sheet_name="Sheet1"):
    """Build a matrix of inverter string counts for each date."""
    matrix_data = []
    all_inverters = set()

    for date_str in dates:
        df = load_snapshot_sheet(date_str, sheet_name)
        if df is None or df.empty:
            continue

        inverter_col = get_inverter_column(df)
        if not inverter_col:
            continue
        grid_col = get_grid_column(df)
        context_cols = [c for c in [inverter_col, "Plot", "Block", grid_col, "SACU"] if c and c in df.columns]

        if "Working String Count" in df.columns:
            counts = df.groupby(context_cols).agg({
                "Working String Count": "sum",
                "Total Active Strings": "sum",
                "Failed String Count": "sum"
            }).reset_index()
        else:
            pv_cols = get_pv_current_columns(df)
            counts = []
            for _, row in df.iterrows():
                working = 0
                for col in pv_cols:
                    val = pd.to_numeric(row.get(col), errors="coerce")
                    if pd.notna(val) and val > WORKING_CURRENT_THRESHOLD:
                        working += 1
                total = get_total_active_strings(row.get("Plot"), row.get("Block"))
                counts.append({
                    inverter_col: row[inverter_col],
                    "Plot": row.get("Plot", ""),
                    "Block": row.get("Block", ""),
                    "Grid": row.get(grid_col, "") if grid_col else "",
                    "SACU": row.get("SACU", ""),
                    "Working String Count": working,
                    "Total Active Strings": total,
                    "Failed String Count": max(0, total - working)
                })
            counts = pd.DataFrame(counts)

        counts["Date"] = date_str
        counts["Inverter"] = counts[inverter_col]
        if grid_col and grid_col != "Grid" and grid_col in counts.columns:
            counts = counts.rename(columns={grid_col: "Grid"})
        if "Grid" not in counts.columns:
            counts["Grid"] = ""
        all_inverters.update(counts["Inverter"].unique())
        matrix_data.append(counts)

    if not matrix_data:
        return pd.DataFrame()

    result = {}
    for inv in all_inverters:
        inv_data = {"Inverter": inv}
        plot, block, grid, sacu = "", "", "", ""
        for df_item in matrix_data:
            row = df_item[df_item["Inverter"] == inv]
            if not row.empty:
                plot = row.iloc[0].get("Plot", "")
                block = row.iloc[0].get("Block", "")
                grid = row.iloc[0].get("Grid", "")
                sacu = row.iloc[0].get("SACU", "")
                break
        inv_data["Plot"] = plot
        inv_data["Block"] = block
        inv_data["Grid"] = grid
        inv_data["SACU"] = sacu

        for df_item in matrix_data:
            date_str = df_item["Date"].iloc[0] if not df_item.empty else ""
            row = df_item[df_item["Inverter"] == inv]
            if not row.empty:
                working = int(row.iloc[0].get("Working String Count", 0))
                total = int(row.iloc[0].get("Total Active Strings", 0))
                failed = int(row.iloc[0].get("Failed String Count", 0))
                inv_data[f"{date_str} - Working"] = working
                inv_data[f"{date_str} - Failed"] = failed
                inv_data[f"{date_str} - Total"] = total
            else:
                inv_data[f"{date_str} - Working"] = 0
                inv_data[f"{date_str} - Failed"] = 0
                inv_data[f"{date_str} - Total"] = 0

        result[inv] = inv_data

    return pd.DataFrame(list(result.values()))


def display_inverter_date_matrix():
    """Display date-wise inverter string counts matrix."""
    st.subheader("📊 Inverter Date-wise String Counts")
    st.caption("Shows every inverter's working/total/failed string counts for each available snapshot date.")

    available_dates = get_available_snapshot_dates()
    if len(available_dates) < 1:
        st.info("📌 No snapshot dates available.")
        return

    min_date = datetime.strptime(available_dates[0], "%Y-%m-%d").date()
    max_date = datetime.strptime(available_dates[-1], "%Y-%m-%d").date()

    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date", value=min_date, min_value=min_date, max_value=max_date, key="matrix_from")
    with col2:
        to_date = st.date_input("To Date", value=max_date, min_value=min_date, max_value=max_date, key="matrix_to")

    if from_date > to_date:
        st.error("❌ From Date must be on or before To Date.")
        return

    dates_in_range = [d for d in available_dates if str(from_date) <= d <= str(to_date)]
    if len(dates_in_range) < 1:
        st.info("📌 No snapshots in the selected date range.")
        return

    with st.spinner("🔄 Building inverter date matrix..."):
        df_matrix = build_inverter_date_matrix_cached(tuple(dates_in_range), "Sheet1")

    if df_matrix.empty:
        st.info("📌 No data available for the selected range.")
        return

    plot_choices = ["All"] + sorted_filter_options(df_matrix["Plot"])
    selected_plot = st.selectbox("Plot", plot_choices, key="matrix_plot_filter")

    if selected_plot != "All":
        df_matrix = df_matrix[df_matrix["Plot"] == selected_plot]

    st.markdown(f"**{len(df_matrix)} inverters** shown across {len(dates_in_range)} dates.")

    display_cols = ["Inverter", "Plot", "Block", "Grid", "SACU"]
    for d in dates_in_range:
        display_cols.append(f"{d} - Working")
        display_cols.append(f"{d} - Failed")
        display_cols.append(f"{d} - Total")

    df_display = df_matrix[display_cols].copy()

    def color_working(val):
        try:
            if isinstance(val, (int, float)) and val > 0:
                return "background-color: #10b981; color: white; font-weight: bold;"
        except:
            pass
        return ""

    def color_failed(val):
        try:
            if isinstance(val, (int, float)) and val > 0:
                return "background-color: #ef4444; color: white; font-weight: bold;"
        except:
            pass
        return ""

    styled = df_display.style
    for d in dates_in_range:
        styled = styled.map(color_working, subset=[f"{d} - Working"])
        styled = styled.map(color_failed, subset=[f"{d} - Failed"])

    # Check if we need pagination for styling
    total_cells = len(df_display) * len(df_display.columns)
    if total_cells > 250000:
        st.warning(f"⚠️ Large dataset ({total_cells:,} cells). Showing first 100 rows with styling. Download full data below.")
        df_display_page = df_display.head(100)
        styled = styled.applymap(lambda x: color_working(x) if isinstance(x, (int, float)) and x > 0 else "", subset=[f"{d} - Working" for d in dates_in_range])
        styled = styled.applymap(lambda x: color_failed(x) if isinstance(x, (int, float)) and x > 0 else "", subset=[f"{d} - Failed" for d in dates_in_range])
        st.dataframe(styled, use_container_width=True, height=500)
    else:
        st.dataframe(styled, use_container_width=True, height=500)

    render_colored_excel_download(
        "📥 Download Inverter Matrix (Excel, color-coded)",
        df_display,
        "inverter_matrix",
        key="matrix_download",
        value_cols=[f"{d} - Working" for d in dates_in_range] + [f"{d} - Failed" for d in dates_in_range] + [f"{d} - Total" for d in dates_in_range],
        pv_value_cols=[f"{d} - Working" for d in dates_in_range]
    )

    st.markdown("---")
    st.markdown("#### 📈 Summary")

    total_working_all = 0
    total_failed_all = 0
    total_strings_all = 0
    
    for d in dates_in_range:
        total_working_all += df_matrix[f"{d} - Working"].sum()
        total_failed_all += df_matrix[f"{d} - Failed"].sum()
        total_strings_all += df_matrix[f"{d} - Total"].sum()

    st.markdown("**Per-Date Summary**")
    date_summary_rows = []
    for d in dates_in_range:
        working = df_matrix[f"{d} - Working"].sum()
        failed = df_matrix[f"{d} - Failed"].sum()
        total = df_matrix[f"{d} - Total"].sum()
        avail = round((working / total * 100), 2) if total > 0 else 0.0
        date_summary_rows.append({
            "Date": d,
            "Working": working,
            "Failed": failed,
            "Total": total,
            "Availability (%)": avail,
        })
    
    df_date_summary = pd.DataFrame(date_summary_rows)
    st.dataframe(df_date_summary, use_container_width=True, hide_index=True,
                 column_config=build_grid_column_config(df_date_summary))

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total Inverters", len(df_matrix))
    col2.metric("Total Working (all dates)", int(total_working_all))
    col3.metric("Total Failed (all dates)", int(total_failed_all))
    col4.metric("Total Strings (all dates)", int(total_strings_all))
    col5.metric("Overall Availability", f"{(total_working_all / total_strings_all * 100):.1f}%" if total_strings_all > 0 else "0%")


def display_inverter_history_comparison():
    """Changed Inverters: Compare two dates."""
    st.subheader("📊 Inverter History Matrix - Changed Inverters (From Date → To Date)")
    st.caption(
        "Every inverter whose Working String Count changed between two snapshot dates - "
        "sorted by Old count (worst first) so the inverters most worth checking surface at the top."
    )

    available_dates = get_available_snapshot_dates()
    if len(available_dates) < 2:
        st.info("📌 At least 2 saved snapshot dates are required for this comparison.")
        return

    min_date = datetime.strptime(available_dates[0], "%Y-%m-%d").date()
    max_date = datetime.strptime(available_dates[-1], "%Y-%m-%d").date()
    sorted_dates = sorted(available_dates)

    col1, col2 = st.columns(2)
    with col1:
        default_old = datetime.strptime(sorted_dates[-2], "%Y-%m-%d").date()
        from_date = st.date_input("From Date", value=default_old, min_value=min_date, max_value=max_date, key="comp_from")
    with col2:
        to_date = st.date_input("To Date", value=max_date, min_value=min_date, max_value=max_date, key="comp_to")

    if from_date >= to_date:
        st.error("❌ From Date must be strictly before To Date.")
        return

    old_date, new_date = str(from_date), str(to_date)

    latest_df = load_snapshot_sheet(new_date, "Sheet1")
    plot_choices = ["All"]
    if latest_df is not None and "Plot" in latest_df.columns:
        plot_choices = ["All"] + sorted_filter_options(latest_df["Plot"])
    selected_plot = st.selectbox("Plot (optional filter)", plot_choices, key="comp_plot")

    with st.spinner("🔄 Comparing snapshots..."):
        result, error = build_snapshot_diff(old_date, new_date, "Sheet1")
    if error:
        st.error(error)
        return

    changed = result["changed_inverters"]
    if selected_plot != "All" and "Plot" in changed.columns:
        changed = changed[changed["Plot"] == selected_plot]

    if changed.empty:
        st.success(f"✅ No inverter's Working String Count changed between {old_date} and {new_date}.")
        return

    improved = int((changed["Change"] > 0).sum())
    worsened = int((changed["Change"] < 0).sum())
    m1, m2, m3 = st.columns(3)
    m1.metric("Inverters Changed", len(changed))
    m2.metric("Improved", improved, delta=improved, delta_color="normal")
    m3.metric("Worsened", worsened, delta=-worsened, delta_color="inverse")

    chart_df = changed[["String Inverter", "Working String Count_Old", "Working String Count_New"]].copy()
    fig = go.Figure()
    fig.add_trace(go.Bar(x=chart_df["String Inverter"], y=chart_df["Working String Count_Old"],
                          name=f"{old_date}", marker_color="#64748b"))
    fig.add_trace(go.Bar(x=chart_df["String Inverter"], y=chart_df["Working String Count_New"],
                          name=f"{new_date}", marker_color="#10b981"))
    fig.update_layout(barmode="group", height=420, title="Working String Count - From Date vs To Date",
                       plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
    st.plotly_chart(fig, use_container_width=True, key="comp_chart")

    display_cols = [c for c in ["Plot", "Block", "Grid", "String Inverter", "Working String Count_Old",
                                 "Working String Count_New", "Change", "Failed String Count_Old",
                                 "Failed String Count_New"] if c in changed.columns]

    rename_map = {
        "Working String Count_Old": f"Working ({old_date})",
        "Working String Count_New": f"Working ({new_date})",
        "Failed String Count_Old": f"Failed ({old_date})",
        "Failed String Count_New": f"Failed ({new_date})",
    }
    display_df = changed[display_cols].rename(columns=rename_map)

    def color_change(val):
        if val > 0:
            return "background-color: #10b981; color: white; font-weight: bold;"
        if val < 0:
            return "background-color: #ef4444; color: white; font-weight: bold;"
        return ""

    styled = display_df.style.map(color_change, subset=["Change"])
    st.dataframe(styled, use_container_width=True, height=450)

    render_colored_excel_download(
        "📥 Download Changed Inverters (Excel, color-coded)",
        display_df,
        f"changed_inverters_{old_date}_to_{new_date}",
        key="comp_download",
        value_cols=[f"Working ({old_date})", f"Working ({new_date})", f"Failed ({old_date})", f"Failed ({new_date})"]
    )


def display_inverter_history_matrix():
    """Main Inverter History Matrix tab with two sub-tabs."""
    tab1, tab2 = st.tabs(["📅 Date-wise Inverter Counts", "📊 Compare Two Dates"])

    with tab1:
        display_inverter_date_matrix()

    with tab2:
        display_inverter_history_comparison()


# ==========================================
# STRING HISTORY MATRIX
# ==========================================
@st.cache_data(ttl=CACHE_LONG_TTL, show_spinner=False)
def build_string_status_matrix_cached(history_hash, plot, block, grid, sacu, inverter,
                                       date_start_str, date_end_str, max_dates):
    """Cached builder for the full Block -> Inverter -> String status matrix."""
    history = load_string_history()
    strings = history.get("strings", {})

    date_start = date_start_str
    date_end = date_end_str

    rows = []
    dates_set = set()

    for inv_id, inv_data in strings.items():
        if inv_id.startswith("_"):
            continue
        metadata = inv_data.get("_metadata", {})
        p = metadata.get("plot", "")
        b = metadata.get("block", "")
        g = metadata.get("grid", "")
        s = metadata.get("sacu", "")

        if plot and plot != "All" and p != plot:
            continue
        if block and block != "All" and b != block:
            continue
        if grid and grid != "All" and g != grid:
            continue
        if sacu and sacu != "All" and s != sacu:
            continue
        if inverter and inverter != "All" and inv_id != inverter:
            continue

        for string_id, sdata in inv_data.items():
            if string_id.startswith("_"):
                continue
            status_history = sdata.get("status_history", [])
            row_statuses = {}
            for record in status_history:
                d = record.get("date", "")
                if date_start and d < date_start:
                    continue
                if date_end and d > date_end:
                    continue
                row_statuses[d] = record.get("status", "unknown")
                dates_set.add(d)

            if row_statuses:
                rows.append({
                    "Plot": p, "Block": b, "Grid": g, "SACU": s,
                    "Inverter": inv_id, "String": string_id,
                    **row_statuses,
                })

    dates_sorted = sorted(dates_set)
    if max_dates and len(dates_sorted) > max_dates:
        dates_sorted = dates_sorted[-max_dates:]

    if not rows:
        return pd.DataFrame(), dates_sorted

    df = pd.DataFrame(rows)
    for d in dates_sorted:
        if d not in df.columns:
            df[d] = "unknown"

    ordered_cols = ["Plot", "Block", "Grid", "SACU", "Inverter", "String"] + dates_sorted
    ordered_cols = [c for c in ordered_cols if c in df.columns]
    df = df[ordered_cols].fillna("unknown")
    sort_cols = [c for c in ["Plot", "Block", "Grid", "Inverter", "String"] if c in df.columns]
    df = df.sort_values(sort_cols).reset_index(drop=True)
    return df, dates_sorted


@st.cache_data(ttl=CACHE_LONG_TTL, show_spinner=False)
def build_tat_report_cached(history_hash, plot, block, grid, sacu, inverter,
                             date_start_str, date_end_str):
    """Build TAT report with failure/restoration analysis."""
    history = load_string_history()
    strings = history.get("strings", {})
    
    date_start = date_start_str
    date_end = date_end_str
    
    report_rows = []
    
    for inv_id, inv_data in strings.items():
        if inv_id.startswith("_"):
            continue
        metadata = inv_data.get("_metadata", {})
        p = metadata.get("plot", "")
        b = metadata.get("block", "")
        g = metadata.get("grid", "")
        s = metadata.get("sacu", "")
        
        if plot and plot != "All" and p != plot:
            continue
        if block and block != "All" and b != block:
            continue
        if grid and grid != "All" and g != grid:
            continue
        if sacu and sacu != "All" and s != sacu:
            continue
        if inverter and inverter != "All" and inv_id != inverter:
            continue
        
        for string_id, sdata in inv_data.items():
            if string_id.startswith("_"):
                continue
                
            status_history = sdata.get("status_history", [])
            
            filtered_history = []
            for record in status_history:
                d = record.get("date", "")
                if date_start and d < date_start:
                    continue
                if date_end and d > date_end:
                    continue
                filtered_history.append(record)
            
            if not filtered_history:
                continue
            
            failure_count = 0
            restored_count = 0
            current_failure_hours = 0
            tat_hours = 0
            last_status = None
            failure_start_date = None
            
            for i, record in enumerate(filtered_history):
                status = record.get("status", "unknown")
                date_str = record.get("date", "")
                
                if status == "failed" and last_status != "failed":
                    failure_count += 1
                    failure_start_date = datetime.strptime(date_str, "%Y-%m-%d")
                elif status == "working" and last_status == "failed":
                    restored_count += 1
                    if failure_start_date:
                        days_diff = (datetime.strptime(date_str, "%Y-%m-%d") - failure_start_date).days
                        tat_hours += days_diff * WORKING_HOURS_PER_DAY
                        failure_start_date = None
                
                last_status = status
            
            if last_status == "failed" and failure_start_date:
                days_diff = (datetime.now() - failure_start_date).days
                current_failure_hours = days_diff * WORKING_HOURS_PER_DAY
            
            date_status = {}
            for record in filtered_history:
                d = record.get("date", "")
                status = record.get("status", "unknown")
                value = record.get("value", 0)
                date_status[d] = {"status": status, "value": value}
            
            row = {
                "Plot": p, "Block": b, "Grid": g, "SACU": s,
                "Inverter": inv_id, "String": string_id,
                "Failure Count": failure_count,
                "Restored Count": restored_count,
                "TAT (Hours)": round(tat_hours, 1),
                "Current Failure (Hours)": round(current_failure_hours, 1),
                "Current Failure (Mins)": round(current_failure_hours * 60, 0),
            }
            
            for d, data in date_status.items():
                row[f"{d} - Status"] = data["status"]
                row[f"{d} - Value (A)"] = data["value"]
            
            report_rows.append(row)
    
    if not report_rows:
        return pd.DataFrame()
    
    df_report = pd.DataFrame(report_rows)
    
    cols = ["Plot", "Block", "Grid", "SACU", "Inverter", "String", 
            "Failure Count", "Restored Count", "TAT (Hours)", 
            "Current Failure (Hours)", "Current Failure (Mins)"]
    
    date_cols = sorted([c for c in df_report.columns if " - Status" in c or " - Value" in c])
    cols.extend(date_cols)
    
    df_report = df_report[[c for c in cols if c in df_report.columns]]
    
    return df_report


def display_string_history_matrix(history, current_df):
    """String Status Change Report with date range selection and 3 sub-tabs."""
    st.subheader("📈 String History Matrix - Status Change Report")
    st.caption(
        "Select a date range - analyze string status changes, TAT, and failure/restoration patterns."
    )

    available_dates = get_available_snapshot_dates()
    if len(available_dates) < 2:
        st.info("📌 At least 2 saved snapshot dates are required for this comparison.")
        return

    min_date = datetime.strptime(available_dates[0], "%Y-%m-%d").date()
    max_date = datetime.strptime(available_dates[-1], "%Y-%m-%d").date()
    sorted_dates = sorted(available_dates)
    
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date", value=datetime.strptime(sorted_dates[-2], "%Y-%m-%d").date(), 
                                   min_value=min_date, max_value=max_date, key="string_range_from")
    with col2:
        to_date = st.date_input("To Date", value=datetime.strptime(sorted_dates[-1], "%Y-%m-%d").date(),
                                 min_value=min_date, max_value=max_date, key="string_range_to")
    
    if from_date >= to_date:
        st.error("❌ From Date must be strictly before To Date.")
        return
    
    old_date, new_date = str(from_date), str(to_date)
    
    latest_df = load_snapshot_sheet(new_date, "Sheet1")
    grid_col = get_grid_column(latest_df) if latest_df is not None else None
    plot_choices, block_choices, grid_choices = ["All"], ["All"], ["All"]
    if latest_df is not None and "Plot" in latest_df.columns:
        plot_choices = ["All"] + sorted_filter_options(latest_df["Plot"])

    col1, col2, col3 = st.columns(3)
    with col1:
        selected_plot = st.selectbox("Plot", plot_choices, key="string_matrix_plot")
    if latest_df is not None and "Block" in latest_df.columns:
        block_source = latest_df if selected_plot == "All" else latest_df[latest_df["Plot"] == selected_plot]
        block_choices = ["All"] + sorted_filter_options(block_source["Block"])
    with col2:
        selected_block = st.selectbox("Block", block_choices, key="string_matrix_block")
    if latest_df is not None and grid_col:
        grid_source = latest_df
        if selected_plot != "All" and "Plot" in grid_source.columns:
            grid_source = grid_source[grid_source["Plot"] == selected_plot]
        if selected_block != "All" and "Block" in grid_source.columns:
            grid_source = grid_source[grid_source["Block"] == selected_block]
        grid_choices = ["All"] + sorted_filter_options(grid_source[grid_col])
    with col3:
        selected_grid = st.selectbox("Grid", grid_choices, key="string_matrix_grid")

    with st.spinner("🔄 Building string status matrix..."):
        history_hash = hashlib.md5(json.dumps(history, sort_keys=True).encode()).hexdigest()
        df_matrix, dates_in_range = build_string_status_matrix_cached(
            history_hash, selected_plot, selected_block, "All", "All", "All",
            old_date, new_date, 30
        )
        
        df_tat_report = build_tat_report_cached(
            history_hash, selected_plot, selected_block, "All", "All", "All",
            old_date, new_date
        )
        df_matrix = add_grid_from_snapshot(df_matrix, latest_df, "Inverter")
        df_tat_report = add_grid_from_snapshot(df_tat_report, latest_df, "Inverter")
        if selected_grid != "All":
            if "Grid" in df_matrix.columns:
                df_matrix = df_matrix[df_matrix["Grid"] == selected_grid]
            if "Grid" in df_tat_report.columns:
                df_tat_report = df_tat_report[df_tat_report["Grid"] == selected_grid]
        
        # Build snapshot diff for status change summary
        diff_result, diff_error = build_snapshot_diff(old_date, new_date, "Sheet1")

    # Create sub-tabs
    sub_tab1, sub_tab2, sub_tab3, sub_tab4 = st.tabs([
        "📊 Status Matrix", 
        "📈 Status Change Summary", 
        "📋 TAT Report"
        , "Restore",
    ])

    with sub_tab1:
        if df_matrix.empty:
            st.info("No data available for the selected range.")
            return
        
        st.markdown(f"**{len(df_matrix)} strings** shown across {len(dates_in_range)} dates.")
        
        display_cols = ["Plot", "Block", "Grid", "SACU", "Inverter", "String"] + dates_in_range
        display_cols = [c for c in display_cols if c in df_matrix.columns]
        df_display = df_matrix[display_cols].copy()
        
        def color_status_value(val):
            if val == "working":
                return "background-color: #10b981; color: white; font-weight: bold;"
            elif val == "failed":
                return "background-color: #ef4444; color: white; font-weight: bold;"
            elif val == "open":
                return "background-color: #94a3b8; color: white; font-weight: bold;"
            elif val == "unknown":
                return "background-color: #64748b; color: white; font-weight: bold;"
            return ""
        
        total_cells = len(df_display) * len(df_display.columns)
        if total_cells > 250000:
            st.warning(f"⚠️ Large dataset ({total_cells:,} cells). Showing first 100 rows with styling. Download full data below.")
            df_display_page = df_display.head(100)
            styled = df_display_page.style
            for d in dates_in_range:
                if d in df_display_page.columns:
                    styled = styled.map(color_status_value, subset=[d])
            st.dataframe(styled, use_container_width=True, height=500)
        else:
            styled = df_display.style
            for d in dates_in_range:
                if d in df_display.columns:
                    styled = styled.map(color_status_value, subset=[d])
            st.dataframe(styled, use_container_width=True, height=500)
        
        render_colored_excel_download(
            "📥 Download Status Matrix (Excel, color-coded)",
            df_display,
            f"status_matrix_{old_date}_to_{new_date}",
            key="string_matrix_download",
            status_cols=dates_in_range,
            value_cols=dates_in_range
        )

    with sub_tab2:
        """Status Change Summary - 6 categories with expandable details."""
        if diff_error:
            st.error(diff_error)
            return
        
        df_changes = diff_result["df_string_changes"]
        
        # Apply plot/block filters
        if selected_plot != "All" and "Plot" in df_changes.columns:
            df_changes = df_changes[df_changes["Plot"] == selected_plot]
        if selected_block != "All" and "Block" in df_changes.columns:
            df_changes = df_changes[df_changes["Block"] == selected_block]
        if selected_grid != "All" and "Grid" in df_changes.columns:
            df_changes = df_changes[df_changes["Grid"] == selected_grid]
        
        
        if df_changes.empty:
            st.info("No data available for the selected filters.")
            return
        
        summary = pd.DataFrame({
            "Category": [
                "Strings Restored (Red -> Working)",
                "Strings Currently Hard Failed (Red)",
                "Strings Currently Low Performance (Yellow)",
                "Strings Regressed (Yellow -> Red)",
                "Strings Regressed (Working -> Red)",
                "Strings Regressed (Working -> Yellow)",
            ],
            "Count": [
                int(((df_changes["Old Status"] == "Red - Failed") & (df_changes["New Status"] == "Green - Working")).sum()),
                int((df_changes["New Status"] == "Red - Failed").sum()),
                int((df_changes["New Status"] == "Yellow - Low Performance").sum()),
                int(((df_changes["Old Status"] == "Yellow - Low Performance") & (df_changes["New Status"] == "Red - Failed")).sum()),
                int(((df_changes["Old Status"] == "Green - Working") & (df_changes["New Status"] == "Red - Failed")).sum()),
                int(((df_changes["Old Status"] == "Green - Working") & (df_changes["New Status"] == "Yellow - Low Performance")).sum()),
            ],
        })
        st.markdown(f"#### Status Changes: {old_date} → {new_date}")
        
        # Summary bar chart
        fig_summary = px.bar(
            summary, x="Count", y="Category", orientation="h", text="Count",
            color="Category",
            color_discrete_map={
                "Strings Restored (Red -> Working)": "#10b981",
                "Strings Currently Hard Failed (Red)": "#ef4444",
                "Strings Currently Low Performance (Yellow)": "#fbbf24",
                "Strings Regressed (Yellow -> Red)": "#f97316",
                "Strings Regressed (Working -> Red)": "#dc2626",
                "Strings Regressed (Working -> Yellow)": "#f59e0b",
            },
        )
        fig_summary.update_layout(height=360, showlegend=False, plot_bgcolor="rgba(0,0,0,0)",
                                   paper_bgcolor="rgba(0,0,0,0)", yaxis_title="", xaxis_title="Strings")
        fig_summary.update_traces(textposition="outside")
        st.plotly_chart(fig_summary, use_container_width=True, key="string_matrix_summary_chart")
        
        # Define the 6 categories
        categories = [
            ("1. Recovered from Hard Failure", "🟢",
             "Restored",
             (df_changes["Old Status"] == "Red - Failed") & (df_changes["New Status"] == "Green - Working")),
            ("2. Currently Hard Failed ", "🔴",
             "Hard_Failed",
             df_changes["New Status"] == "Red - Failed"),
            ("3. Currently Low Performance (Yellow)", "🟡",
             "Low_Performance",
             df_changes["New Status"] == "Yellow - Low Performance"),
            ("4. Regressed: Low Performance → Hard Failure (Yellow → Red)", "🟠",
             "Regressed_Yellow_Red",
             (df_changes["Old Status"] == "Yellow - Low Performance") & (df_changes["New Status"] == "Red - Failed")),
            ("5. Regressed: Working → Hard Failure (Working → Red)", "🔻",
             "Regressed_Working_Red",
             (df_changes["Old Status"] == "Green - Working") & (df_changes["New Status"] == "Red - Failed")),
            ("6. Regressed: Working → Low Performance ", "🔸",
             "Regressed_Working_Yellow",
             (df_changes["Old Status"] == "Green - Working") & (df_changes["New Status"] == "Yellow - Low Performance")),
        ]
        
        display_cols = [c for c in ["Plot", "Block", "Grid", "String Inverter", "MPPT PV No",
                                     "Old PV Value", "New PV Value", "Old Status", "New Status"] 
                        if c in df_changes.columns]
        
        st.markdown("---")
        for title, icon, category_name, mask in categories:
            subset = df_changes[mask][display_cols].copy()
            with st.expander(f"{icon} {title} - {len(subset)} string(s)", 
                            expanded=len(subset) > 0 and len(subset) <= 20):
                if subset.empty:
                    st.caption("None.")
                else:
                    page = paginate_dataframe(subset, page_size=20, key_prefix=f"string_matrix_{category_name}")
                    st.dataframe(page, use_container_width=True)
                    
                    if not subset.empty:
                        file_prefix = f"{category_name}_{old_date}_to_{new_date}"
                        render_colored_excel_download(
                            f"📥 Download {title} (Excel, color-coded)",
                            subset,
                            file_prefix,
                            key=f"download_{category_name}",
                            pv_value_cols=["Old PV Value", "New PV Value"],
                            value_cols=["Old PV Value", "New PV Value"]
                        )
        
        st.markdown("---")
        # Full download for status change summary
        render_colored_excel_download(
            "📥 Download Full Status Change Report (Excel, color-coded)",
            df_changes[display_cols],
            f"full_status_change_{old_date}_to_{new_date}",
            key="string_matrix_full_download",
            pv_value_cols=["Old PV Value", "New PV Value"],
            value_cols=["Old PV Value", "New PV Value"]
        )

    with sub_tab3:
        """TAT Report - Failure/Restoration analysis."""
        if df_tat_report.empty:
            st.info("No TAT data available for the selected range.")
            return
        
        st.markdown("**TAT & Failure Analysis Report**")
        st.caption("Shows failure/restoration counts, TAT hours, and current failure hours for each string.")
        
        tat_display_cols = [c for c in ["Plot", "Block", "Grid", "SACU", "Inverter", "String",
                                         "Failure Count", "Restored Count", "TAT (Hours)",
                                         "Current Failure (Hours)", "Current Failure (Mins)"] 
                            if c in df_tat_report.columns]
        
        for d in dates_in_range:
            if f"{d} - Value (A)" in df_tat_report.columns:
                tat_display_cols.append(f"{d} - Value (A)")
        
        df_tat_display = df_tat_report[tat_display_cols].copy()
        
        def color_tat(val):
            if isinstance(val, (int, float)):
                if val == 0:
                    return "background-color: #10b981; color: white; font-weight: bold;"
                elif val > 0:
                    return "background-color: #ef4444; color: white; font-weight: bold;"
            return ""
        
        def color_count(val):
            if isinstance(val, (int, float)):
                if val > 0:
                    return "background-color: #fbbf24; color: black; font-weight: bold;"
            return ""
        
        tat_cells = len(df_tat_display) * len(df_tat_display.columns)
        if tat_cells > 250000:
            st.warning(f"⚠️ Large dataset ({tat_cells:,} cells). Showing first 100 rows with styling. Download full data below.")
            df_tat_page = df_tat_display.head(100)
            styled_tat = df_tat_page.style
            styled_tat = styled_tat.map(color_tat, subset=["TAT (Hours)", "Current Failure (Hours)", "Current Failure (Mins)"])
            styled_tat = styled_tat.map(color_count, subset=["Failure Count", "Restored Count"])
            for d in dates_in_range:
                col_name = f"{d} - Value (A)"
                if col_name in df_tat_page.columns:
                    styled_tat = styled_tat.map(lambda x: f"background-color: {_get_pv_value_color(x)}; color: white; font-weight: bold;" 
                                                if x is not None and x != "" and isinstance(x, (int, float)) else "", 
                                                subset=[col_name])
            st.dataframe(styled_tat, use_container_width=True, height=450)
        else:
            styled_tat = df_tat_display.style
            styled_tat = styled_tat.map(color_tat, subset=["TAT (Hours)", "Current Failure (Hours)", "Current Failure (Mins)"])
            styled_tat = styled_tat.map(color_count, subset=["Failure Count", "Restored Count"])
            for d in dates_in_range:
                col_name = f"{d} - Value (A)"
                if col_name in df_tat_display.columns:
                    styled_tat = styled_tat.map(lambda x: f"background-color: {_get_pv_value_color(x)}; color: white; font-weight: bold;" 
                                                if x is not None and x != "" and isinstance(x, (int, float)) else "", 
                                                subset=[col_name])
            st.dataframe(styled_tat, use_container_width=True, height=450)
        
        total_failures = df_tat_report["Failure Count"].sum()
        total_restored = df_tat_report["Restored Count"].sum()
        total_tat = df_tat_report["TAT (Hours)"].sum()
        total_current_failure = df_tat_report["Current Failure (Hours)"].sum()
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Failures", int(total_failures))
        col2.metric("Total Restored", int(total_restored))
        col3.metric("Total TAT (Hrs)", f"{total_tat:.1f}")
        col4.metric("Total Current Failure (Hrs)", f"{total_current_failure:.1f}")
        
        render_colored_excel_download(
            "📥 Download TAT Report (Excel, color-coded)",
            df_tat_display,
            f"tat_report_{old_date}_to_{new_date}",
            key="tat_report_download",
            value_cols=["TAT (Hours)", "Current Failure (Hours)", "Current Failure (Mins)", "Failure Count", "Restored Count"],
            pv_value_cols=[f"{d} - Value (A)" for d in dates_in_range if f"{d} - Value (A)" in df_tat_display.columns]
        )

    with sub_tab4:
        if diff_error:
            st.error(diff_error)
            return

        df_restore = diff_result["df_string_changes"].copy()
        if selected_plot != "All" and "Plot" in df_restore.columns:
            df_restore = df_restore[df_restore["Plot"] == selected_plot]
        if selected_block != "All" and "Block" in df_restore.columns:
            df_restore = df_restore[df_restore["Block"] == selected_block]
        if selected_grid != "All" and "Grid" in df_restore.columns:
            df_restore = df_restore[df_restore["Grid"] == selected_grid]

        restored_mask = (
            (df_restore["Old Status"] == "Red - Failed")
            & (df_restore["New Status"] == "Green - Working")
        )
        restored = df_restore[restored_mask].copy()

        st.markdown("**Restored Strings**")
        st.caption(f"Strings that moved from failed on {old_date} to working on {new_date}.")
        st.metric("Restored strings", len(restored))

        restore_cols = [c for c in [
            "Plot", "Block", "Grid", "String Inverter", "MPPT PV No",
            "Old PV Value", "New PV Value", "Old Status", "New Status",
            "TAT - Hard Failure Recovery (Working Hours)",
        ] if c in restored.columns]

        if restored.empty:
            st.info("No restored strings found for the selected range and filters.")
        else:
            restored_display = restored[restore_cols]
            page_restore = paginate_dataframe(restored_display, page_size=25, key_prefix="string_matrix_restore")
            st.dataframe(page_restore, width="stretch",
                         column_config=build_grid_column_config(page_restore))

            render_colored_excel_download(
                "Download Restored Strings (Excel, color-coded)",
                restored_display,
                f"restored_strings_{old_date}_to_{new_date}",
                key="string_matrix_restore_download",
                pv_value_cols=["Old PV Value", "New PV Value"],
                value_cols=["Old PV Value", "New PV Value", "TAT - Hard Failure Recovery (Working Hours)"]
            )

# ==========================================
# SNAPSHOT-TO-SNAPSHOT COMPARISON
# ==========================================
def _pick_old_new_dates(key_prefix):
    """Shared Old/New snapshot date picker used by the comparison-based tabs."""
    available_dates = get_available_snapshot_dates()
    if len(available_dates) < 2:
        st.info("📌 At least 2 saved snapshot dates are required for this comparison.")
        return None, None
    min_date = datetime.strptime(available_dates[0], "%Y-%m-%d").date()
    max_date = datetime.strptime(available_dates[-1], "%Y-%m-%d").date()
    sorted_dates = sorted(available_dates)
    default_old_date = datetime.strptime(sorted_dates[-2], "%Y-%m-%d").date()

    col1, col2 = st.columns(2)
    with col1:
        old_date = st.date_input("From Date", value=default_old_date, min_value=min_date, max_value=max_date, key=f"{key_prefix}_old")
    with col2:
        new_date = st.date_input("To Date", value=max_date, min_value=min_date, max_value=max_date, key=f"{key_prefix}_new")

    if old_date >= new_date:
        st.error("❌ From Date must be strictly before To Date.")
        return None, None
    return str(old_date), str(new_date)


# def _classify_pv_strings_long(df, id_cols, pv_cols):
#     """Classifies every PV string in `df` into the 4-status scheme."""
#     present_cols = [c for c in pv_cols if c in df.columns]
#     id_cols_present = [c for c in id_cols if c in df.columns]
#     if not present_cols or not id_cols_present:
#         return pd.DataFrame()

#     numeric = df[present_cols].apply(pd.to_numeric, errors="coerce")
#     working_mask = numeric > WORKING_CURRENT_THRESHOLD
#     row_avg_working = numeric.where(working_mask).mean(axis=1)
#     low_perf_threshold = row_avg_working * 0.8
#     low_perf_mask = working_mask & numeric.lt(low_perf_threshold, axis=0)

#     status = pd.DataFrame("Green - Working", index=numeric.index, columns=present_cols)
#     status[numeric <= WORKING_CURRENT_THRESHOLD] = "Red - Failed"
#     status[low_perf_mask] = "Yellow - Low Performance"
#     status[numeric < 0] = "Blinking Red - Negative Value"

#     context = df[id_cols_present].copy()
#     long_values = numeric.join(context).reset_index(drop=True).melt(
#         id_vars=id_cols_present, value_vars=present_cols, var_name="MPPT PV No", value_name="PV Value",
#     )
#     long_status = status.join(context).reset_index(drop=True).melt(
#         id_vars=id_cols_present, value_vars=present_cols, var_name="MPPT PV No", value_name="Status",
#     )
#     long_df = long_values.merge(long_status, on=id_cols_present + ["MPPT PV No"])
#     return long_df.dropna(subset=["PV Value"]).reset_index(drop=True)

def _classify_pv_strings_long(df, id_cols, pv_cols):
    """
    Classifies every PV string in `df` into the 4-status scheme.
    Uses per-inverter average working current for low performance detection.
    """
    import numpy as np
    
    present_cols = [c for c in pv_cols if c in df.columns]
    id_cols_present = [c for c in id_cols if c in df.columns]
    if not present_cols or not id_cols_present:
        return pd.DataFrame()

    numeric = df[present_cols].apply(pd.to_numeric, errors="coerce")
    
    # Create inverter key for each row
    inverter_keys = df[id_cols_present].apply(
        lambda row: tuple(row.values), axis=1
    )
    
    # Calculate per-inverter average working current
    inverter_avg_map = {}
    inverter_working_counts = {}
    
    for idx, row in df.iterrows():
        inv_key = inverter_keys[idx]
        if inv_key not in inverter_avg_map:
            # Get all rows for this inverter
            inv_mask = inverter_keys == inv_key
            inv_numeric = numeric[inv_mask]
            
            # Get working values (> threshold)
            working_vals = inv_numeric[inv_numeric > WORKING_CURRENT_THRESHOLD].stack().values
            
            if len(working_vals) > 0:
                inverter_avg_map[inv_key] = np.mean(working_vals)
                inverter_working_counts[inv_key] = len(working_vals)
            else:
                inverter_avg_map[inv_key] = 0
                inverter_working_counts[inv_key] = 0
    
    # Create series for inverter averages
    inverter_avg_series = pd.Series(
        [inverter_avg_map.get(key, 0) for key in inverter_keys],
        index=df.index
    )
    
    # Low performance threshold: 80% of inverter's average working current
    low_perf_threshold = inverter_avg_series * 0.8
    
    # Determine status for each PV column
    status = pd.DataFrame("Green - Working", index=numeric.index, columns=present_cols)
    
    # Red - Failed: value <= threshold
    status[numeric <= WORKING_CURRENT_THRESHOLD] = "Red - Failed"
    
    # Yellow - Low Performance: working but below 80% of inverter average
    # Only apply if inverter has at least one working string
    working_mask = numeric > WORKING_CURRENT_THRESHOLD
    low_perf_mask = working_mask & numeric.lt(low_perf_threshold, axis=0)
    # Don't mark as low performance if inverter has no working strings
    no_working_mask = inverter_avg_series == 0
    low_perf_mask = low_perf_mask & ~no_working_mask.values[:, None]
    status[low_perf_mask] = "Yellow - Low Performance"
    
    # Blinking Red - Negative Value (highest priority)
    status[numeric < 0] = "Blinking Red - Negative Value"

    # Convert to long format
    context = df[id_cols_present].copy()
    long_values = numeric.join(context).reset_index(drop=True).melt(
        id_vars=id_cols_present, value_vars=present_cols, 
        var_name="MPPT PV No", value_name="PV Value",
    )
    long_status = status.join(context).reset_index(drop=True).melt(
        id_vars=id_cols_present, value_vars=present_cols, 
        var_name="MPPT PV No", value_name="Status",
    )
    long_df = long_values.merge(long_status, on=id_cols_present + ["MPPT PV No"])
    return long_df.dropna(subset=["PV Value"]).reset_index(drop=True)

@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def build_snapshot_diff(old_date, new_date, sheet_name="Sheet1"):
    """Core comparison engine shared by all tabs."""
    df_old = load_snapshot_sheet(old_date, sheet_name)
    df_new = load_snapshot_sheet(new_date, sheet_name)
    if df_old is None or df_new is None:
        return None, f"No saved snapshot data found for '{sheet_name}' on {old_date} and/or {new_date}."

    inverter_col = get_inverter_column(df_old) or get_inverter_column(df_new)
    if not inverter_col:
        return None, "Could not detect an inverter identifier column in the snapshots."

    grid_col = get_grid_column(df_old) or get_grid_column(df_new)
    id_candidates = ["Plot", "Block"]
    if grid_col:
        id_candidates.append(grid_col)
    id_candidates.append(inverter_col)
    id_cols = [c for c in id_candidates if c in df_old.columns and c in df_new.columns]
    pv_cols = get_pv_current_columns(df_old)

    try:
        delta_days = (datetime.strptime(str(new_date), "%Y-%m-%d") - datetime.strptime(str(old_date), "%Y-%m-%d")).days
    except Exception:
        delta_days = 0
    if delta_days <= 0:
        return None, "New Date must be strictly after Old Date."

    working_hours_used = delta_days * WORKING_HOURS_PER_DAY
    working_minutes_used = working_hours_used * 60

    df_old_prep = df_old.copy()
    df_new_prep = df_new.copy()
    
    for prep_df in (df_old_prep, df_new_prep):
        for col in id_cols:
            prep_df[col] = prep_df[col].astype(str).str.strip()
        prep_df["Total Active Strings"] = prep_df.apply(
            lambda row: get_total_active_strings(row.get("Plot"), row.get("Block")), axis=1
        )
        prep_df.drop_duplicates(subset=id_cols, keep="last", inplace=True)

    old_long = _classify_pv_strings_long(df_old_prep, id_cols, pv_cols).rename(
        columns={"PV Value": "Old PV Value", "Status": "Old Status"})
    new_long = _classify_pv_strings_long(df_new_prep, id_cols, pv_cols).rename(
        columns={"PV Value": "New PV Value", "Status": "New Status"})

    join_cols = id_cols + ["MPPT PV No"]
    df_string_changes = pd.merge(old_long, new_long, on=join_cols, how="inner")
    rename_map = {inverter_col: "String Inverter"}
    if grid_col and grid_col != "Grid" and grid_col in df_string_changes.columns:
        rename_map[grid_col] = "Grid"
    df_string_changes = df_string_changes.rename(columns=rename_map)

    recovered_hard_to_working = df_string_changes[
        (df_string_changes["Old Status"] == "Red - Failed") & (df_string_changes["New Status"] == "Green - Working")]
    currently_hard_failed = df_string_changes[df_string_changes["New Status"] == "Red - Failed"]
    currently_low_performance = df_string_changes[df_string_changes["New Status"] == "Yellow - Low Performance"]
    regressed_lp_to_hard_failure = df_string_changes[
        (df_string_changes["Old Status"] == "Yellow - Low Performance") & (df_string_changes["New Status"] == "Red - Failed")]
    regressed_to_hard_failure = df_string_changes[
        (df_string_changes["Old Status"] == "Green - Working") & (df_string_changes["New Status"] == "Red - Failed")]
    regressed_to_low_performance = df_string_changes[
        (df_string_changes["Old Status"] == "Green - Working") & (df_string_changes["New Status"] == "Yellow - Low Performance")]

    status_change_summary = pd.DataFrame({
        "Category": [
            "Strings Restored (Red -> Working)",
            "Strings Currently Hard Failed (Red)",
            "Strings Currently Low Performance (Yellow)",
            "Strings Regressed (Yellow -> Red)",
            "Strings Regressed (Working -> Red)",
            "Strings Regressed (Working -> Yellow)",
        ],
        "Count": [
            len(recovered_hard_to_working), len(currently_hard_failed), len(currently_low_performance),
            len(regressed_lp_to_hard_failure), len(regressed_to_hard_failure), len(regressed_to_low_performance),
        ],
    })

    df_string_changes["TAT - Hard Failure Recovery (Working Hours)"] = 0.0
    df_string_changes["Current Failure (Working Mins)"] = 0.0
    df_string_changes.loc[recovered_hard_to_working.index, "TAT - Hard Failure Recovery (Working Hours)"] = working_hours_used
    currently_failed_mask = df_string_changes["New Status"].isin(
        ["Red - Failed", "Yellow - Low Performance", "Blinking Red - Negative Value"])
    df_string_changes.loc[currently_failed_mask, "Current Failure (Working Mins)"] = working_minutes_used

    tat_summary = pd.DataFrame({
        "Metric": ["Strings Restored (Hard Failure -> Working)", "Strings Currently Failed / Low Performance"],
        "Count": [len(recovered_hard_to_working), int(currently_failed_mask.sum())],
        "Total": [
            f"{len(recovered_hard_to_working) * working_hours_used:,.0f} Hrs",
            f"{int(currently_failed_mask.sum()) * working_minutes_used:,.0f} Mins",
        ],
    })

    rename_cols = ["Failed String Count", "Total Active Strings", "Working String Count"]
    old_counts = df_old_prep[id_cols + rename_cols].rename(columns={c: f"{c}_Old" for c in rename_cols})
    new_counts = df_new_prep[id_cols + rename_cols].rename(columns={c: f"{c}_New" for c in rename_cols})
    counts_merged = pd.merge(old_counts, new_counts, on=id_cols, how="inner")
    for suffix in ["_Old", "_New"]:
        for c in rename_cols:
            counts_merged[f"{c}{suffix}"] = pd.to_numeric(counts_merged[f"{c}{suffix}"], errors="coerce").fillna(0).astype(int)

    changed_inverters = counts_merged[
        counts_merged["Working String Count_Old"] != counts_merged["Working String Count_New"]
    ].copy()
    changed_inverters = changed_inverters.rename(columns=rename_map)
    changed_inverters["Change"] = changed_inverters["Working String Count_New"] - changed_inverters["Working String Count_Old"]
    changed_inverters = changed_inverters.sort_values("Working String Count_Old", ascending=True).reset_index(drop=True)

    def _overall(df_prep):
        total_strings = pd.to_numeric(df_prep.get("Total Active Strings"), errors="coerce").fillna(0).sum()
        working = pd.to_numeric(df_prep.get("Working String Count"), errors="coerce").fillna(0).sum()
        failed = pd.to_numeric(df_prep.get("Failed String Count"), errors="coerce").fillna(0).sum()
        return {
            "Total Inverters": df_prep[inverter_col].nunique() if inverter_col in df_prep.columns else len(df_prep),
            "Total Active Strings": int(total_strings),
            "Total Working Strings": int(working),
            "Total Failed Strings": int(failed),
            "Availability (%)": round((working / total_strings) * 100, 2) if total_strings else 0.0,
        }

    return {
        "df_string_changes": df_string_changes,
        "changed_inverters": changed_inverters,
        "status_change_summary": status_change_summary,
        "tat_summary": tat_summary,
        "overall_old": _overall(df_old_prep),
        "overall_new": _overall(df_new_prep),
        "working_hours_used": working_hours_used,
        "working_minutes_used": working_minutes_used,
        "delta_days": delta_days,
        "inverter_col": inverter_col,
    }, None


def compare_two_snapshots_by_date(old_date, new_date, sheet_name="Sheet1"):
    """Per-inverter roll-up of the old_date -> new_date string-level diff."""
    diff, error = build_snapshot_diff(old_date, new_date, sheet_name)
    if error:
        return None, error

    inverter_col = "String Inverter"
    dfc = diff["df_string_changes"].copy()
    group_cols = [c for c in ["Plot", "Block", "Grid", inverter_col] if c in dfc.columns]
    if not group_cols or dfc.empty:
        return None, "No comparable string data found for this date pair."

    dfc["_f2w"] = ((dfc["Old Status"] == "Red - Failed") & (dfc["New Status"] == "Green - Working")).astype(int)
    dfc["_w2f"] = ((dfc["Old Status"] == "Green - Working") &
                   (dfc["New Status"].isin(["Red - Failed", "Yellow - Low Performance"]))).astype(int)
    dfc["_old_failed"] = dfc["Old Status"].isin(
        ["Red - Failed", "Yellow - Low Performance", "Blinking Red - Negative Value"]).astype(int)
    dfc["_new_failed"] = dfc["New Status"].isin(
        ["Red - Failed", "Yellow - Low Performance", "Blinking Red - Negative Value"]).astype(int)

    counts = dfc.groupby(group_cols, as_index=False).agg(
        Failed_to_Working=("_f2w", "sum"),
        Working_to_Failed=("_w2f", "sum"),
        **{"Failed String Count_old": ("_old_failed", "sum")},
        **{"Failed String Count_new": ("_new_failed", "sum")},
    )

    tat_cols = {}
    if "TAT - Hard Failure Recovery (Working Hours)" in dfc.columns:
        tat_cols["Restoration_TAT_Hours"] = ("TAT - Hard Failure Recovery (Working Hours)", "sum")
    if "Current Failure (Working Mins)" in dfc.columns:
        tat_cols["_current_failure_mins"] = ("Current Failure (Working Mins)", "sum")
    tat_grouped = dfc.groupby(group_cols, as_index=False).agg(**tat_cols) if tat_cols else None
    if tat_grouped is not None:
        if "_current_failure_mins" in tat_grouped.columns:
            tat_grouped["Current_Failure_Hours"] = tat_grouped["_current_failure_mins"] / 60.0
            tat_grouped = tat_grouped.drop(columns=["_current_failure_mins"])
        counts = counts.merge(tat_grouped, on=group_cols, how="left")
    for col in ("Restoration_TAT_Hours", "Current_Failure_Hours"):
        if col not in counts.columns:
            counts[col] = 0.0

    def _joined_ids_table(mask_col, out_name):
        subset = dfc[dfc[mask_col] == 1]
        if subset.empty or "MPPT PV No" not in subset.columns:
            return pd.DataFrame(columns=group_cols + [out_name])
        return (
            subset.groupby(group_cols)["MPPT PV No"]
            .apply(lambda ids: ", ".join(str(x) for x in ids))
            .reset_index(name=out_name)
        )

    restored_names = _joined_ids_table("_f2w", "Restored_Strings")
    failed_names = _joined_ids_table("_w2f", "Newly_Failed_Strings")

    df_history = counts.merge(restored_names, on=group_cols, how="left").merge(failed_names, on=group_cols, how="left")
    df_history["Restored_Strings"] = df_history["Restored_Strings"].fillna("")
    df_history["Newly_Failed_Strings"] = df_history["Newly_Failed_Strings"].fillna("")
    df_history[["Restoration_TAT_Hours", "Current_Failure_Hours"]] = df_history[
        ["Restoration_TAT_Hours", "Current_Failure_Hours"]
    ].fillna(0.0)

    working_vals = pd.to_numeric(
        dfc.loc[dfc["New Status"] == "Green - Working", "New PV Value"], errors="coerce"
    ) if "New PV Value" in dfc.columns else pd.Series(dtype=float)
    baseline_avg_working_current = float(working_vals.mean()) if len(working_vals.dropna()) else 0.0

    return {
        "df_history": df_history,
        "inverter_col": inverter_col,
        "baseline_avg_working_current": baseline_avg_working_current,
    }, None


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def build_range_trend_data_cached(from_date, to_date, sheet_name="Sheet1"):
    """Cached version of build_range_trend_data."""
    from_str, to_str = str(from_date), str(to_date)
    dates_in_range = [d for d in storage.get_available_snapshot_dates() if from_str <= d <= to_str]

    rows = []
    for d in sorted(dates_in_range):
        df = load_snapshot_sheet(d, sheet_name)
        if df is None or df.empty:
            continue
        needed = {"Plot", "Block", "Working String Count", "Failed String Count", "Total Active Strings"}
        if not needed.issubset(df.columns):
            continue

        group_cols = [c for c in ["Plot", "Block", get_grid_column(df)] if c and c in df.columns]
        grouped = df.groupby(group_cols, as_index=False).agg(
            Working=("Working String Count", "sum"),
            Failed=("Failed String Count", "sum"),
            Total=("Total Active Strings", "sum"),
        )
        grid_col = get_grid_column(grouped)
        if grid_col and grid_col != "Grid":
            grouped = grouped.rename(columns={grid_col: "Grid"})
        grouped["Date"] = d
        grouped["Availability (%)"] = (grouped["Working"] / grouped["Total"] * 100).fillna(0).round(2)
        rows.append(grouped)

    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True)


def build_range_trend_data(from_date, to_date, sheet_name="Sheet1"):
    """Build trend data with caching."""
    return build_range_trend_data_cached(from_date, to_date, sheet_name)


# ==========================================
# UI FUNCTIONS
# ==========================================
def display_upload_registry(user_role, upload_handler=None):
    """Upload registry + health check with fixed backup/restore."""
    st.subheader("📁 Snapshot Upload Registry")
    st.caption(
        "SCADA workbooks are uploaded from the main sidebar (admin only) and "
        "automatically saved here, day-wise, so they can be compared later."
    )

    if can_upload(user_role) and upload_handler:
        with st.expander("📤 Backfill a Previous Date's Snapshot", expanded=False):
            st.caption(
                "Upload a SCADA workbook for any past calendar date. It will be "
                "processed and stored exactly as if it had been uploaded that day."
            )
            backfill_date = st.date_input(
                "Snapshot Date", value=datetime.now().date(),
                max_value=datetime.now().date(), key="restore_backfill_date",
            )
            backfill_file = st.file_uploader(
                "SCADA Excel (.xlsx)", type=["xlsx"], key="restore_backfill_file",
            )
            if backfill_file is not None and st.button("🔄 Process & Save Snapshot", key="restore_backfill_btn", width="stretch"):
                with st.spinner("Processing snapshot..."):
                    ok, msg = upload_handler(backfill_file.getvalue(), backfill_file.name, backfill_date)
                if ok:
                    st.success(msg)
                    get_available_snapshot_dates_cached.clear()
                    load_snapshot_sheet_cached.clear()
                    st.rerun()
                else:
                    st.error(msg)

    uploads = storage.get_all_uploads()
    if not uploads:
        st.info("📌 No snapshots uploaded yet.")
        return

    with st.spinner("🔍 Checking registry integrity..."):
        report = storage.get_upload_registry_report()

    ok_count = sum(1 for r in report if r["integrity_ok"])
    broken_count = len(report) - ok_count

    h1, h2, h3 = st.columns(3)
    h1.metric("Registered Uploads", len(report))
    h2.metric("Healthy", ok_count)
    h3.metric("Broken / Missing Files", broken_count)

    if broken_count > 0:
        st.error(
            f"⚠️ {broken_count} registry entr{'y is' if broken_count == 1 else 'ies are'} missing their backing "
            f"file(s) on disk. Re-upload the affected date(s)."
        )
    else:
        st.success("✅ Upload registry is healthy - every entry's files are present on disk.")

    df_report = pd.DataFrame(report)[
        ["upload_id", "snapshot_date", "original_filename", "uploaded_by", "upload_timestamp",
         "saved_sheets", "integrity_ok", "integrity_message"]
    ].sort_values("snapshot_date", ascending=False)

    display_report = df_report.rename(columns={
        "snapshot_date": "Snapshot Date", "original_filename": "File Name",
        "uploaded_by": "Uploaded By", "upload_timestamp": "Upload Time",
        "saved_sheets": "Sheets Saved", "integrity_ok": "Files OK", "integrity_message": "Integrity Detail",
    })

    def color_integrity(val):
        if val is True:
            return "background-color: #10b981; color: white; font-weight: bold;"
        if val is False:
            return "background-color: #ef4444; color: white; font-weight: bold;"
        return ""

    st.dataframe(
        display_report.drop(columns=["upload_id"]).style.map(color_integrity, subset=["Files OK"]),
        use_container_width=True,
    )

    # Admin-only: view/delete snapshots
    if can_upload(user_role):
        st.markdown("---")
        st.markdown("#### Manage a Snapshot")

        snapshot_labels = [
            f"{row['Snapshot Date']} · {row['File Name']} (uploaded by {row['Uploaded By']})"
            for _, row in display_report.iterrows()
        ]
        label_to_upload_id = dict(zip(snapshot_labels, display_report["upload_id"]))

        if snapshot_labels:
            selected_label = st.selectbox("Select a snapshot", snapshot_labels, key="registry_manage_select")
            selected_upload_id = label_to_upload_id[selected_label]
            entry = storage.get_upload_by_id(selected_upload_id)

            with st.expander("👁️ View Snapshot Data", expanded=False):
                if entry and entry.get("saved_sheets"):
                    sheet_pick = st.selectbox(
                        "Sheet", entry["saved_sheets"], key=f"registry_view_sheet_{selected_upload_id}",
                    )
                    with st.spinner("Loading sheet data..."):
                        df_view = storage.load_sheet_csv(selected_upload_id, sheet_pick)
                    if df_view is not None:
                        st.dataframe(df_view, use_container_width=True, height=400)
                    else:
                        st.warning("Could not load this sheet's saved data.")
                else:
                    st.info("No saved sheets found for this snapshot.")

            mcol1, mcol2 = st.columns([3, 1])
            with mcol2:
                confirm_delete = st.checkbox("Confirm", key=f"confirm_delete_{selected_upload_id}")
                if st.button("🗑️ Delete Snapshot", key=f"delete_upload_{selected_upload_id}",
                             disabled=not confirm_delete, width="stretch", type="primary"):
                    ok, msg = storage.delete_upload(selected_upload_id)
                    storage.log_audit_event(
                        st.session_state.get("user", {}).get("username", "unknown"),
                        user_role, "snapshot_deleted", {"upload_id": selected_upload_id, "result": msg},
                    )
                    if ok:
                        st.success(msg)
                        get_available_snapshot_dates_cached.clear()
                        load_snapshot_sheet_cached.clear()
                        st.rerun()
                    else:
                        st.error(msg)

        st.markdown("---")
        st.markdown("#### 💾 Full Data & Website Backup")
        st.caption(
            "This app has no database - every snapshot, user account, and history file "
            "lives on disk under ./data. Download a full backup regularly."
        )

        # Get backup manifest
        manifest = get_backup_manifest()
        
        # Show backup info
        b1, b2, b3 = st.columns(3)
        b1.metric("Files to Backup", manifest["total_files"])
        b2.metric("Backup Size", f"{manifest['total_size_kb']:,.0f} KB")
        if manifest["file_types"]:
            file_types_str = ", ".join([f"{k}: {v}" for k, v in manifest["file_types"].items()])
            b3.metric("File Types", file_types_str[:20] + "..." if len(file_types_str) > 20 else file_types_str)
        else:
            b3.metric("File Types", "None")

        backup_col1, backup_col2 = st.columns(2)

        with backup_col1:
            st.markdown("**Download Backup**")
            if manifest["total_files"] == 0:
                st.warning("No data files found to backup. Upload some data first.")
            
            if st.button("📦 Prepare Full Backup (.zip)", key="full_backup_prepare_btn", width="stretch"):
                with st.spinner(f"Zipping up {manifest['total_files']} files..."):
                    backup_bytes = export_full_backup_bytes()
                    if backup_bytes:
                        st.session_state["full_backup_bytes"] = backup_bytes
                        storage.log_audit_event(
                            st.session_state.get("user", {}).get("username", "unknown"),
                            user_role, "full_backup_exported", {"files": manifest["total_files"]},
                        )
                        st.success(f"✅ Backup prepared successfully! {manifest['total_files']} files included.")
                    else:
                        st.error("❌ Failed to create backup. No files were included.")
            
            if st.session_state.get("full_backup_bytes"):
                file_size = len(st.session_state["full_backup_bytes"]) / 1024
                st.caption(f"Backup size: {file_size:.1f} KB")
                st.download_button(
                    "⬇️ Download Full Backup", 
                    data=st.session_state["full_backup_bytes"],
                    file_name=f"scada_app_full_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip", 
                    key="full_backup_download_btn", 
                    width="stretch",
                )

        with backup_col2:
            st.markdown("**Restore From Backup**")
            st.caption("Upload a previously downloaded backup .zip to bring the data back.")
            restore_zip = st.file_uploader("Backup file (.zip)", type=["zip"], key="full_backup_restore_upload")
            
            if restore_zip is not None:
                # Check if the zip is valid
                try:
                    import zipfile
                    import io
                    test_buffer = io.BytesIO(restore_zip.getvalue())
                    with zipfile.ZipFile(test_buffer, 'r') as test_zip:
                        file_count = len(test_zip.namelist())
                        st.info(f"Found {file_count} files in the backup.")
                except:
                    st.error("Invalid zip file. Please select a valid backup.")
                    restore_zip = None
            
            overwrite_existing = st.checkbox(
                "⚠️ Overwrite files that already exist",
                key="full_backup_overwrite_checkbox",
            )
            confirm_restore = st.checkbox("✅ I understand and want to restore this backup", key="full_backup_confirm_checkbox")
            
            if restore_zip is not None and st.button(
                "🔁 Restore Backup", 
                key="full_backup_restore_btn",
                disabled=not confirm_restore, 
                width="stretch", 
                type="primary",
            ):
                with st.spinner("Restoring application data..."):
                    restored_count = import_full_backup_bytes(restore_zip.getvalue(), overwrite=overwrite_existing)
                    storage.log_audit_event(
                        st.session_state.get("user", {}).get("username", "unknown"),
                        user_role, "full_backup_restored",
                        {"overwrite": overwrite_existing, "restored_count": restored_count},
                    )
                    
                    if restored_count > 0:
                        get_available_snapshot_dates_cached.clear()
                        load_snapshot_sheet_cached.clear()
                        load_string_history_cached.clear()
                        st.success(f"✅ Restored {restored_count} files successfully!")
                        st.rerun()
                    else:
                        st.warning("⚠️ No files were restored. The backup might be empty or already up to date.")

def display_summary_dashboard(sheet_name="Sheet1"):
    """Last-3-days summary with day-over-day comparisons."""
    st.subheader("📊 Summary Dashboard - Last 3 Days")
    st.caption("Inverter and string counts for the most recent snapshots, with day-over-day restored/failed comparison.")
    with st.expander("ℹ️ How to read this tab", expanded=False):
        st.markdown(
            "- **Daily Snapshot Counts**: for each of the last 3 uploaded dates, how many strings "
            "were Working vs Failed, and the resulting Availability %.\n"
            "- **Day-over-Day Comparison**: looks at each *consecutive pair* of those dates and counts "
            "how many strings **Restored** vs went **Newly Failed** in that gap.\n"
            "- **Restored / Newly Failed - Detail** tables list the exact inverters behind those counts."
        )

    available_dates = get_available_snapshot_dates()
    if not available_dates:
        st.info("📌 No snapshot data available yet.")
        return

    last_dates = sorted(available_dates)[-3:]

    # Per-day counts
    daily_summaries = []
    with st.spinner("📊 Loading recent snapshots..."):
        for d in last_dates:
            df = load_snapshot_sheet(d, sheet_name)
            if df is None or df.empty:
                daily_summaries.append({
                    "Date": d, "Inverters": 0, "Total Strings": 0,
                    "Working": 0, "Failed": 0, "Availability (%)": 0.0,
                })
                continue
            
            # Get inverter column - use cached version
            inverter_col = get_inverter_column(df)
            
            # Count unique inverters - use the correct column
            if inverter_col and inverter_col in df.columns:
                # Convert to string to ensure proper grouping
                total_inverters = df[inverter_col].astype(str).nunique()
            else:
                # Fallback: try to find any inverter-like column
                for col in INVERTER_ID_COLS:
                    if col in df.columns:
                        total_inverters = df[col].astype(str).nunique()
                        inverter_col = col
                        break
                else:
                    total_inverters = 0
            
            # Get string counts - ensure we use the correct columns
            total_strings = int(df["Total Active Strings"].sum()) if "Total Active Strings" in df.columns else 0
            working = int(df["Working String Count"].sum()) if "Working String Count" in df.columns else 0
            failed = int(df["Failed String Count"].sum()) if "Failed String Count" in df.columns else 0
            
            # If Failed String Count is not available, calculate it
            if "Failed String Count" not in df.columns and total_strings > 0:
                failed = total_strings - working
            
            availability = round((working / total_strings * 100), 2) if total_strings > 0 else 0.0
            
            daily_summaries.append({
                "Date": d, 
                "Inverters": total_inverters, 
                "Total Strings": total_strings,
                "Working": working, 
                "Failed": failed, 
                "Availability (%)": availability,
            })

    df_daily = pd.DataFrame(daily_summaries)

    st.markdown("#### Daily Snapshot Counts")
    day_cols = st.columns(len(df_daily))
    for idx, row in df_daily.iterrows():
        with day_cols[idx]:
            st.metric(row["Date"], f"{row['Availability (%)']:.1f}% avail")
            st.caption(f"Inverters: {row['Inverters']:,} | Total Strings: {row['Total Strings']:,}")
            st.write(f"Working: **{row['Working']:,}**  ·  Failed: **{row['Failed']:,}**")

    st.dataframe(df_daily, use_container_width=True, hide_index=True,
                 column_config=build_grid_column_config(df_daily))

    fig_daily = go.Figure()
    fig_daily.add_trace(go.Bar(x=df_daily["Date"], y=df_daily["Working"], name="Working", marker_color="#10b981"))
    fig_daily.add_trace(go.Bar(x=df_daily["Date"], y=df_daily["Failed"], name="Failed", marker_color="#ef4444"))
    fig_daily.update_layout(barmode="group", title="Working vs Failed Strings (Last 3 Days)",
                             height=380, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
    st.plotly_chart(fig_daily, use_container_width=True, key="summary_daily_bar")

    if len(last_dates) < 2:
        st.info("📌 Need at least 2 snapshot dates to calculate restored/failed comparisons.")
        return

    # Day-over-day comparisons
    st.markdown("---")
    st.markdown("#### Day-over-Day Restored / Failed Comparison")

    comparison_rows = []
    restored_detail_frames = []
    failed_detail_frames = []

    with st.spinner("🔄 Comparing consecutive snapshots..."):
        for i in range(1, len(last_dates)):
            prev_date, curr_date = last_dates[i - 1], last_dates[i]
            
            # Get the actual failed count from the snapshot directly
            df_curr = load_snapshot_sheet(curr_date, sheet_name)
            if df_curr is not None and not df_curr.empty:
                if "Failed String Count" in df_curr.columns:
                    current_failed_total = int(df_curr["Failed String Count"].sum())
                else:
                    # Calculate from PV columns if Failed String Count not available
                    pv_cols = get_pv_current_columns(df_curr)
                    total_failed = 0
                    for _, row in df_curr.iterrows():
                        total_active = get_total_active_strings(row.get("Plot"), row.get("Block"))
                        working = 0
                        for col in pv_cols:
                            val = pd.to_numeric(row.get(col), errors="coerce")
                            if pd.notna(val) and val > WORKING_CURRENT_THRESHOLD:
                                working += 1
                        total_failed += max(0, total_active - working)
                    current_failed_total = total_failed
            else:
                current_failed_total = 0
            
            result, error = compare_two_snapshots_by_date(prev_date, curr_date, sheet_name)
            if error:
                st.warning(f"{prev_date} -> {curr_date}: {error}")
                continue

            df_hist = result["df_history"]
            inverter_col = result["inverter_col"]
            pair_label = f"{prev_date} → {curr_date}"

            restored_total = int(df_hist["Failed_to_Working"].sum())
            newly_failed_total = int(df_hist["Working_to_Failed"].sum())

            comparison_rows.append({
                "Comparison": pair_label,
                "Restored Strings": restored_total,
                "Newly Failed Strings": newly_failed_total,
                "Failed as of To-Date": current_failed_total,
            })

            restored_cols = [c for c in ["Plot", "Block", "Grid", inverter_col,
                                          "Failed_to_Working", "Restored_Strings"] if c in df_hist.columns]
            restored_rows = df_hist[df_hist["Failed_to_Working"] > 0][restored_cols].copy()
            if not restored_rows.empty:
                restored_rows.insert(0, "Comparison", pair_label)
                restored_rows = restored_rows.rename(columns={
                    inverter_col: "Inverter", "Failed_to_Working": "Restored Count",
                })
                restored_detail_frames.append(restored_rows)

            failed_cols = [c for c in ["Plot", "Block", "Grid", inverter_col,
                                       "Working_to_Failed", "Newly_Failed_Strings"] if c in df_hist.columns]
            failed_rows = df_hist[df_hist["Working_to_Failed"] > 0][failed_cols].copy()
            if not failed_rows.empty:
                failed_rows.insert(0, "Comparison", pair_label)
                failed_rows = failed_rows.rename(columns={
                    inverter_col: "Inverter", "Working_to_Failed": "Newly Failed Count",
                })
                failed_detail_frames.append(failed_rows)

    if not comparison_rows:
        st.info("📌 No valid day-over-day comparisons could be computed.")
        return

    df_comparison = pd.DataFrame(comparison_rows)

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Restored (window)", int(df_comparison["Restored Strings"].sum()))
    c2.metric("Total Newly Failed (window)", int(df_comparison["Newly Failed Strings"].sum()))
    c3.metric("Failed as of Latest Date", int(df_comparison["Failed as of To-Date"].iloc[-1]))

    st.dataframe(df_comparison, use_container_width=True, hide_index=True,
                 column_config=build_grid_column_config(df_comparison))

    fig_compare = go.Figure()
    fig_compare.add_trace(go.Bar(x=df_comparison["Comparison"], y=df_comparison["Restored Strings"],
                                  name="Restored", marker_color="#10b981"))
    fig_compare.add_trace(go.Bar(x=df_comparison["Comparison"], y=df_comparison["Newly Failed Strings"],
                                  name="Newly Failed", marker_color="#ef4444"))
    fig_compare.add_trace(go.Scatter(x=df_comparison["Comparison"], y=df_comparison["Failed as of To-Date"],
                                      name="Failed as of To-Date", mode="lines+markers",
                                      line=dict(color="#fbbf24", width=3)))
    fig_compare.update_layout(
        barmode="group", title="Restored vs Newly Failed Strings by Day-Pair",
        height=420, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    st.plotly_chart(fig_compare, use_container_width=True, key="summary_compare_chart")

    # Restored strings detail
    st.markdown("---")
    st.markdown("#### ✅ Restored Strings - Detail")
    if restored_detail_frames:
        df_restored = pd.concat(restored_detail_frames, ignore_index=True)
        df_restored = df_restored.sort_values(["Comparison", "Restored Count"], ascending=[True, False])
        page_restored = paginate_dataframe(df_restored, page_size=15, key_prefix="summary_restored")
        st.dataframe(page_restored, use_container_width=True,
                     column_config=build_grid_column_config(page_restored))

        render_colored_excel_download(
            "📥 Download Restored Strings (Excel, color-coded)",
            df_restored,
            "restored_strings",
            key="summary_restored_download",
            value_cols=["Restored Count"]
        )
    else:
        st.success("✅ No strings restored in this window.")

    st.markdown("#### ❌ Newly Failed Strings - Detail")
    if failed_detail_frames:
        df_failed = pd.concat(failed_detail_frames, ignore_index=True)
        df_failed = df_failed.sort_values(["Comparison", "Newly Failed Count"], ascending=[True, False])
        page_failed = paginate_dataframe(df_failed, page_size=15, key_prefix="summary_failed")
        st.dataframe(page_failed, use_container_width=True,
                     column_config=build_grid_column_config(page_failed))

        render_colored_excel_download(
            "📥 Download Newly Failed Strings (Excel, color-coded)",
            df_failed,
            "newly_failed_strings",
            key="summary_failed_download",
            value_cols=["Newly Failed Count"]
        )
    else:
        st.success("✅ No new failures in this window.")
def display_working_hours_analysis():
    """Comprehensive TAT (Turnaround Time) & Current Failure summary."""
    st.subheader("⏰ Working Hours - TAT & Current Failure Summary")
    st.caption(
        f"Working hours window: {WORKING_HOURS_START}:00-{WORKING_HOURS_END}:00 "
        f"({WORKING_HOURS_PER_DAY} hrs/day). TAT and Current Failure figures assume "
        "the full gap between the two chosen snapshots."
    )

    old_date, new_date = _pick_old_new_dates("working_hours")
    if not old_date:
        return

    with st.spinner("⏳ Calculating..."):
        result, error = build_snapshot_diff(old_date, new_date, "Sheet1")
    if error:
        st.error(error)
        return

    st.markdown(f"#### {old_date} → {new_date}  ({result['delta_days']} day(s), {result['working_hours_used']:.0f} working hours)")

    st.markdown("##### Comprehensive String Status Change Summary")
    st.dataframe(result["status_change_summary"], use_container_width=True, hide_index=True,
                 column_config=build_grid_column_config(result["status_change_summary"]))

    render_colored_excel_download(
        "📥 Download Status Change Summary (Excel, color-coded)",
        result["status_change_summary"],
        "status_change_summary",
        key="wh_status_download",
        value_cols=["Count"]
    )

    st.markdown("##### Turnaround Time (TAT) & Current Failure Hours")
    st.dataframe(result["tat_summary"], use_container_width=True, hide_index=True,
                 column_config=build_grid_column_config(result["tat_summary"]))

    render_colored_excel_download(
        "📥 Download TAT Summary (Excel, color-coded)",
        result["tat_summary"],
        "tat_summary",
        key="wh_tat_download",
    )

    st.markdown("---")
    st.markdown("##### Overall Performance Overview")
    ov_old, ov_new = result["overall_old"], result["overall_new"]
    o1, o2 = st.columns(2)
    with o1:
        st.markdown(f"**📅 From Date ({old_date})**")
        st.write(f"Total Inverters: **{ov_old['Total Inverters']:,}**")
        st.write(f"Total Active Strings: **{ov_old['Total Active Strings']:,}**")
        st.write(f"Working Strings: **{ov_old['Total Working Strings']:,}**")
        st.write(f"Failed Strings: **{ov_old['Total Failed Strings']:,}**")
        st.write(f"Availability: **{ov_old['Availability (%)']}%**")
    with o2:
        st.markdown(f"**📅 To Date ({new_date})**")
        avail_delta = round(ov_new["Availability (%)"] - ov_old["Availability (%)"], 2)
        st.write(f"Total Inverters: **{ov_new['Total Inverters']:,}**")
        st.write(f"Total Active Strings: **{ov_new['Total Active Strings']:,}**")
        st.write(f"Working Strings: **{ov_new['Total Working Strings']:,}**")
        st.write(f"Failed Strings: **{ov_new['Total Failed Strings']:,}**")
        st.write(f"Availability: **{ov_new['Availability (%)']}%** ({'+' if avail_delta >= 0 else ''}{avail_delta}%)")

    fig = go.Figure()
    fig.add_trace(go.Bar(name="Working", x=["From Date", "To Date"],
                          y=[ov_old["Total Working Strings"], ov_new["Total Working Strings"]], marker_color="#10b981"))
    fig.add_trace(go.Bar(name="Failed", x=["From Date", "To Date"],
                          y=[ov_old["Total Failed Strings"], ov_new["Total Failed Strings"]], marker_color="#ef4444"))
    fig.update_layout(barmode="group", height=360, title="Working vs Failed Strings - From Date vs To Date",
                       plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
    st.plotly_chart(fig, use_container_width=True, key="working_hours_overview_chart")

    perf_df = pd.DataFrame({
        "Metric": ["Total Inverters", "Total Active Strings", "Working Strings", "Failed Strings", "Availability (%)"],
        f"From Date ({old_date})": [ov_old["Total Inverters"], ov_old["Total Active Strings"], 
                                     ov_old["Total Working Strings"], ov_old["Total Failed Strings"], 
                                     ov_old["Availability (%)"]],
        f"To Date ({new_date})": [ov_new["Total Inverters"], ov_new["Total Active Strings"],
                                   ov_new["Total Working Strings"], ov_new["Total Failed Strings"],
                                   ov_new["Availability (%)"]]
    })
    render_colored_excel_download(
        "📥 Download Performance Overview (Excel, color-coded)",
        perf_df,
        "performance_overview",
        key="wh_perf_download",
        value_cols=[f"From Date ({old_date})", f"To Date ({new_date})"]
    )


def display_calendar_comparison(sheet_name="Sheet1"):
    """Calendar-wise comparison with trend analysis and color-coded exports."""
    st.subheader("📅 Calendar-wise Comparison (From Date → To Date)")
    with st.expander("ℹ️ How to read this tab", expanded=False):
        st.markdown(
            "Pick any **From Date** and **To Date** from the uploaded snapshots.\n"
            "- **Restored Strings**: were Failed on the From Date, Working on the To Date.\n"
            "- **Newly Failed Strings**: were Working on the From Date, Failed on the To Date.\n"
            "- **Restoration TAT Summary**: assumes every string that recovered took the *entire* gap.\n"
            "- **Trend Across the Selected Range** charts every day in between."
        )

    available_dates = get_available_snapshot_dates()
    if len(available_dates) < 2:
        st.info("📌 At least 2 saved snapshot dates are required for comparison.")
        return

    min_date = datetime.strptime(available_dates[0], "%Y-%m-%d").date()
    max_date = datetime.strptime(available_dates[-1], "%Y-%m-%d").date()

    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date", value=min_date, min_value=min_date, max_value=max_date, key="restore_range_from")
    with col2:
        to_date = st.date_input("To Date", value=max_date, min_value=min_date, max_value=max_date, key="restore_range_to")

    if from_date > to_date:
        st.error("❌ From Date must be on or before To Date.")
        return

    with st.spinner("🔄 Loading comparison data..."):
        result, error = compare_two_snapshots_by_date(str(from_date), str(to_date), sheet_name)

    if error:
        st.error(error)
        return

    df_history = result["df_history"]
    inverter_col = result["inverter_col"]
    baseline_current = result["baseline_avg_working_current"]

    restored_total = int(df_history["Failed_to_Working"].sum())
    newly_failed_total = int(df_history["Working_to_Failed"].sum())
    failed_from = int(df_history["Failed String Count_old"].sum())
    failed_to = int(df_history["Failed String Count_new"].sum())

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("From → To", f"{from_date} → {to_date}")
    c2.metric("✅ Restored Strings", restored_total)
    c3.metric("❌ Newly Failed Strings", newly_failed_total)
    c4.metric("Failed: From-Date", failed_from)
    c5.metric("Failed: To-Date", failed_to, delta=int(failed_to - failed_from), delta_color="inverse")

    st.caption(f"Baseline average working current: {baseline_current:.2f} A · Time-based metrics assume {WORKING_HOURS_START}:00-{WORKING_HOURS_END}:00 working hours per day.")

    st.markdown("---")
    st.markdown("#### ⏱️ Restoration TAT Summary")
    total_tat_hours = float(df_history["Restoration_TAT_Hours"].sum())
    avg_tat_hours = float(df_history.loc[df_history["Restoration_TAT_Hours"] > 0, "Restoration_TAT_Hours"].mean()) if (df_history["Restoration_TAT_Hours"] > 0).any() else 0.0
    still_failing_hours = float(df_history["Current_Failure_Hours"].sum())

    t1, t2, t3 = st.columns(3)
    t1.metric("Total Restoration TAT (hrs)", f"{total_tat_hours:,.0f}")
    t2.metric("Avg TAT per Restored String (hrs)", f"{avg_tat_hours:,.1f}")
    t3.metric("Cumulative Ongoing-Failure Hours", f"{still_failing_hours:,.0f}")

    st.markdown("---")
    st.markdown("#### 📋 Full Comparison Detail")
    display_history = df_history.rename(columns={inverter_col: "Inverter"}).sort_values(
        by="Current_Failure_Hours", ascending=False
    )
    page_history = paginate_dataframe(display_history, page_size=25, key_prefix="calendar_compare")
    st.dataframe(page_history, use_container_width=True,
                 column_config=build_grid_column_config(page_history))

    render_colored_excel_download(
        "📥 Download Full Comparison (Excel, color-coded)",
        display_history,
        f"calendar_compare_{from_date}_{to_date}",
        key="calendar_compare_download",
        value_cols=["Failed_to_Working", "Working_to_Failed", "Failed String Count_old", "Failed String Count_new",
                    "Restoration_TAT_Hours", "Current_Failure_Hours"]
    )

    st.markdown("---")
    st.markdown("#### 📈 Trend Across the Selected Range")

    with st.spinner("📊 Loading trend data..."):
        df_trend = build_range_trend_data(from_date, to_date, sheet_name)

    if df_trend.empty:
        st.info("📌 No trend data available for this range.")
        return

    fig_avail = px.line(
        df_trend, x="Date", y="Availability (%)", color="Plot", line_group="Block",
        markers=True, title="Availability Trend by Plot / Block",
    )
    fig_avail.update_layout(height=450, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
    st.plotly_chart(fig_avail, use_container_width=True, key="cal_trend_avail")

    fig_wf = px.bar(
        df_trend, x="Date", y=["Working", "Failed"], facet_col="Plot",
        barmode="stack", title="Working vs Failed Strings Over Time (by Plot)",
        color_discrete_sequence=["#10b981", "#ef4444"],
    )
    fig_wf.update_layout(height=450, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
    st.plotly_chart(fig_wf, use_container_width=True, key="cal_trend_stacked")

    render_colored_excel_download(
        "📥 Download Trend Data (Excel, color-coded)",
        df_trend,
        f"trend_data_{from_date}_{to_date}",
        key="cal_trend_download",
        value_cols=["Working", "Failed", "Total", "Availability (%)"]
    )


# ==========================================
# MAIN ENTRY
# ==========================================
def display_tat_dashboard(processed_dataframes=None, current_df=None, sheet_name="Sheet1",
                           user_role="viewer", username="unknown", upload_handler=None,
                           snapshot_date=None):
    """Main entry point for the Restore & TAT dashboard."""
    st.title("🔄 Restore & TAT Analysis")
    st.caption("Day-wise SCADA snapshots power history, TAT, and calendar comparisons.")
    
    init_history()
    history = load_string_history()

    history_source_df = None
    if processed_dataframes and sheet_name in processed_dataframes:
        history_source_df = processed_dataframes[sheet_name]
    elif current_df is not None:
        history_source_df = current_df

    history_date = str(snapshot_date or datetime.now().strftime("%Y-%m-%d"))
    history_key = f"{history_date}:{sheet_name}:{len(history_source_df) if history_source_df is not None else 0}"
    if history_source_df is not None and not history_source_df.empty and st.session_state.get("restore_history_key") != history_key:
        update_string_history(history_source_df, history_date)
        st.session_state.restore_history_key = history_key
        history = load_string_history()

    tabs = st.tabs([
        "📁 Upload Registry",
        "📊 Summary Dashboard",
        "📈 String History Matrix",
        "📉 Inverter History Matrix",
        "⏰ Working Hours",
        "📅 Calendar Compare",
    ])

    with tabs[0]:
        display_upload_registry(user_role, upload_handler=upload_handler)

    with tabs[1]:
        display_summary_dashboard(sheet_name)

    with tabs[2]:
        display_string_history_matrix(history, current_df)

    with tabs[3]:
        display_inverter_history_matrix()

    with tabs[4]:
        display_working_hours_analysis()

    with tabs[5]:
        display_calendar_comparison(sheet_name)


# ==========================================
# BACKWARD COMPATIBILITY WRAPPER
# ==========================================
def get_restore_tab(processed_dataframes=None, filtered_df=None, sheet_name="Sheet1",
                     user_role="viewer", username="unknown", process_scada_excel=None,
                     upload_handler=None, snapshot_date=None):
    """Kept for backward compatibility with older app.py calls."""
    return display_tat_dashboard(
        processed_dataframes=processed_dataframes,
        current_df=filtered_df,
        sheet_name=sheet_name,
        user_role=user_role,
        username=username,
        upload_handler=upload_handler,
        snapshot_date=snapshot_date,
    )
